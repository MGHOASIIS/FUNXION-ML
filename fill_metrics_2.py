import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ─────────────────────────────────────────────────────────────────────────────
# 1) LOAD YOUR “LONG” RESULTS (from CSV or Python list of dicts)
# ─────────────────────────────────────────────────────────────────────────────

# TODO: change here
df_res = pd.read_csv("rcvni-righthandused.csv")

# ─────────────────────────────────────────────────────────────────────────────
# 2) MAP NUMERIC OR INTERNAL CODES TO EXACT EXCEL LABELS
# ─────────────────────────────────────────────────────────────────────────────

# paradigm_map = {
#     2: "1. Rotatary Cuff vs No Injury",
#     1: "2. Injury vs No Injury (# 37 vs 17)",
#     3: "4. Other Injury vs No Injury",
#     4: "3. Rotatary vs Other Injury"
# }

hand_map = {
    0: "Left hand",
    1: "Right hand"
}

dominance_map = {
    0: "Left hand",
    1: "Right hand"
}

# attempt_map = {
#     2: "One attempts in LL",
#     1: "One attempts in LU",
#     3: "One attempts in LL+LU",
#     4: "Multiple attempts in LL+LU"
# }

feature_filter_map = {
    1: "Time series (with padding)",
    2: "Time series (with truncating)",
    3: "Time series (with DTW)",
    4: "Time series (variable)"
}

model_map = {
    1: "LR (LOO)",
    2: "RF (LOO)",
    3: "HMM (LOO)",
    4: "CNN (LOO)",
    5: "RNN (LOO)"
}

movement_map = {
    1: "Pronation", # Lock unlocked *****
    2: "Supination", # Lock locked **** this is wrong but to make it easy !
}

# df_res["ParadigmLabel"] = df_res["Paradigm"].map(paradigm_map)
df_res["HandUsedLabel"] = df_res["HandUsed"].map(hand_map)
df_res["DominanceLabel"] = df_res["DominantHand"].map(dominance_map)
# df_res["AttemptLabel"] = df_res["OneAttempt"].map(attempt_map)
df_res["FeatureFilterLabel"] = df_res["FeatureFilter"].map(feature_filter_map)
df_res["ModelLabel"] = df_res["Model"].map(model_map)
df_res["KeyLabel"] = df_res["Keys"].map(movement_map)
print(df_res)

# ─────────────────────────────────────────────────────────────────────────────
# 3) READ YOUR EXISTING EXCEL “TEMPLATE”
# ─────────────────────────────────────────────────────────────────────────────

# Adjust the sheet_name= if your sheet is named differently.
output_filename = "results-table.xlsx"
sheet_name = "Task2-rcvni"
df_tpl = pd.read_excel(output_filename, sheet_name=sheet_name)

# For clarity, let’s display the first few columns so you can see column names:
print("=== TEMPLATED SHEET COLUMNS ===")
print(df_tpl.columns.tolist())
print("===============================")


# ─────────────────────────────────────────────────────────────────────────────
# 4) LOOP OVER EACH ROW OF THE TEMPLATE AND “FILL IN” THE AUC/BA/RECALL CELLS
# ─────────────────────────────────────────────────────────────────────────────

print("\n=== RESULT KEYS ===")
print(df_res.columns.tolist())
print("===================\n")

filled_cells = []
print("<<<<<", df_res["HandUsedLabel"], df_res["KeyLabel"], df_res["DominanceLabel"], df_res["FeatureFilterLabel"], df_res["ModelLabel"])

# Now iterate row by row through the template:
for idx, tpl_row in df_tpl.iterrows():
    # 1) Read “Filter – Number of attempts” from the template
    handused_label = tpl_row["Paradigm"]
    key_label = tpl_row["Movement"]
    dom_label = tpl_row["Dominance"]
    feature_filter_label = tpl_row["Signal length"]
    model_label = tpl_row["Model"]

    # print(handused_label == df_res["HandUsedLabel"])
    # print((df_res["KeyLabel"] == key_label.strip().split()[0]))
    # print((df_res["DominanceLabel"] == dom_label))
    # print(feature_filter_label.split('[', 1)[0].rstrip() == df_res["FeatureFilterLabel"])
    # print((df_res["ModelLabel"] == model_label))

    match = df_res[
        (df_res["HandUsedLabel"] == handused_label) & 
        (df_res["KeyLabel"] == key_label.strip().split()[0]) &
        (df_res["DominanceLabel"] == dom_label) &
        (df_res["FeatureFilterLabel"] == feature_filter_label.split('[', 1)[0].rstrip()) &
        (df_res["ModelLabel"] == model_label)
    ]
    print(">>>>>", handused_label, key_label, dom_label, feature_filter_label, model_label)
    print("OLAYYYYYYYYYY", match)
    if match.shape[0] == 0:
        # No matching result for this (paradigm,attempt). Skip.
        continue

    
    row = match.iloc[-1]
    print("@@@@@@@@@@@", row)

    auc_val    = row["auc"]
    ba_val     = row["ba"]
    recall_val = row["recall"]


    print("#######", df_tpl.at[idx+1, 'AUC'])
    print("#######", df_tpl.at[idx+1, 'BA'])
    print("#######", df_tpl.at[idx+1, 'Recall'])

    number_strings = re.findall(r"\d+\.\d+", auc_val)
    numbers = [float(x) for x in number_strings]
    df_tpl.at[idx+1, 'AUC'] = str(round(numbers[0], 3))+"["+str(round(numbers[1], 3))+"-"+str(round(numbers[2], 3))+"]"
    df_tpl.at[idx+1, 'BA'] = ba_val
    df_tpl.at[idx+1, 'Recall'] = recall_val

    print("#######", df_tpl.at[idx+1, 'AUC'])
    print("#######", df_tpl.at[idx+1, 'BA'])
    print("#######", df_tpl.at[idx+1, 'Recall'])


    filled_cells.append((idx+1, 'AUC'))
    filled_cells.append((idx+1, 'BA'))
    filled_cells.append((idx+1, 'Recall'))

# ─────────────────────────────────────────────────────────────────────────────
# 5) SAVE THE UPDATED EXCEL
# ─────────────────────────────────────────────────────────────────────────────


df_tpl.to_excel(output_filename, sheet_name=sheet_name, index=False)
print(f"\n✅ Wrote updated sheet to '{output_filename}'")

# ─────────────────────────────────────────────────────────────────────────────
# 5) REOPEN WITH openpyxl AND APPLY FILL TO RECORDED CELLS
# ─────────────────────────────────────────────────────────────────────────────

wb = load_workbook(output_filename)
ws = wb.active

# Create a yellow fill style
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Map column names to Excel column letters
# (e.g. "1. Rotator Cuff vs No Injury" → "D" if that is column D in the sheet)
col_letter_map = {}
for col_idx, col_name in enumerate(df_tpl.columns, start=1):
    col_letter_map[col_name] = ws.cell(row=1, column=col_idx).column_letter

# Now iterate over filled_cells to color each one
for row_idx, col_name in filled_cells:
    # Excel row = DataFrame index + 2 (because header is row 1, and index=0→Excel row 2)
    excel_row    = row_idx + 2
    excel_column = col_letter_map[col_name]
    ws[f"{excel_column}{excel_row}"].fill = yellow_fill

# Save the workbook
wb.save(output_filename)
print(f"✅ Filled cells highlighted in yellow and saved to '{output_filename}'")