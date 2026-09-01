"""
extract_subject_features.py
============================
Extracts per-subject features from raw pickled datasets for
Paradigm 2 (RCT vs Controls) across all 6 tasks.

Features
--------
Extended features from the features/ folder (~560 features).
All keys are prefixed with 'ext_' in the output.
Modules called directly (same logic as MasterFeatureExtractor):

  biomechanical.py
    BiomechanicalFeatures(sensor_names)
      .extract_biomechanical_features(signal, time_elapsed)

  joint_kinematics.py  (inline)
    ROM / mean / max / min on rotation channels per sensor
    Keys: {sensor}_{plane}_rom/mean/max/min
          planes: roll, pitch, yaw   sensors: Head, LeftHand, RightHand

  asymmetry_compensation_metrics.py
    calculate_asymmetry(left_data, right_data)   keys prefixed 'hand_'

  movement_quality.py
    calculate_smoothness_jerk(velocity, dt) / calculate_sparc(velocity, sr)
    keys prefixed '{sensor}_'

  temporal.py
    extract_temporal_features(head_diff_magnitude, sr)

  time_domain_stats.py
    extract_time_domain_features(signal_1d)  per channel (all 18)
    keys prefixed '{channel_display_name}_'

  frequency_domain.py
    extract_frequency_features(pos_mag, sr)  per sensor
    keys prefixed '{sensor}_'

  wavelet_features.py
    extract_wavelet_features(head_pos_magnitude, wavelet='db4', level=5)
    keys prefixed 'head_'

  complexity_entropy.py
    extract_complexity_features(pos_mag)  per sensor
    keys prefixed '{sensor}_'

Clinical data (DASH scores, Q1-Q11, diagnosis) are joined from
px_details and kept as identity columns — not used as features.

Usage
-----
  python extract_subject_features.py \\
      --data-dir storage/pickled/xdash \\
      --px-details storage/raw/xdash/xdash_px_details.xlsx \\
      --features-dir features/ \\
      --out-dir storage/results/xdash/hmm/subject_features/
"""

import argparse
import pickle
import sys
import warnings
from pathlib import Path

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore", category=RuntimeWarning)

# =============================================================================
# Constants
# =============================================================================

SAMPLING_RATE = 50  # Hz

SENSOR_NAMES   = ["Head", "LeftHand", "RightHand"]
SENSOR_OFFSETS = {"Head": 0, "LeftHand": 6, "RightHand": 12}

CHAN_DISPLAY = [
    "H_PosX", "H_PosY", "H_PosZ", "H_RotX", "H_RotY", "H_RotZ",
    "L_PosX", "L_PosY", "L_PosZ", "L_RotX", "L_RotY", "L_RotZ",
    "R_PosX", "R_PosY", "R_PosZ", "R_RotX", "R_RotY", "R_RotZ",
]

TASK_LABELS = {
    1: "Jar Opening", 
    # 2: "Key Turning",  3: "Cleaning",
    # 4: "Back Washing", 5: "Cutting",     6: "Hammering",
}

TASK_Q_COL = {
    1: ("q1_jar",          "Q1: Open jar"),
    2: ("q2_key",          "Q2: Turn key"),
    3: ("q3_household",    "Q3: Household"),
    4: ("q4_back_wash",    "Q4: Wash back"),
    5: ("q5_knife",        "Q5: Cut food"),
    6: ("q6_recreational", "Q6: Recreational"),
}

Q_COLS = [
    "q1_jar", "q2_key", "q3_household", "q4_back_wash", "q5_knife",
    "q6_recreational", "q7_social", "q8_work",
    "q9_pain", "q10_tingling", "q11_sleep",
]

DIA_NAMES    = {0: "Healthy/Control", 1: "RCT", 2: "Arthritis",
                3: "Bursitis", 4: "Tendonitis"}
RCT_DIA_CODE = 1


# =============================================================================
# Extended feature modules — lazy import
# =============================================================================

_EXT_AVAILABLE       = False
_BiomechanicalFeatures   = None
_calculate_asymmetry     = None
_calculate_smoothness    = None
_calculate_sparc         = None
_extract_temporal        = None
_extract_time_domain     = None
_extract_frequency       = None
_extract_wavelet         = None
_extract_complexity      = None


def _init_extended_modules(features_dir: Path | None) -> bool:
    """
    Import all functions from the features/ folder.
    Sets module-level globals so extract_extended_features() can use them.
    Returns True if at least one module loaded successfully.
    """
    global _EXT_AVAILABLE
    global _BiomechanicalFeatures, _calculate_asymmetry
    global _calculate_smoothness, _calculate_sparc
    global _extract_temporal, _extract_time_domain
    global _extract_frequency, _extract_wavelet, _extract_complexity

    if features_dir is not None and features_dir.exists():
        sys.path.insert(0, str(features_dir.resolve()))

    errors = []

    try:
        from biomechanical import BiomechanicalFeatures
        _BiomechanicalFeatures = BiomechanicalFeatures
    except Exception as e:
        errors.append(f"biomechanical: {e}")

    try:
        from asymmetry_compensation_metrics import calculate_asymmetry
        _calculate_asymmetry = calculate_asymmetry
    except Exception as e:
        errors.append(f"asymmetry_compensation_metrics: {e}")

    try:
        from movement_quality import calculate_smoothness_jerk, calculate_sparc
        _calculate_smoothness = calculate_smoothness_jerk
        _calculate_sparc      = calculate_sparc
    except Exception as e:
        errors.append(f"movement_quality: {e}")

    try:
        from temporal import extract_temporal_features
        _extract_temporal = extract_temporal_features
    except Exception as e:
        errors.append(f"temporal: {e}")

    try:
        from time_domain_stats import extract_time_domain_features
        _extract_time_domain = extract_time_domain_features
    except Exception as e:
        errors.append(f"time_domain_stats: {e}")

    try:
        from frequency_domain import extract_frequency_features
        _extract_frequency = extract_frequency_features
    except Exception as e:
        errors.append(f"frequency_domain: {e}")

    try:
        from wavelet_features import extract_wavelet_features
        _extract_wavelet = extract_wavelet_features
    except Exception as e:
        errors.append(f"wavelet_features: {e}")

    try:
        from complexity_entropy import extract_complexity_features
        _extract_complexity = extract_complexity_features
    except Exception as e:
        errors.append(f"complexity_entropy: {e}")

    # joint_kinematics is computed inline — no import needed

    if errors:
        print(f"  [Extended] Partial load — {len(errors)} module(s) missing:")
        for e in errors:
            print(f"    x {e}")
        n_ok = 9 - len(errors)
        print(f"  [Extended] {n_ok}/9 modules active")
        _EXT_AVAILABLE = n_ok > 0
    else:
        print("  [Extended] All 9 feature modules loaded — ~560 features per subject")
        _EXT_AVAILABLE = True

    return _EXT_AVAILABLE


# =============================================================================
# Extended feature extraction
# =============================================================================

def extract_extended_features(signal: np.ndarray,
                               time_elapsed: np.ndarray) -> dict:
    """
    Call every function from the features/ folder and return a flat dict.
    All keys are prefixed with 'ext_'.
    Any individual module failure is caught silently — the subject row
    is never dropped because one module errors.
    """
    if not _EXT_AVAILABLE:
        return {}

    feats = {}
    sr = SAMPLING_RATE

    # ------------------------------------------------------------------
    # 1. BIOMECHANICAL
    # ------------------------------------------------------------------
    if _BiomechanicalFeatures is not None:
        try:
            extractor = _BiomechanicalFeatures(SENSOR_NAMES)
            bio = extractor.extract_biomechanical_features(signal, time_elapsed)
            feats.update(bio)
        except Exception:
            pass

    # ------------------------------------------------------------------
    # 2. JOINT KINEMATICS  (inline)
    #    Rotation channels: Head=3-5, LeftHand=9-11, RightHand=15-17
    #    Planes: roll(0), pitch(1), yaw(2)
    # ------------------------------------------------------------------
    try:
        planes = ["roll", "pitch", "yaw"]
        for sensor_name, offset in SENSOR_OFFSETS.items():
            rot_start = offset + 3
            rotations = signal[:, rot_start:rot_start + 3]
            for j, plane in enumerate(planes):
                rot_col = rotations[:, j]
                feats[f"{sensor_name}_{plane}_rom"]  = float(np.ptp(rot_col))
                feats[f"{sensor_name}_{plane}_mean"] = float(np.mean(rot_col))
                feats[f"{sensor_name}_{plane}_max"]  = float(np.max(rot_col))
                feats[f"{sensor_name}_{plane}_min"]  = float(np.min(rot_col))
    except Exception:
        pass

    # ------------------------------------------------------------------
    # 3. ASYMMETRY
    #    left_data  = signal[:, 6:12]   LeftHand all 6 channels
    #    right_data = signal[:, 12:18]  RightHand all 6 channels
    # ------------------------------------------------------------------
    if _calculate_asymmetry is not None:
        try:
            asym = _calculate_asymmetry(signal[:, 6:12], signal[:, 12:18])
            for k, v in asym.items():
                feats[f"hand_{k}"] = v
        except Exception:
            pass

    # ------------------------------------------------------------------
    # 4. MOVEMENT QUALITY
    #    velocity = |diff(positions)| * sr  (scalar magnitude per sensor)
    # ------------------------------------------------------------------
    for sensor_name, offset in SENSOR_OFFSETS.items():
        positions = signal[:, offset:offset + 3]
        velocity  = np.linalg.norm(np.diff(positions, axis=0), axis=1) * sr
        velocity  = np.concatenate([[velocity[0]], velocity])  # restore length T

        if _calculate_smoothness is not None:
            try:
                jerk_feats = _calculate_smoothness(velocity, 1.0 / sr)
                for k, v in jerk_feats.items():
                    feats[f"{sensor_name}_{k}"] = v
            except Exception:
                pass

        if _calculate_sparc is not None:
            try:
                feats[f"{sensor_name}_sparc"] = _calculate_sparc(velocity, sr)
            except Exception:
                pass

    # ------------------------------------------------------------------
    # 5. TEMPORAL
    #    signal_1d = head position magnitude of diff
    # ------------------------------------------------------------------
    if _extract_temporal is not None:
        try:
            head_pos    = signal[:, 0:3]
            head_diff   = np.linalg.norm(np.diff(head_pos, axis=0), axis=1)
            head_signal = np.concatenate([[head_diff[0]], head_diff])
            feats.update(_extract_temporal(head_signal, sr))
        except Exception:
            pass

    # ------------------------------------------------------------------
    # 6. TIME DOMAIN STATS  — all 18 channels
    # ------------------------------------------------------------------
    if _extract_time_domain is not None:
        try:
            for ch_idx, ch_name in enumerate(CHAN_DISPLAY):
                td_feats = _extract_time_domain(signal[:, ch_idx])
                for k, v in td_feats.items():
                    feats[f"{ch_name}_{k}"] = v
        except Exception:
            pass

    # ------------------------------------------------------------------
    # 7. FREQUENCY DOMAIN  — per sensor
    # ------------------------------------------------------------------
    if _extract_frequency is not None:
        for sensor_name, offset in SENSOR_OFFSETS.items():
            try:
                pos_mag    = np.linalg.norm(signal[:, offset:offset + 3], axis=1)
                freq_feats = _extract_frequency(pos_mag, sr)
                for k, v in freq_feats.items():
                    feats[f"{sensor_name}_{k}"] = v
            except Exception:
                pass

    # ------------------------------------------------------------------
    # 8. WAVELET  — head position magnitude
    # ------------------------------------------------------------------
    if _extract_wavelet is not None:
        try:
            head_mag  = np.linalg.norm(signal[:, 0:3], axis=1)
            wav_feats = _extract_wavelet(head_mag, wavelet="db4", level=5)
            for k, v in wav_feats.items():
                feats[f"head_{k}"] = v
        except Exception:
            pass

    # ------------------------------------------------------------------
    # 9. COMPLEXITY / ENTROPY  — per sensor
    # ------------------------------------------------------------------
    if _extract_complexity is not None:
        for sensor_name, offset in SENSOR_OFFSETS.items():
            try:
                pos_mag    = np.linalg.norm(signal[:, offset:offset + 3], axis=1)
                cplx_feats = _extract_complexity(pos_mag)
                for k, v in cplx_feats.items():
                    feats[f"{sensor_name}_{k}"] = v
            except Exception:
                pass

    return {f"ext_{k}": v for k, v in feats.items()}


# =============================================================================
# Data loading
# =============================================================================

def load_pickled(data_dir: Path, task: int):
    patient_pkl = data_dir / f"patient_data_task{task}.pkl"
    control_pkl = data_dir / f"control_data_task{task}.pkl"
    if not patient_pkl.exists():
        raise FileNotFoundError(f"Missing: {patient_pkl}")
    if not control_pkl.exists():
        raise FileNotFoundError(f"Missing: {control_pkl}")
    with open(patient_pkl, "rb") as f:
        patients = pickle.load(f)
    with open(control_pkl, "rb") as f:
        controls = pickle.load(f)
    return patients, controls


def load_subject_metadata(px_details_path: Path) -> dict:
    if not px_details_path.exists():
        print("  [WARN] px_details not found — no clinical data joined")
        return {}

    df = None
    for header_row in range(6):
        candidate = pd.read_excel(px_details_path, sheet_name="Sheet1",
                                  header=header_row)
        if "dia_code" in candidate.columns and "id" in candidate.columns:
            df = candidate
            break

    if df is None:
        print("  [WARN] Could not find 'dia_code' column")
        return {}

    df.columns = [str(c).strip() for c in df.columns]
    meta = {}

    for _, row in df.iterrows():
        sid = str(row.get("id", "")).strip().upper()
        if not sid or sid in ("NAN", "NONE", ""):
            continue

        dia_code = row.get("dia_code", None)
        try:
            dia_code = int(dia_code)
        except (ValueError, TypeError):
            dia_code = None

        diag_raw = row.get("diagnosis", None)
        if pd.isna(diag_raw) or str(diag_raw).strip() in ("", "nan", "NA", "NaN"):
            diagnosis = DIA_NAMES.get(dia_code, "Unknown") \
                        if dia_code is not None else "Healthy/Control"
        else:
            diagnosis = str(diag_raw).strip()

        def _get(col_variants):
            for col in col_variants:
                for c in df.columns:
                    if c.strip().lower() == col.lower():
                        val = row.get(c)
                        try:
                            return float(val) if not pd.isna(val) else None
                        except (TypeError, ValueError):
                            return None
            return None

        c_dash   = _get(["C_DASH", "c_dash"])
        x_dash   = _get(["X_DASH", "x_dash"])
        avg_dash = _get(["c and x", "avg_dash"])
        q_vals   = {q: _get([q]) for q in Q_COLS}

        dash_stored = _get(["DASH_survey", "dash_survey"])
        if dash_stored is not None:
            dash_survey = dash_stored
        else:
            q_num = [v for v in q_vals.values() if v is not None]
            dash_survey = ((sum(q_num) / len(q_num)) - 1) * 25 \
                          if len(q_num) == 11 else None

        meta[sid] = {
            "dia_code": dia_code, "diagnosis": diagnosis,
            "C_DASH": c_dash, "X_DASH": x_dash,
            "avg_dash": avg_dash, "DASH_survey": dash_survey,
            **q_vals,
        }

    rct_n  = sum(1 for v in meta.values() if v.get("dia_code") == RCT_DIA_CODE)
    dash_n = sum(1 for v in meta.values() if v.get("C_DASH") is not None)
    print(f"  [px_details] {len(meta)} subjects | RCT={rct_n} | DASH={dash_n}")
    return meta


def to_numpy(tensor) -> np.ndarray:
    try:
        import torch
        arr = tensor.detach().cpu().numpy() if isinstance(tensor, torch.Tensor) \
              else np.asarray(tensor, dtype=float)
    except ImportError:
        arr = np.asarray(tensor, dtype=float)
    if arr.ndim == 2 and arr.shape[1] == 19:
        arr = arr[:, 1:]   # drop TimeElapsed column if present
    return arr


# =============================================================================
# Main extraction loop
# =============================================================================

def run_extraction(
    data_dir:        Path,
    px_details_path: Path,
    out_dir:         Path,
    tasks:           list[int],
    features_dir:    Path | None = None,
):
    print("\n" + "=" * 70)
    print("XDash Subject Feature Extraction — Paradigm 2 (RCT vs Controls)")
    print("=" * 70)

    subject_meta = load_subject_metadata(px_details_path)
    rct_ids = {k for k, v in subject_meta.items()
               if v.get("dia_code") == RCT_DIA_CODE} if subject_meta else None

    _init_extended_modules(features_dir)

    all_rows = []

    for task in tasks:
        print(f"\n[Task {task}] {TASK_LABELS[task]}")
        print("-" * 45)

        try:
            patients, controls = load_pickled(data_dir, task)
        except FileNotFoundError as e:
            print(f"  [SKIP] {e}")
            continue

        rct_patients = {k: v for k, v in patients.items()
                        if str(k).upper() in rct_ids} if rct_ids else patients
        print(f"  Patients: {len(patients)} total -> {len(rct_patients)} RCT")
        print(f"  Controls: {len(controls)}")

        q_col, _ = TASK_Q_COL.get(task, ("q1_jar", "Q1"))

        for group_label, group_dict, y_val in [
            ("RCT",     rct_patients, 1),
            ("Control", controls,     0),
        ]:
            for sid, tensor in group_dict.items():
                try:
                    signal = to_numpy(tensor)
                except Exception as e:
                    print(f"    [{sid}] Conversion error: {e}")
                    continue

                if signal.ndim != 2 or signal.shape[1] != 18:
                    print(f"    [{sid}] Bad shape {signal.shape} — skip")
                    continue

                T            = signal.shape[0]
                time_elapsed = np.arange(T) / SAMPLING_RATE

                ext = extract_extended_features(signal, time_elapsed)

                meta    = subject_meta.get(str(sid).upper(), {})
                dash_ft = {
                    "dia_code":     meta.get("dia_code"),
                    "diagnosis":    meta.get("diagnosis"),
                    "C_DASH":       meta.get("C_DASH"),
                    "X_DASH":       meta.get("X_DASH"),
                    "avg_dash":     meta.get("avg_dash"),
                    "DASH_survey":  meta.get("DASH_survey"),
                    "task_q_score": meta.get(q_col),
                    **{q: meta.get(q) for q in Q_COLS},
                }

                row = {
                    "subject_id": str(sid),
                    "group":      group_label,
                    "y":          y_val,
                    "task":       task,
                    "task_name":  TASK_LABELS[task],
                    **dash_ft,
                    **ext,
                }
                all_rows.append(row)

                dash_str = f"C_DASH={dash_ft['C_DASH']}" \
                           if dash_ft["C_DASH"] is not None else "no DASH"
                print(f"    [{group_label:7s}] {str(sid):8s}  "
                      f"T={T:5d}  +{len(ext)} ext  {dash_str}")

    if not all_rows:
        print("\n[ERROR] No data extracted. Check --data-dir.")
        sys.exit(1)

    # ==========================================================================
    # Assemble DataFrame
    # ==========================================================================
    df = pd.DataFrame(all_rows)

    id_cols   = ["subject_id", "group", "y", "task", "task_name"]
    dash_cols = ["dia_code", "diagnosis", "C_DASH", "X_DASH", "avg_dash",
                 "DASH_survey", "task_q_score"] + Q_COLS
    ext_cols  = sorted(c for c in df.columns if c.startswith("ext_"))
    fixed     = id_cols + dash_cols
    remaining = [c for c in df.columns
                 if c not in fixed and not c.startswith("ext_")]
    df = df[[c for c in fixed + remaining + ext_cols if c in df.columns]]

    # ==========================================================================
    # Summary
    # ==========================================================================
    print("\n" + "=" * 70)
    print(f"SUMMARY  rows={len(df)}  subjects={df['subject_id'].nunique()}"
          f"  tasks={df['task'].nunique()}"
          f"  total_cols={len(df.columns)}  ext_cols={len(ext_cols)}")
    print("=" * 70)
    summary_feats = ["C_DASH", "X_DASH", "DASH_survey"]
    print(f"\n{'Feature':<30}  {'RCT':>22}  {'Control':>22}")
    print("-" * 78)
    for feat in summary_feats:
        if feat not in df.columns:
            continue
        rv = df[df["group"] == "RCT"][feat].dropna()
        cv = df[df["group"] == "Control"][feat].dropna()
        rs = f"{rv.mean():.3f}+/-{rv.std():.3f}" if not rv.empty else "N/A"
        cs = f"{cv.mean():.3f}+/-{cv.std():.3f}" if not cv.empty else "N/A"
        print(f"  {feat:<28}  {rs:>22}  {cs:>22}")

    # ==========================================================================
    # Z-score normalisation (within-group, per task)
    # ==========================================================================
    ID_COLS_ZSCORE = ["subject_id", "group", "y", "task", "task_name",
                      "dia_code", "diagnosis"]

    numeric_cols = [c for c in df.columns
                    if c not in ID_COLS_ZSCORE
                    and pd.api.types.is_numeric_dtype(df[c])]

    def compute_group_zscores(df_grp: pd.DataFrame) -> pd.DataFrame:
        """
        Z-score each numeric feature per task within the group.
        Subjects where within-task std == 0 receive z = 0.
        """
        out = df_grp[ID_COLS_ZSCORE].copy()
        for col in numeric_cols:
            z_col = []
            for _task_id, sub in df_grp.groupby("task"):
                vals = sub[col].values.astype(float)
                mu   = np.nanmean(vals)
                sd   = np.nanstd(vals, ddof=0)
                z    = (vals - mu) / sd if sd > 1e-12 else np.zeros_like(vals)
                z_col.extend(zip(sub.index, z))
            out[col] = pd.Series(dict(z_col), name=col)
        return out.reset_index(drop=True)

    df_rct = df[df["group"] == "RCT"].reset_index(drop=True)
    df_ctl = df[df["group"] == "Control"].reset_index(drop=True)

    df_z_rct = compute_group_zscores(df_rct)
    df_z_ctl = compute_group_zscores(df_ctl)

    print(f"\n[Z-scores] RCT rows={len(df_z_rct)}  Control rows={len(df_z_ctl)}"
          f"  features z-scored={len(numeric_cols)}")

    # ==========================================================================
    # Save
    # ==========================================================================
    out_dir.mkdir(parents=True, exist_ok=True)
    csv_path  = out_dir / f"subject_features_P2_T{tasks}.csv"
    xlsx_path = out_dir / f"subject_features_P2_T{tasks}.xlsx"

    df.to_csv(csv_path, index=False)
    print(f"\n[Saved] CSV  -> {csv_path}")

    z_rct_csv = out_dir / "subject_features_P2_zscores_RCT.csv"
    z_ctl_csv = out_dir / "subject_features_P2_zscores_Control.csv"
    df_z_rct.to_csv(z_rct_csv, index=False)
    df_z_ctl.to_csv(z_ctl_csv, index=False)
    print(f"[Saved] Z-score CSV (RCT)     -> {z_rct_csv}")
    print(f"[Saved] Z-score CSV (Control) -> {z_ctl_csv}")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.utils import get_column_letter

        wb       = Workbook()
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        rct_fill = PatternFill("solid", fgColor="FFE0CC")
        ctl_fill = PatternFill("solid", fgColor="CCE5FF")
        ext_fill = PatternFill("solid", fgColor="E8F4E8")

        # ------------------------------------------------------------------
        # Sheet 1: full feature table
        # ------------------------------------------------------------------
        ws = wb.active
        ws.title = "Features"
        for r_idx, row_data in enumerate(
                dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row_data)
            if r_idx == 1:
                for ci, cell in enumerate(ws[r_idx], 1):
                    col_name       = df.columns[ci - 1]
                    cell.font      = Font(bold=True, color="FFFFFF")
                    cell.fill      = ext_fill if col_name.startswith("ext_") \
                                     else hdr_fill
                    cell.alignment = Alignment(horizontal="center")
            else:
                fill = rct_fill if ws.cell(r_idx, 2).value == "RCT" else ctl_fill
                for cell in ws[r_idx]:
                    cell.fill = fill
        for ci, col in enumerate(df.columns, 1):
            max_len = max(len(str(col)),
                          df[col].astype(str).str.len().max() if len(df) else 0)
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 30)
        ws.freeze_panes = "F2"

        # ------------------------------------------------------------------
        # Sheet 2: summary stats
        # ------------------------------------------------------------------
        ws2 = wb.create_sheet("Summary Stats")
        ws2.append(["Feature", "Group", "Mean", "Std", "Min", "Median", "Max", "N"])
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = hdr_fill
        for feat in summary_feats + ext_cols[:20]:
            if feat not in df.columns:
                continue
            for grp in ["RCT", "Control"]:
                v = df[df["group"] == grp][feat].dropna()
                if v.empty:
                    continue
                ws2.append([feat, grp,
                            round(v.mean(), 4), round(v.std(), 4),
                            round(v.min(), 4), round(v.median(), 4),
                            round(v.max(), 4), len(v)])

        # ------------------------------------------------------------------
        # Sheet 3: DASH scores (one row per subject)
        # ------------------------------------------------------------------
        ws3 = wb.create_sheet("DASH Scores")
        df_u = df.drop_duplicates("subject_id")[
            ["subject_id", "group", "diagnosis",
             "C_DASH", "X_DASH", "avg_dash", "DASH_survey"] + Q_COLS
        ].copy().reset_index(drop=True)
        for t in tasks:
            qc, _ = TASK_Q_COL.get(t, ("q1_jar", ""))
            tdf = df[df["task"] == t][["subject_id", "task_q_score"]].rename(
                columns={"task_q_score": f"T{t}_{qc}"})
            df_u = df_u.merge(tdf, on="subject_id", how="left")
        ws3.append(list(df_u.columns))
        for cell in ws3[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center")
        for r_idx, row_data in enumerate(
                dataframe_to_rows(df_u, index=False, header=False), 2):
            ws3.append(row_data)
            fill = rct_fill if ws3.cell(r_idx, 2).value == "RCT" else ctl_fill
            for cell in ws3[r_idx]:
                cell.fill = fill
        for ci, col in enumerate(df_u.columns, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = \
                max(len(str(col)) + 2, 12)
        ws3.freeze_panes = "D2"

        # ------------------------------------------------------------------
        # Sheet 4: extended feature group comparison (Cohen's d)
        # ------------------------------------------------------------------
        if ext_cols:
            ws4 = wb.create_sheet("Extended Features")
            ws4.append(["Feature", "Module", "RCT mean", "RCT std",
                        "Control mean", "Control std", "Diff", "Cohen d"])
            for cell in ws4[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2D5016")

            def _module_tag(col_name):
                n = col_name.replace("ext_", "")
                if n.startswith("hand_"):
                    return "asymmetry"
                if n.startswith("head_"):
                    return "wavelet/joint"
                for s in SENSOR_NAMES:
                    if n.startswith(f"{s}_"):
                        if "entropy" in n or "sample" in n:
                            return "complexity"
                        if "freq" in n or "dominant" in n:
                            return "frequency"
                        if "sparc" in n or "jerk" in n:
                            return "movement_quality"
                        return "biomechanical/joint"
                for cd in CHAN_DISPLAY:
                    if n.startswith(f"{cd}_"):
                        return "time_domain"
                return "temporal/other"

            for feat in ext_cols:
                rv = df[df["group"] == "RCT"][feat].dropna()
                cv = df[df["group"] == "Control"][feat].dropna()
                if rv.empty or cv.empty:
                    continue
                pooled = np.sqrt((rv.std()**2 + cv.std()**2) / 2 + 1e-10)
                d = (rv.mean() - cv.mean()) / pooled
                ws4.append([feat, _module_tag(feat),
                            round(rv.mean(), 4), round(rv.std(), 4),
                            round(cv.mean(), 4), round(cv.std(), 4),
                            round(rv.mean() - cv.mean(), 4), round(d, 3)])
            for ci in range(1, 9):
                ws4.column_dimensions[get_column_letter(ci)].width = 30
            ws4.freeze_panes = "B2"

        # ------------------------------------------------------------------
        # Sheets 5 & 6: Z-scores (RCT and Control)
        # ------------------------------------------------------------------
        def _write_zscore_sheet(wb, df_z, sheet_title, row_fill, threshold=1.5):
            from openpyxl.styles import PatternFill as PF

            high_fill = PF("solid", fgColor="FF8C00")   # orange — high outlier
            low_fill  = PF("solid", fgColor="4472C4")   # blue   — low outlier
            zero_fill = PF("solid", fgColor="F2F2F2")   # grey   — near mean

            ws_z = wb.create_sheet(sheet_title)
            ws_z.sheet_view.showGridLines = False

            all_cols = list(df_z.columns)
            id_end   = len(ID_COLS_ZSCORE)

            ws_z.append(all_cols)
            for ci, cell in enumerate(ws_z[1], 1):
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = hdr_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                ws_z.column_dimensions[get_column_letter(ci)].width = \
                    max(len(str(all_cols[ci - 1])) + 1, 10)
            ws_z.row_dimensions[1].height = 50

            for r_idx, row_data in enumerate(
                    dataframe_to_rows(df_z, index=False, header=False), 2):
                ws_z.append(row_data)
                for ci, cell in enumerate(ws_z[r_idx], 1):
                    if ci <= id_end:
                        cell.fill = row_fill
                        continue
                    try:
                        z_val = float(cell.value) if cell.value is not None else None
                    except (ValueError, TypeError):
                        z_val = None
                    if z_val is None or np.isnan(z_val):
                        cell.value = ""
                        continue
                    cell.number_format = "+0.00;-0.00;0.00"
                    if z_val > threshold:
                        cell.fill = high_fill
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif z_val < -threshold:
                        cell.fill = low_fill
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif abs(z_val) < 0.25:
                        cell.fill = zero_fill
                    else:
                        cell.fill = row_fill

            for ci in range(1, id_end + 1):
                ws_z.column_dimensions[get_column_letter(ci)].width = 14

            # Legend banner
            ws_z.insert_rows(1)
            ws_z.merge_cells(f"A1:{get_column_letter(len(all_cols))}1")
            ws_z["A1"].value = (
                f"Z-scores (within-group, per task) -- "
                f"Orange = z > +{threshold}s (high outlier)   "
                f"Blue = z < -{threshold}s (low outlier)   "
                f"Grey = |z| < 0.25 (near mean)"
            )
            ws_z["A1"].font      = Font(bold=True, color="1F3864", size=10)
            ws_z["A1"].fill      = PF("solid", fgColor="D9E1F2")
            ws_z["A1"].alignment = Alignment(horizontal="left", vertical="center")
            ws_z.row_dimensions[1].height = 20
            ws_z.freeze_panes = ws_z.cell(row=3, column=id_end + 1)

        _write_zscore_sheet(wb, df_z_rct,
                            sheet_title="Z-scores RCT",
                            row_fill=rct_fill)
        _write_zscore_sheet(wb, df_z_ctl,
                            sheet_title="Z-scores Control",
                            row_fill=ctl_fill)

        wb.save(xlsx_path)
        print(f"[Saved] XLSX -> {xlsx_path}")

    except ImportError:
        print("  [NOTE] openpyxl unavailable — only CSV saved")

    print(f"\nDone.  rows={len(df)}  cols={len(df.columns)}"
          f"  ({len(ext_cols)} extended)\n")
    return df


# =============================================================================
# CLI
# =============================================================================

def parse_args():
    p = argparse.ArgumentParser(
        description="Extract per-subject extended features for Paradigm 2")
    p.add_argument("--dataset",      default="xdash",
                   help="Dataset name (must match datasets/ folder). Default: xdash")
    p.add_argument("--data-dir",     default=None,
                   help="Default: storage/pickled/<dataset>")
    p.add_argument("--px-details",   default=None,
                   help="Default: storage/raw/<dataset>/<dataset>_px_details.xlsx")
    p.add_argument("--features-dir", default="features",
                   help="Path to features/ folder with all module .py files")
    p.add_argument("--out-dir",      default=None,
                   help="Default: storage/results/<dataset>/hmm/subject_features")
    p.add_argument("--tasks",        nargs="+", type=int, default=list(range(1, 2)))
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()

    for _root in (Path.cwd(), *Path(__file__).resolve().parents):
        if (_root / "config" / "paths.py").exists():
            sys.path.insert(0, str(_root))
            break
    from config.paths import get_pickled_dir, get_raw_dir, get_results_dir

    data_dir  = Path(args.data_dir) if args.data_dir else get_pickled_dir(args.dataset)
    px_path   = Path(args.px_details) if args.px_details else get_raw_dir(args.dataset) / f"{args.dataset}_px_details.xlsx"
    if args.out_dir:
        out_dir = Path(args.out_dir)
    else:
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = get_results_dir(args.dataset) / "hmm" / f"subject_features_{timestamp}"

    run_extraction(
        data_dir        = data_dir,
        px_details_path = px_path,
        out_dir         = out_dir,
        tasks           = args.tasks,
        features_dir    = Path(args.features_dir) if args.features_dir else None,
    )