"""
generate_hmm_analysis_table.py
================================
Generates a 4-sheet Excel analysis table for one HMM task×paradigm result.

Bug fixes in this version:
  1. FP/FN subject IDs are now correct (fx not PX for controls).
     Root cause: y_true/y_pred in the checkpoint are stored in
     np.unique('g0_i_sid', 'g1_i_sid') order — controls FIRST
     (g0_ < g1_ lexicographically), within each group in lexicographic
     index order (0,10,11,...,2,3,...).  get_ckpt_subject_order()
     reproduces this exact ordering for correct mapping.

  2. Per-subject sheet now shows patients first then controls with correct
     Group, True Label, Predicted, and HMM Prob columns.

  3. State % columns are created dynamically per n_states value — if a
     model has 3 states, State 0 %, State 1 %, State 2 % are all shown.

  4. Sheet 3 now has an Emission Importance column with a clear note
     explaining where to find the values (diagnostics PNG) and how to
     add them numerically.

Usage:
    python scripts/generate_hmm_analysis_table.py --task 1 --paradigm 1
    python scripts/generate_hmm_analysis_table.py --all
"""
import argparse
import glob
import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from sklearn.metrics import (precision_recall_fscore_support,
                             accuracy_score, balanced_accuracy_score,
                             confusion_matrix)

# ── Constants ─────────────────────────────────────────────────────────────────

TASK_NAMES   = {1:"jar_opening",2:"key_turning",3:"cleaning",
                4:"back_washing",5:"cutting",6:"hammering"}
TASK_LABELS  = {1:"Jar Opening",2:"Key Turning",3:"Cleaning",
                4:"Back Washing",5:"Cutting",6:"Hammering"}
PARADIGM_NAMES = {1:"patients_vs_controls",2:"rct_vs_controls",
                  3:"other_conditions_vs_controls",4:"rct_vs_other_conditions"}
PARADIGM_G1  = {1:"Patients",      2:"RCT Patients",   3:"Other Patients", 4:"RCT Patients"}
PARADIGM_G0  = {1:"Controls",      2:"Controls",       3:"Controls",       4:"Other Patients"}
TASK_Q_COL   = {1:("q1_jar","Q1: Open jar"),    2:("q2_key","Q2: Turn key"),
                3:("q3_household","Q3: Household"),4:("q4_back_wash","Q4: Wash back"),
                5:("q5_knife","Q5: Cut food"),   6:("q6_recreational","Q6: Recreational")}
# Exact dia_code → diagnosis from xdash_px_details.xlsx
DIA_NAMES    = {0:"Healthy/Control", 1:"Rotator cuff tear",
                2:"Glenohumeral arthritis", 3:"Biceps tendonitis", 4:"Bursitis"}

C = {"hdr_id":"1F4E79","hdr_dash":"7B5C1E","hdr_clf":"7B2C2C",
     "hdr_state":"4A3060","hdr_align":"3B3B7B",
     "g1_row":"DDEEFF","g0_row":"E8F5E0",
     "fn_row":"FFD7D7","fp_row":"FFE5B4","flat_row":"FF9999","hdr_notes":"2D5016"}


# ── Style helpers ─────────────────────────────────────────────────────────────

def _bdr():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, r, c, val, bg, fc="FFFFFF", bold=True, size=9, wrap=True, ha="center"):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = Font(bold=bold, name="Arial", size=size, color=fc)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    cell.border    = _bdr()
    return cell

def dat(ws, r, c, val, bg=None, bold=False, fc="000000", ha="center", fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = Font(name="Arial", size=9, bold=bold, color=fc)
    cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=True)
    cell.border    = _bdr()
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    if fmt:
        cell.number_format = fmt
    return cell

def grad(ws, col_letter, r1, r2, lo, hi):
    ws.conditional_formatting.add(
        f"{col_letter}{r1}:{col_letter}{r2}",
        ColorScaleRule(start_type="min", start_color=lo,
                       end_type="max",   end_color=hi))


# ── File discovery ────────────────────────────────────────────────────────────

def _find(pattern):
    m = sorted(glob.glob(pattern, recursive=True))
    return Path(m[-1]) if m else None

def load_results_json(hmm_dir, t, p):
    for pat in [
        str(hmm_dir/f"T{t}-P{p}-HMM*"/"results"/f"results_T{t}_P{p}_HMM_variable_length.json"),
        str(hmm_dir/"**"/f"results_T{t}_P{p}_HMM_variable_length.json"),
    ]:
        f = _find(pat)
        if f:
            return json.load(open(f))
    return None

def load_checkpoint(hmm_dir, t, p):
    for pat in [
        str(hmm_dir/f"T{t}-P{p}-HMM*"/"model_checkpoints"/f"HMM_T{t}_P{p}_BA*.json"),
        str(hmm_dir/"**"/f"HMM_T{t}_P{p}_BA*.json"),
    ]:
        f = _find(pat)
        if f:
            return json.load(open(f))
    return None

def load_alignment(hmm_dir, t, p):
    for pat in [
        str(hmm_dir/f"T{t}-P{p}-HMM*"/"diagnostics"/f"alignment_T{t}_P{p}.csv"),
        str(hmm_dir/"**"/f"alignment_T{t}_P{p}.csv"),
    ]:
        f = _find(pat)
        if f:
            return pd.read_csv(f)
    return None

def load_state_seq(state_seq_dir, t, p):
    p_ = Path(state_seq_dir) / f"T{t}_P{p}" / "summary.csv"
    return pd.read_csv(p_) if p_.exists() else None

def load_px_details(path):
    if not Path(path).exists():
        return None
    for h in range(6):
        df = pd.read_excel(path, sheet_name="Sheet1", header=h)
        if "dia_code" in df.columns:
            df["id"] = df["id"].astype(str).str.strip().str.upper()
            return df
    return None


def load_emission_csv(path):
    """Load optional emission importance CSV. Returns None if path is None or missing."""
    if path is None: return None
    p = Path(path)
    if not p.exists():
        print(f"  [WARN] Emission CSV not found: {p}")
        return None
    return pd.read_csv(p)


def build_emission_map(df_emission):
    """feature → {global: float} or {state0: float, state1: float, ...}"""
    if df_emission is None: return {}
    out = {}
    df = df_emission.copy()
    df.columns = [c.strip().lower() for c in df.columns]
    feat_col = next((c for c in df.columns if "feature" in c), df.columns[0])
    for _, row in df.iterrows():
        feat = str(row[feat_col]).strip()
        if "global_importance" in df.columns:
            out[feat] = {"global": float(row["global_importance"])}
        else:
            sc = [c for c in df.columns if c.startswith("state") and "importance" in c]
            out[feat] = {c.replace("_importance",""): float(row[c]) for c in sc if not pd.isna(row[c])}
    return out



# ── Subject list builder ──────────────────────────────────────────────────────

def get_subject_lists(paradigm, df_px, n_g1_fallback, n_g0_fallback, df_align=None):
    """
    Return (g1_ids, g0_ids) for display order: group 1 first, group 0 second.
    Uses per-paradigm alignment CSV when available and correctly sized.
    Falls back to px_details with paradigm-aware filtering.
    """
    if df_px is not None:
        px   = df_px[df_px["id"].str.startswith("PX")].copy()
        fx   = df_px[~df_px["id"].str.startswith("PX")].copy()
        rct  = px[px["dia_code"] == 1]["id"].tolist()
        oth  = px[px["dia_code"] != 1]["id"].tolist()
        fxl  = ["fx" + r["id"].lstrip("FX").lstrip("fx").zfill(2)
                 for _, r in fx.iterrows()]
        expected = {1: len(px)+len(fxl), 2: len(rct)+len(fxl),
                    3: len(oth)+len(fxl), 4: len(rct)+len(oth)}
    else:
        rct = oth = fxl = []
        expected = {}

    # Validate alignment CSV covers the right subjects
    if df_align is not None:
        unique_sids = df_align["subject_id"].unique()
        n_csv = len(unique_sids)
        exp_n = expected.get(paradigm)
        if exp_n is not None and n_csv == exp_n:
            # Use alignment CSV order (ground truth for this paradigm)
            seen = {}
            for sid in df_align["subject_id"]:
                if sid not in seen:
                    seen[sid] = len(seen)
            ordered = sorted(seen, key=seen.get)
            if paradigm in (1, 2, 3):
                g1 = [s for s in ordered if s.upper().startswith("PX")]
                g0 = [s for s in ordered if not s.upper().startswith("PX")]
            else:  # P4: split by RCT
                rct_set = set(rct) if df_px is not None else set()
                g1 = [s for s in ordered if s.upper() in {r.upper() for r in rct_set}]
                g0 = [s for s in ordered if s.upper() not in {r.upper() for r in rct_set}]
            return g1, g0
        elif exp_n is not None:
            print(f"  [WARN] Alignment CSV has {n_csv} subjects but P{paradigm} "
                  f"expects {exp_n} — using px_details fallback")

    # px_details fallback
    if df_px is not None:
        if paradigm == 1:  return px["id"].tolist(), fxl
        elif paradigm == 2: return rct, fxl
        elif paradigm == 3: return oth, fxl
        elif paradigm == 4: return rct, oth

    # Last resort
    return ([f"PX{str(i+1).zfill(2)}" for i in range(n_g1_fallback)],
            [f"fx{str(i+1).zfill(2)}" for i in range(n_g0_fallback)])


# ── KEY FIX: Reproduce checkpoint y_true/y_pred ordering ─────────────────────

def get_ckpt_subject_order(g1_ids, g0_ids):
    """
    The checkpoint stores y_true/y_pred in the order produced by
    np.unique() on subject tag strings built as:
       Group 1 → 'g1_{idx}_{sid}'   e.g. 'g1_0_PX01', 'g1_1_PX02', ...
       Group 0 → 'g0_{idx}_{sid}'   e.g. 'g0_0_fx01', 'g0_1_fx02', ...

    np.unique() sorts lexicographically:
      - 'g0_...' < 'g1_...'  → controls come FIRST
      - Within each group, numeric index as string: '0','1','10','11',...,'2',...

    This function reproduces that exact ordering so y_true[i] maps
    correctly to the returned subject ID at position i.
    """
    raw = ([f"g0_{i}_{sid}" for i, sid in enumerate(g0_ids)] +
           [f"g1_{i}_{sid}" for i, sid in enumerate(g1_ids)])
    return [s.split("_", 2)[2] for s in np.unique(raw)]


def compute_per_class_metrics(y_true, y_pred, g1_lbl, g0_lbl):
    """
    Per-class precision/recall/F1/support from checkpoint y_true/y_pred.
    Source: sklearn.metrics.precision_recall_fscore_support
    Returns dict with keys 'accuracy', 'g0', 'g1'.
    Returns None if y_true is empty.
    """
    if len(y_true) == 0: return None
    p, r, f, s = precision_recall_fscore_support(y_true, y_pred, labels=[0,1])
    acc = accuracy_score(y_true, y_pred)
    return {
        "accuracy": acc,
        "g0": {"label":g0_lbl, "precision":p[0], "recall":r[0], "f1":f[0], "support":int(s[0])},
        "g1": {"label":g1_lbl, "precision":p[1], "recall":r[1], "f1":f[1], "support":int(s[1])},
    }


# ── Lookup maps ───────────────────────────────────────────────────────────────

def build_alignment_map(df_align):
    """subject_id.upper() → {n_events, match_rate, mean_error_s}"""
    if df_align is None:
        return {}
    out = {}
    for sid, grp in df_align.groupby("subject_id"):
        n_ev   = len(grp)
        n_ok   = int(grp["matched"].sum())
        errors = grp["temporal_error_s"].dropna()
        out[sid.upper()] = {
            "n_events":     n_ev,
            "match_rate":   round(n_ok / max(n_ev, 1), 3),
            "mean_error_s": round(float(errors.mean()), 3) if len(errors) else None,
        }
    return out

def build_state_seq_map(ss_df):
    """
    Returns (map, state_cols).
    map: subject_id.upper() → {total_s, n_transitions, state0_pct, state1_pct, ...}
    state_cols: e.g. ['state0_pct','state1_pct'] or ['state0_pct','state1_pct','state2_pct']
    Dynamically detects all stateN_pct columns so >2-state models work correctly.
    """
    if ss_df is None:
        return {}, ["state0_pct", "state1_pct"]

    state_cols = sorted(
        [c for c in ss_df.columns if c.startswith("state") and c.endswith("_pct")],
        key=lambda c: int(c.replace("state","").replace("_pct",""))
    ) or ["state0_pct", "state1_pct"]

    out = {}
    for _, row in ss_df.iterrows():
        sid = str(row["subject_id"]).upper()
        entry = {"total_s": row.get("total_s"), "n_transitions": row.get("n_transitions")}
        for sc in state_cols:
            entry[sc] = row.get(sc)
        out[sid] = entry
    return out, state_cols

def build_dash_map(df_px, q_col):
    """subject_id.upper() → {dia_code, diagnosis, c_dash, x_dash, avg_dash, q_task}"""
    if df_px is None:
        return {}
    out = {}
    for _, row in df_px.iterrows():
        sid      = str(row.get("id","")).upper()
        dia_code = row.get("dia_code", None)
        diag_raw = row.get("diagnosis", None)
        if pd.isna(diag_raw) or str(diag_raw).strip() in ("","nan","NA","NaN"):
            diag_str = DIA_NAMES.get(int(dia_code), "Unknown") if dia_code is not None else "Healthy/Control"
        else:
            diag_str = str(diag_raw).strip()
        out[sid] = {
            "dia_code": dia_code,
            "diagnosis": diag_str,
            "c_dash":    row.get("C_DASH"),
            "x_dash":    row.get("X_DASH"),
            "avg_dash":  row.get("c and x"),
            "q_task":    row.get(q_col),
        }
    return out


def build_emission_map(df_emission):
    """feature → {global: float} or {state0: float, state1: float, ...}"""
    if df_emission is None: return {}
    out = {}
    df = df_emission.copy()
    df.columns = [c.strip().lower() for c in df.columns]
    feat_col = next((c for c in df.columns if "feature" in c), df.columns[0])
    for _, row in df.iterrows():
        feat = str(row[feat_col]).strip()
        if "global_importance" in df.columns:
            out[feat] = {"global": float(row["global_importance"])}
        else:
            sc = [c for c in df.columns if c.startswith("state") and "importance" in c]
            out[feat] = {c.replace("_importance",""): float(row[c]) for c in sc}
    return out




# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — MODEL SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════

def sheet_summary(ws, results, ckpt, task, paradigm, g1_ids, g0_ids, pcm=None):
    g1_lbl = PARADIGM_G1[paradigm]
    g0_lbl = PARADIGM_G0[paradigm]
    bp     = results.get("best_params", {})
    n_st   = bp.get("n_components", "?")
    m      = results.get("metrics", {})
    ev     = (ckpt or {}).get("metrics", {})

    # Title
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = (f"HMM Variable Length  —  Task {task}: {TASK_LABELS[task]}  |  "
               f"Paradigm {paradigm}: {g1_lbl} vs {g0_lbl}  |  n_states={n_st}")
    c.font      = Font(bold=True, name="Arial", size=12, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color=C["hdr_clf"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # ── Metrics ───────────────────────────────────────────────────────────────
    metrics_rows = [
        ("Metric", "Value", "Notes"),
        ("Balanced Accuracy",
         f"{m.get('ba', ev.get('balanced_accuracy','?')):.4f}" if isinstance(m.get('ba'), float) else "?",
         "Primary — handles class imbalance"),
        ("Accuracy",
         f"{ev.get('accuracy','?'):.4f}" if isinstance(ev.get('accuracy'), float) else "?",
         "Overall proportion correct"),
        ("Recall (Sensitivity)",
         f"{m.get('recall','?'):.4f}" if isinstance(m.get('recall'), float) else "?",
         f"{g1_lbl} correctly identified"),
        ("Precision",
         f"{m.get('precision','?'):.4f}" if isinstance(m.get('precision'), float) else "?",
         "Positive predictive value"),
        ("F1 Score",
         f"{m.get('f1','?'):.4f}" if isinstance(m.get('f1'), float) else "?",
         "Harmonic mean precision/recall"),
        ("AUC-ROC",
         f"{m.get('auc','?'):.4f}" if isinstance(m.get('auc'), float) else "?",
         "Area under ROC curve"),
        ("AUC 95% CI",
         f"[{m.get('auc_ci_low','?'):.3f}, {m.get('auc_ci_high','?'):.3f}]"
         if isinstance(m.get('auc_ci_low'), float) else "?",
         "Wide CI reflects small N"),
    ]
    for i, (a, b, note) in enumerate(metrics_rows):
        r = i + 3
        is_h = (i == 0)
        hdr(ws, r, 1, a, "2C4770" if is_h else "EAEAEA",
            fc="FFFFFF" if is_h else "000000", size=9)
        dat(ws, r, 2, b, bold=not is_h)
        dat(ws, r, 3, note, ha="left")


    # ── Per-class metrics ─────────────────────────────────────────────────────
    pc_r = 14
    ws.merge_cells(f"A{pc_r}:D{pc_r}")
    hdr(ws, pc_r, 1,
        "PER-CLASS METRICS  (source: checkpoint y_true/y_pred → sklearn)" if pcm
        else "PER-CLASS METRICS  (checkpoint not found — N/A)",
        "2C4770", size=10)
    for ci, h in enumerate(["Class","Label","Precision","Recall","F1","Support"]):
        hdr(ws, pc_r+1, ci+1, h, "2C4770", fc="FFFFFF", size=9)
    if pcm:
        for i,(cls_key, bg) in enumerate([("g0","E8F5E0"),("g1","DDEEFF")]):
            d_ = pcm[cls_key]; r = pc_r+2+i
            dat(ws,r,1,f"Class {'0' if cls_key=='g0' else '1'}",bg=bg,bold=True)
            dat(ws,r,2,d_["label"],bg=bg)
            dat(ws,r,3,f"{d_['precision']:.4f}",bg=bg)
            dat(ws,r,4,f"{d_['recall']:.4f}",bg=bg)
            dat(ws,r,5,f"{d_['f1']:.4f}",bg=bg)
            dat(ws,r,6,d_["support"],bg=bg)
    else:
        ws.merge_cells(f"A{pc_r+2}:F{pc_r+2}")
        dat(ws,pc_r+2,1,"Run with checkpoint JSON to populate.",fc="888888",ha="left")

    # ── Hyperparameters ───────────────────────────────────────────────────────
    hp_r = 20
    ws.merge_cells(f"A{hp_r}:C{hp_r}")
    hdr(ws, hp_r, 1, "Best Hyperparameters", C["hdr_state"], size=10)
    hdr(ws, hp_r+1, 1, "Parameter",  "2C4770", fc="FFFFFF", size=9)
    hdr(ws, hp_r+1, 2, "Value",      "2C4770", fc="FFFFFF", size=9)
    hdr(ws, hp_r+1, 3, "Description","2C4770", fc="FFFFFF", size=9)
    hp_rows = [
        ("n_components (n_states)", n_st,
         "Number of hidden states — optimised per T×P via LOOCV grid search"),
        ("covariance_type", bp.get("covariance_type","?"),
         "Gaussian emission covariance structure"),
        ("n_iter",  bp.get("n_iter","?"),   "Maximum EM iterations"),
        ("method",  results.get("preprocessing_method","variable_length"),
         "Full sequence kept per subject (no truncation)"),
    ]
    for i, (k, v, d) in enumerate(hp_rows):
        bg = "FFF2CC" if i == 0 else None
        dat(ws, hp_r+2+i, 1, k, bg=bg, bold=(i==0))
        dat(ws, hp_r+2+i, 2, str(v), bg=bg, bold=(i==0))
        dat(ws, hp_r+2+i, 3, d, bg=bg, ha="left")

    # ── Dataset composition ───────────────────────────────────────────────────
    ds_r = hp_r + 8
    ws.merge_cells(f"A{ds_r}:C{ds_r}")
    hdr(ws, ds_r, 1, "Dataset Composition", C["hdr_id"], size=10)
    hdr(ws, ds_r+1, 1, "Group",    "2C4770", fc="FFFFFF", size=9)
    hdr(ws, ds_r+1, 2, "N",        "2C4770", fc="FFFFFF", size=9)
    hdr(ws, ds_r+1, 3, "Subjects", "2C4770", fc="FFFFFF", size=9)
    dat(ws, ds_r+2, 1, f"Group 1: {g1_lbl}", bold=True)
    dat(ws, ds_r+2, 2, len(g1_ids), bold=True)
    dat(ws, ds_r+2, 3, ", ".join(g1_ids[:8]) + ("..." if len(g1_ids)>8 else ""), ha="left")
    dat(ws, ds_r+3, 1, f"Group 0: {g0_lbl}", bold=True)
    dat(ws, ds_r+3, 2, len(g0_ids), bold=True)
    dat(ws, ds_r+3, 3, ", ".join(g0_ids[:8]) + ("..." if len(g0_ids)>8 else ""), ha="left")
    dat(ws, ds_r+4, 1, "Total", bold=True)
    dat(ws, ds_r+4, 2, len(g1_ids)+len(g0_ids), bold=True)

    # ── Confusion matrix ──────────────────────────────────────────────────────
    preds  = (ckpt or {}).get("predictions", {})
    y_true = np.array(preds.get("y_true", []))
    y_pred = np.array(preds.get("y_pred", []))

    cm_r = ds_r + 7
    ws.merge_cells(f"A{cm_r}:C{cm_r}")
    hdr(ws, cm_r, 1, "Confusion Matrix", C["hdr_clf"], size=10)
    hdr(ws, cm_r+1, 1, "",                  "DDDDDD", fc="000000")
    hdr(ws, cm_r+1, 2, f"Pred: {g0_lbl}",  "2C6E49")
    hdr(ws, cm_r+1, 3, f"Pred: {g1_lbl}",  C["hdr_clf"])
    hdr(ws, cm_r+2, 1, f"True: {g0_lbl}",  "2C6E49")
    hdr(ws, cm_r+3, 1, f"True: {g1_lbl}",  C["hdr_clf"])

    if len(y_true):
        tn = int(((y_true==0)&(y_pred==0)).sum())
        fp = int(((y_true==0)&(y_pred==1)).sum())
        fn = int(((y_true==1)&(y_pred==0)).sum())
        tp = int(((y_true==1)&(y_pred==1)).sum())
    else:
        tn=fp=fn=tp="?"

    dat(ws, cm_r+2, 2, tn, bg="C8E6C9", bold=True)
    dat(ws, cm_r+2, 3, fp, bg=C["fp_row"])
    dat(ws, cm_r+3, 2, fn, bg=C["fn_row"])
    dat(ws, cm_r+3, 3, tp, bg="C8E6C9", bold=True)

    # ── Misclassified subjects — uses get_ckpt_subject_order ─────────────────
    mc_r = cm_r + 6
    fn_ids, fp_ids = [], []
    if len(y_true):
        ckpt_order = get_ckpt_subject_order(g1_ids, g0_ids)
        for i, sid in enumerate(ckpt_order):
            if i >= len(y_true):
                break
            yt, yp = int(y_true[i]), int(y_pred[i])
            if yt == 1 and yp == 0:
                fn_ids.append(sid)
            elif yt == 0 and yp == 1:
                fp_ids.append(sid)

    ws.merge_cells(f"A{mc_r}:C{mc_r}")
    hdr(ws, mc_r, 1, "Misclassified Subjects", C["hdr_clf"], size=10)
    hdr(ws, mc_r+1, 1, "Type",        "2C4770", fc="FFFFFF")
    hdr(ws, mc_r+1, 2, "Subject IDs", "2C4770", fc="FFFFFF")
    hdr(ws, mc_r+1, 3, "Count",       "2C4770", fc="FFFFFF")
    dat(ws, mc_r+2, 1, f"False Negatives (missed {g1_lbl})", bold=True, fc="CC0000")
    dat(ws, mc_r+2, 2, ", ".join(fn_ids) if fn_ids else "None", bg=C["fn_row"], ha="left")
    dat(ws, mc_r+2, 3, len(fn_ids), bg=C["fn_row"], bold=True)
    dat(ws, mc_r+3, 1, f"False Positives ({g0_lbl} → {g1_lbl})", bold=True, fc="FF6600")
    dat(ws, mc_r+3, 2, ", ".join(fp_ids) if fp_ids else "None", bg=C["fp_row"], ha="left")
    dat(ws, mc_r+3, 3, len(fp_ids), bg=C["fp_row"], bold=True)

    for col, w in zip("ABCDEFG", [28, 18, 55, 10, 10, 10, 10]):
        ws.column_dimensions[col].width = w


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — PER-SUBJECT ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════════

def sheet_per_subject(ws, task, paradigm, g1_ids, g0_ids,
                      dash_map, alignment_map, state_seq_map, state_cols,
                      y_true, y_pred, y_proba, n_states):

    g1_lbl   = PARADIGM_G1[paradigm]
    g0_lbl   = PARADIGM_G0[paradigm]
    q_col, q_desc = TASK_Q_COL.get(task, ("q1_jar","Q: Task score"))

    # ── Build prediction lookup using correct checkpoint order ────────────────
    ckpt_order = get_ckpt_subject_order(g1_ids, g0_ids)
    pred_map = {}
    for i, sid in enumerate(ckpt_order):
        if i < len(y_true):
            pred_map[sid.upper()] = {
                "y_true":  int(y_true[i]),
                "y_pred":  int(y_pred[i]),
                "y_proba": float(y_proba[i]) if i < len(y_proba) else None,
            }

    fn_set, fp_set = set(), set()
    for sid in g1_ids + g0_ids:
        pm = pred_map.get(sid.upper(), {})
        yt, yp = pm.get("y_true"), pm.get("y_pred")
        if yt == 1 and yp == 0:
            fn_set.add(sid.upper())
        elif yt == 0 and yp == 1:
            fp_set.add(sid.upper())

    state_pct_hdrs = [(f"State {i} %", 10) for i in range(len(state_cols))]

    COLS = [
        ("IDENTITY",       "Subject ID",                 10, C["hdr_id"]),
        ("IDENTITY",       "Group",                      12, C["hdr_id"]),
        ("IDENTITY",       "Diagnosis",                  22, C["hdr_id"]),
        ("IDENTITY",       f"n_states\n(HMM)",            9, C["hdr_id"]),
        ("DASH SCORES",    "C_DASH",                     10, C["hdr_dash"]),
        ("DASH SCORES",    "X_DASH",                     10, C["hdr_dash"]),
        ("DASH SCORES",    "Avg DASH",                   10, C["hdr_dash"]),
        ("DASH SCORES",    f"{q_desc}\n(self-report)",   13, C["hdr_dash"]),
        ("CLASSIFICATION", "True Label",                 13, C["hdr_clf"]),
        ("CLASSIFICATION", "Predicted",                  13, C["hdr_clf"]),
        ("CLASSIFICATION", "Correct?",                    8, C["hdr_clf"]),
        ("CLASSIFICATION", "HMM Prob",                   10, C["hdr_clf"]),
        ("STATE SEQUENCE", "Duration (s)",               10, C["hdr_state"]),
        ("STATE SEQUENCE", "N Transitions",              11, C["hdr_state"]),
    ] + [
        ("STATE SEQUENCE", lbl, w, C["hdr_state"]) for lbl, w in state_pct_hdrs
    ] + [
        ("EVENT ALIGNMENT","N Events",                    9, C["hdr_align"]),
        ("EVENT ALIGNMENT","Match Rate\n@0.5s",          11, C["hdr_align"]),
        ("EVENT ALIGNMENT","Mean Err (s)",               11, C["hdr_align"]),
        ("KEY FINDINGS",   "Notes",                      45, C["hdr_notes"]),
    ]

    sec_info = {}
    for i, (sec, _, _, _) in enumerate(COLS):
        if sec not in sec_info:
            sec_info[sec] = {"s": i+1, "e": i+1, "bg": COLS[i][3]}
        else:
            sec_info[sec]["e"] = i+1
    for sec, info in sec_info.items():
        s, e = info["s"], info["e"]
        if s != e:
            ws.merge_cells(start_row=1, start_column=s, end_row=1, end_column=e)
        hdr(ws, 1, s, sec, info["bg"], size=10)

    for i, (_, lbl, w, bg) in enumerate(COLS):
        hdr(ws, 2, i+1, lbl, bg, size=9)
        ws.column_dimensions[get_column_letter(i+1)].width = w
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "A3"

    for row_i, sid in enumerate(g1_ids + g0_ids):
        r      = row_i + 3
        sid_up = sid.upper()

        pm  = pred_map.get(sid_up, {})
        yt  = pm.get("y_true")
        yp  = pm.get("y_pred")
        yb  = pm.get("y_proba")

        if yt is None:
            yt = 1 if row_i < len(g1_ids) else 0

        ok    = (yt == yp) if yp is not None else None
        group = g1_lbl if yt == 1 else g0_lbl
        t_lbl = g1_lbl if yt == 1 else g0_lbl
        p_lbl = (g1_lbl if yp == 1 else g0_lbl) if yp is not None else "—"

        d        = dash_map.get(sid_up, {})
        dia_code = d.get("dia_code")
        diag_str = DIA_NAMES.get(int(dia_code), str(dia_code)) if dia_code is not None else "—"

        ss       = state_seq_map.get(sid_up, {})
        dur      = ss.get("total_s")
        n_trans  = ss.get("n_transitions")
        is_flat  = (n_trans == 0) if n_trans is not None else False
        s_pcts   = [ss.get(sc) for sc in state_cols]

        al       = alignment_map.get(sid_up, {})
        n_ev     = al.get("n_events")
        match_r  = al.get("match_rate")
        mean_err = al.get("mean_error_s")
        if n_ev and n_ev > 0 and mean_err is None:
            mean_err = "NaN"

        if   sid_up in fn_set:  row_bg = C["fn_row"]
        elif sid_up in fp_set:  row_bg = C["fp_row"]
        elif is_flat:            row_bg = C["flat_row"]
        elif yt == 1:            row_bg = C["g1_row"]
        else:                    row_bg = C["g0_row"]

        vals = [
            (sid,      row_bg, sid_up in fn_set|fp_set, "000000"),
            (group,    row_bg, False,  "000000"),
            (diag_str, row_bg, False,  "000000"),
            (n_states, row_bg, False,  "000000"),
            (d.get("c_dash"),   row_bg, False, "000000"),
            (d.get("x_dash"),   row_bg, False, "000000"),
            (d.get("avg_dash"), row_bg, False, "000000"),
            (d.get("q_task"),   row_bg, False, "000000"),
            (t_lbl,    row_bg, False,  "000000"),
            (p_lbl,    row_bg, ok is False, "CC0000" if ok is False else "000000"),
            ("✓" if ok else "✗" if ok is False else "—",
             row_bg, True,
             "006400" if ok else "CC0000" if ok is False else "888888"),
            (yb,       row_bg, False,  "000000"),
            (dur,      row_bg, False,  "000000"),
            (n_trans,  row_bg, is_flat, "CC0000" if is_flat else "000000"),
        ] + [
            (v, row_bg, False, "000000") for v in s_pcts
        ] + [
            (n_ev,     row_bg, False, "000000"),
            (match_r,  row_bg, False, "000000"),
            (mean_err, row_bg, False, "000000"),
        ]

        state_ci_start = 14
        align_match_ci = state_ci_start + len(state_cols) + 1

        for ci, (val, bg, bold, fc) in enumerate(vals):
            ha = "left" if ci == 2 else "center"
            c  = dat(ws, r, ci+1, val, bg=bg, bold=bold, fc=fc, ha=ha)
            if ci in (4,5,6) and isinstance(val, (int,float)):
                c.number_format = "0.00"
            if ci == 11 and isinstance(val, float):
                c.number_format = "0.000"
            if state_ci_start <= ci < state_ci_start + len(state_cols):
                if isinstance(val, (int,float)):
                    c.number_format = "0.0"
            if ci == align_match_ci and isinstance(val, (int,float)):
                c.number_format = "0.000"

        ws.row_dimensions[r].height = 16

    nd = len(g1_ids) + len(g0_ids)
    if nd > 0:
        er = 2 + nd
        grad(ws, get_column_letter(7),  3, er, "FFFFFF", "C00000")
        grad(ws, get_column_letter(12), 3, er, "FFFFFF", "7B2C2C")
        grad(ws, get_column_letter(14 + len(state_cols) + 1), 3, er, "FFFFFF", "2E75B6")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — FEATURE IMPORTANCE
# ═══════════════════════════════════════════════════════════════════════════════

def sheet_feature_importance(ws, results, task, paradigm, emission_map=None):
    n_st = results.get("best_params",{}).get("n_components","?")
    ws.merge_cells("A1:E1")
    hdr(ws, 1, 1,
        f"Feature Importance — Task {task} P{paradigm} | n_states={n_st} | "
        f"Permutation (classification) and Emission (structural, see note)",
        C["hdr_clf"], size=11)
    ws.row_dimensions[1].height = 22

    hdrs = ["Rank", "Feature", "Permutation\nImportance",
            "Normalised\nRank", "Emission Importance\n(fill from diagnostics PNG)"]
    for i, h in enumerate(hdrs):
        hdr(ws, 2, i+1, h, C["hdr_clf"])

    fi = sorted(results.get("feature_importance",{}).items(),
                key=lambda x: x[1], reverse=True)
    n = len(fi)
    for rank, (feat, imp) in enumerate(fi, 1):
        bg = "FFF2CC" if rank <= 6 else None
        dat(ws, rank+2, 1, rank, bg=bg)
        dat(ws, rank+2, 2, feat, bg=bg, ha="left")
        c = dat(ws, rank+2, 3, imp, bg=bg)
        c.number_format = "0.0000"
        norm = dat(ws, rank+2, 4, f"=C{rank+2}/MAX($C$3:$C${n+2})", bg=bg)
        norm.number_format = "0.0%"
        dat(ws, rank+2, 5, None, bg=bg)

    grad(ws, "C", 3, 3+n, "FFFFFF", "1F4E79")
    for col, w in zip("ABCDE", [8, 24, 22, 16, 28]):
        ws.column_dimensions[col].width = w

    note_r = n + 4
    ws.merge_cells(f"A{note_r}:E{note_r}")
    nc = ws.cell(row=note_r, column=1,
                 value=("Permutation Importance (col C): drop in BA when this channel is "
                        "randomly shuffled — higher = more discriminative for classification. "
                        "Emission Importance (col E): mean absolute emission magnitude per hidden "
                        "state, reflecting kinematic signature of each movement phase. "
                        "Values saved as diagnostics/state_importance_patient.png and "
                        "state_importance_control.png. "
                        "To populate numerically: add a CSV export in your diagnostics "
                        "pipeline (model.compute_state_specific_importance → save to CSV), "
                        "then re-run this script with --emission-csv path/to/file.csv."))
    nc.font      = Font(name="Arial", size=9, italic=True, color="555555")
    nc.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[note_r].height = 52


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — LEGEND
# ═══════════════════════════════════════════════════════════════════════════════

def sheet_legend(ws, paradigm, n_states):
    g1_lbl = PARADIGM_G1[paradigm]
    g0_lbl = PARADIGM_G0[paradigm]

    ws.merge_cells("A1:C1")
    hdr(ws, 1, 1, f"Legend — Paradigm {paradigm}: {g1_lbl} vs {g0_lbl}", "333333", size=12)
    ws.row_dimensions[1].height = 22

    hdr(ws, 3, 1, "ROW COLOURS", "2C4770", fc="FFFFFF", size=9)
    ws.merge_cells("B3:C3")
    colour_rows = [
        (C["g1_row"],  "000000", f"Group 1 — {g1_lbl} — correctly classified (True Positive)"),
        (C["g0_row"],  "000000", f"Group 0 — {g0_lbl} — correctly classified (True Negative)"),
        (C["fn_row"],  "CC0000", f"False Negative — {g1_lbl} subject predicted as {g0_lbl}"),
        (C["fp_row"],  "FF6600", f"False Positive — {g0_lbl} subject predicted as {g1_lbl}"),
        (C["flat_row"],"CC0000", "Flat sequence — zero HMM transitions detected"),
    ]
    for i, (bg, fc, txt) in enumerate(colour_rows):
        r = i + 4
        sw = ws.cell(row=r, column=1, value="  ")
        sw.fill   = PatternFill("solid", start_color=bg)
        sw.border = _bdr()
        ws.merge_cells(f"B{r}:C{r}")
        ws.cell(row=r, column=2, value=txt).font = Font(name="Arial", size=10, color=fc)
        ws.row_dimensions[r].height = 16

    pg_r = len(colour_rows) + 6
    hdr(ws, pg_r, 1, "PARADIGM GROUPS", "2C4770", fc="FFFFFF", size=9)
    ws.merge_cells(f"B{pg_r}:C{pg_r}")
    p_info = {
        1: [("Group 1 (label=1)", "All 40 PX patients — RCT (n=25), GA (n=10), Bursitis (n=2+1), Biceps (n=3)"),
            ("Group 0 (label=0)", "All 20 healthy controls (fx01–fx20)")],
        2: [("Group 1 (label=1)", "25 RCT patients only (dia_code=1)"),
            ("Group 0 (label=0)", "All 20 healthy controls (fx01–fx20)")],
        3: [("Group 1 (label=1)", "15 non-RCT patients — GA (n=10), Bursitis, Biceps (dia_code≠1)"),
            ("Group 0 (label=0)", "All 20 healthy controls (fx01–fx20)")],
        4: [("Group 1 (label=1)", "25 RCT patients (dia_code=1)"),
            ("Group 0 (label=0)", "15 non-RCT patients — GA, Bursitis, Biceps (dia_code≠1)"),
            ("⚠ Note", "No healthy controls in P4 — both groups are patients")],
    }
    for i, (k, v) in enumerate(p_info.get(paradigm, [])):
        r = pg_r + 1 + i
        dat(ws, r, 1, k, bold=True)
        ws.merge_cells(f"B{r}:C{r}")
        ws.cell(row=r, column=2, value=v).font = Font(name="Arial", size=10)
        ws.row_dimensions[r].height = 16

    cd_r = pg_r + len(p_info.get(paradigm,[])) + 3
    hdr(ws, cd_r, 1, "COLUMN DEFINITIONS", "2C4770", fc="FFFFFF", size=9)
    ws.merge_cells(f"B{cd_r}:C{cd_r}")
    col_defs = [
        ("n_states", f"Number of hidden states = {n_states}, optimised per T×P via LOOCV"),
        ("C_DASH", "Computer-administered DASH (0–100, higher = more disability)"),
        ("X_DASH", "XR-administered DASH — answered inside the XR headset"),
        ("Avg DASH", "Mean of C_DASH and X_DASH"),
        ("Q score", "Self-reported difficulty for this specific task: 1=No difficulty … 5=Unable"),
        ("HMM Prob", "Sigmoid of log P(seq|HMM_g1)−log P(seq|HMM_g0). >0.5 → predicted Group 1"),
        ("N Transitions", "Viterbi-decoded state switches. 0 = flat/degenerate sequence"),
        ("State N %", "% of frames in hidden state N. States ordered 0,1,...,n_states−1"),
        ("Match Rate @0.5s", "Fraction of annotated events with HMM transition within 0.5s"),
        ("Mean Err (s)", "Mean temporal error (s) between events and nearest HMM transition. NaN = flat"),
        ("Dia Code", "1=Rotator cuff tear, 2=Glenohumeral arthritis, 3=Biceps tendonitis, 4=Bursitis"),
        ("✓/✗", "Correctly (✓) or incorrectly (✗) classified by LOOCV"),
    ]
    for i, (k, v) in enumerate(col_defs):
        r = cd_r + 1 + i
        dat(ws, r, 1, k, bold=True)
        ws.merge_cells(f"B{r}:C{r}")
        ws.cell(row=r, column=2, value=v).font = Font(name="Arial", size=9)
        ws.row_dimensions[r].height = 16

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 75


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════

def generate(task, paradigm, hmm_dir, state_seq_dir, px_details_path, out_dir, emission_csv_path=None):
    tag = f"T{task}_P{paradigm}"
    print(f"\n{'='*60}")
    print(f"  {tag} — {TASK_LABELS[task]} | "
          f"{PARADIGM_G1[paradigm]} vs {PARADIGM_G0[paradigm]}")
    print(f"{'='*60}")

    results  = load_results_json(hmm_dir, task, paradigm)
    if results is None:
        print(f"  [SKIP] No results JSON found for {tag}")
        return None

    ckpt     = load_checkpoint(hmm_dir, task, paradigm)
    df_align = load_alignment(hmm_dir, task, paradigm)
    ss_df    = load_state_seq(state_seq_dir, task, paradigm)
    df_px    = load_px_details(px_details_path)

    n_st = results["best_params"].get("n_components","?")
    print(f"  BA={results['metrics'].get('ba','?'):.4f}  n_states={n_st}")
    print(f"  ckpt={'✓' if ckpt else '✗'}  "
          f"align={'✓' if df_align is not None else '✗'}  "
          f"state_seq={'✓' if ss_df is not None else '✗'}  "
          f"px_details={'✓' if df_px is not None else '✗'}")

    preds  = (ckpt or {}).get("predictions", {})
    y_true = np.array(preds.get("y_true", []))
    y_pred = np.array(preds.get("y_pred", []))
    y_proba= np.array(preds.get("y_proba",[]))
    n_g1   = int(y_true.sum())  if len(y_true) else 0
    n_g0   = int((y_true==0).sum()) if len(y_true) else 0

    g1_ids, g0_ids = get_subject_lists(paradigm, df_px, n_g1, n_g0, df_align)
    print(f"  G1 ({PARADIGM_G1[paradigm]}): n={len(g1_ids)}  "
          f"G0 ({PARADIGM_G0[paradigm]}): n={len(g0_ids)}")

    g1_lbl       = PARADIGM_G1[paradigm]
    g0_lbl       = PARADIGM_G0[paradigm]
    pcm          = (compute_per_class_metrics(y_true, y_pred, g1_lbl, g0_lbl)
                    if len(y_true) else None)
    df_emission  = load_emission_csv(emission_csv_path)
    q_col, _     = TASK_Q_COL.get(task, ("q1_jar",""))
    dash_map     = build_dash_map(df_px, q_col)
    alignment_map         = build_alignment_map(df_align)
    state_seq_map, state_cols = build_state_seq_map(ss_df)
    emission_map          = build_emission_map(df_emission)
    print(f"  State columns: {state_cols}  |  "
          f"emission_csv={'OK' if df_emission is not None else 'N/A'}")

    wb  = Workbook()
    ws1 = wb.active;                  ws1.title = "Model Summary"
    ws2 = wb.create_sheet("Per-Subject Analysis")
    ws3 = wb.create_sheet("Feature Importance")
    ws4 = wb.create_sheet("Legend")

    sheet_summary(ws1, results, ckpt, task, paradigm, g1_ids, g0_ids, pcm=pcm)
    sheet_per_subject(ws2, task, paradigm, g1_ids, g0_ids,
                      dash_map, alignment_map, state_seq_map, state_cols,
                      y_true, y_pred, y_proba, n_st)
    sheet_feature_importance(ws3, results, task, paradigm, emission_map)
    sheet_legend(ws4, paradigm, n_st)

    Path(out_dir).mkdir(parents=True, exist_ok=True)
    out_path = Path(out_dir) / f"HMM_{tag}_Analysis.xlsx"
    wb.save(out_path)
    print(f"  → {out_path}")
    return out_path


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="Generate HMM comprehensive analysis Excel table per T×P")
    ap.add_argument("--task",     type=int, choices=range(1,7))
    ap.add_argument("--paradigm", type=int, choices=range(1,5))
    ap.add_argument("--all",      action="store_true")
    ap.add_argument("--hmm-dir",       default="hmm-results")
    ap.add_argument("--state-seq-dir", default="hmm-results/state_seqs")
    ap.add_argument("--px-details",    default="data/xdash_px_details.xlsx")
    ap.add_argument("--out",           default="hmm-results/reports")
    ap.add_argument("--emission-csv",  default=None,
                    help="CSV with emission importance from save_emission_importance.py")
    args = ap.parse_args()

    if not args.all and (args.task is None or args.paradigm is None):
        ap.error("Provide --task and --paradigm, or use --all")

    combos = ([(t,p) for t in range(1,7) for p in range(1,5)]
              if args.all else [(args.task, args.paradigm)])

    generated = []
    for task, paradigm in combos:
        p = generate(task, paradigm,
                     Path(args.hmm_dir), Path(args.state_seq_dir),
                     Path(args.px_details), Path(args.out))
        if p:
            generated.append(p)

    print(f"\n{'='*60}")
    print(f"Done. {len(generated)} file(s) → {args.out}/")
    for p in generated:
        print(f"  {p.name}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()