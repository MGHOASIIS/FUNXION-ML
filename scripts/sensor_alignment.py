"""
sensor_alignment.py

Compares sensor/feature importance RANKINGS between:
  OLD: CSV with short names (HPosY, LRotZ…), partial list, raw scores (unnormalised)
  NEW: JSON summaries with full names (head_pos_y, left_rot_z…), all 18 features, normalised

Key design decisions
────────────────────
• Raw scores differ wildly in scale (old ~1–4, new ~0.01–0.11) → compare RANKS only.
• Old CSV lists only top-k features per experiment (k varies, typically 5–8).
  Unreported features are treated as tied-last; comparisons are restricted to
  the reported set wherever that gives a cleaner signal.
• Exact match key: (Task, Paradigm, Model).  FeatureFilter / Method are retained
  for context but not used as a join key.

Sheets
──────
1. Raw Scores Side-by-Side   – normalised scores for every feature, old & new, colour-scaled
2. Rank Comparison           – rank 1-18 for both, Δ highlighted, unreported greyed
3. Top-K Sensor Agreement    – for each experiment: which of old's top-K appear in new's top-K
4. Alignment Heatmap         – sensors × experiments, cell = rank delta (green=aligned, red=not)
5. Spearman per Experiment   – ρ over reported features, sorted best→worst
6. Sensor-Level Summary      – per sensor: avg old rank, avg new rank, avg |Δ rank|
7. Experiment Dashboard      – aggregate stats + interpretation notes
"""

import json, glob, os, ast, re
import numpy as np
import pandas as pd
from scipy.stats import spearmanr
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ── CONFIG ───────────────────────────────────────────────────────────────────
OLD_CSV         = "scripts/old_sensor_results.csv"
EXPERIMENTS_DIR = "experiments_from_hpc"
OUTPUT_FILE     = "sensor_alignment.xlsx"

# ── NAME MAPPING ─────────────────────────────────────────────────────────────
SHORT_TO_FULL = {
    "HPosX": "head_pos_x",        "HPosY": "head_pos_y",        "HPosZ": "head_pos_z",
    "HRotX": "head_rot_x",        "HRotY": "head_rot_y",        "HRotZ": "head_rot_z",
    "RPosX": "right_hand_pos_x",  "RPosY": "right_hand_pos_y",  "RPosZ": "right_hand_pos_z",
    "RRotX": "right_hand_rot_x",  "RRotY": "right_hand_rot_y",  "RRotZ": "right_hand_rot_z",
    "LPosX": "left_hand_pos_x",   "LPosY": "left_hand_pos_y",   "LPosZ": "left_hand_pos_z",
    "LRotX": "left_hand_rot_x",   "LRotY": "left_hand_rot_y",   "LRotZ": "left_hand_rot_z",
}
FULL_TO_SHORT = {v: k for k, v in SHORT_TO_FULL.items()}
ALL_FEATS     = list(SHORT_TO_FULL.values())   # canonical order, 18 features

# Sensor group labels for display
GROUPS = {
    "head_pos_x":"Head Pos","head_pos_y":"Head Pos","head_pos_z":"Head Pos",
    "head_rot_x":"Head Rot","head_rot_y":"Head Rot","head_rot_z":"Head Rot",
    "left_hand_pos_x":"L.Hand Pos","left_hand_pos_y":"L.Hand Pos","left_hand_pos_z":"L.Hand Pos",
    "left_hand_rot_x":"L.Hand Rot","left_hand_rot_y":"L.Hand Rot","left_hand_rot_z":"L.Hand Rot",
    "right_hand_pos_x":"R.Hand Pos","right_hand_pos_y":"R.Hand Pos","right_hand_pos_z":"R.Hand Pos",
    "right_hand_rot_x":"R.Hand Rot","right_hand_rot_y":"R.Hand Rot","right_hand_rot_z":"R.Hand Rot",
}
GROUP_FILLS = {
    "Head Pos":    PatternFill("solid", fgColor="D9E1F2"),
    "Head Rot":    PatternFill("solid", fgColor="BDD7EE"),
    "L.Hand Pos":  PatternFill("solid", fgColor="FFF2CC"),
    "L.Hand Rot":  PatternFill("solid", fgColor="FCE4D6"),
    "R.Hand Pos":  PatternFill("solid", fgColor="E2EFDA"),
    "R.Hand Rot":  PatternFill("solid", fgColor="C6EFCE"),
}

# Model integer → name  (from constants.py)
OLD_MODEL_MAP = {3: "HMM", 4: "CNN", 5: "RNN"}
OLD_FF_MAP    = {1: "padding", 2: "truncating", 3: "dtw_embedding", 4: "sliding_window"}

PARADIGM_NAMES = {
    1:"patients_vs_controls", 2:"rct_vs_controls",
    3:"other_conditions_vs_controls", 4:"rct_vs_other_conditions",
}
TASK_NAMES = {1:"jar_opening",2:"key_turning",3:"cleaning",
              4:"back_washing",5:"cutting",6:"hammering"}

# ── STYLES ───────────────────────────────────────────────────────────────────
THIN    = Side(style="thin",   color="BFBFBF")
MEDIUM  = Side(style="medium", color="444444")
THIN_B  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HDR_FILL      = PatternFill("solid", fgColor="1F3864")
SUBHDR_FILL   = PatternFill("solid", fgColor="2E75B6")
OLD_HDR_FILL  = PatternFill("solid", fgColor="7F6000")   # dark gold header
NEW_HDR_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue header
OLD_DATA_FILL = PatternFill("solid", fgColor="FFF2CC")   # light gold
NEW_DATA_FILL = PatternFill("solid", fgColor="D6E4F7")   # light blue
GREY_FILL     = PatternFill("solid", fgColor="D9D9D9")   # unreported
GREEN_FILL    = PatternFill("solid", fgColor="C6EFCE")
AMBER_FILL    = PatternFill("solid", fgColor="FFEB9C")
RED_FILL      = PatternFill("solid", fgColor="FFC7CE")
NEUTRAL_FILL  = PatternFill("solid", fgColor="F2F2F2")

def _hdr(cell, txt=None, fill=HDR_FILL, size=10, wrap=True):
    if txt is not None:
        cell.value = txt
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=size)
    cell.fill      = fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border    = THIN_B

def _dat(cell, val=None, bold=False, align="center", fill=None, fmt=None, size=9, color="000000"):
    if val is not None:
        cell.value = val
    cell.font      = Font(bold=bold, name="Arial", size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = THIN_B
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt

def _rho_fill(rho):
    if rho is None: return NEUTRAL_FILL
    if rho >= 0.7:  return GREEN_FILL
    if rho >= 0.4:  return AMBER_FILL
    return RED_FILL

def _delta_fill(delta):
    if delta is None: return NEUTRAL_FILL
    d = abs(delta)
    if d == 0:  return GREEN_FILL
    if d <= 2:  return PatternFill("solid", fgColor="E2EFDA")
    if d <= 4:  return AMBER_FILL
    if d <= 7:  return PatternFill("solid", fgColor="FFD966")
    return RED_FILL

# ── PARSING ──────────────────────────────────────────────────────────────────

def _parse_fi_str(s):
    """Parse feature_imp string → {short_name: score}. Handles partial dicts."""
    try:
        d = ast.literal_eval(str(s).strip())
        if isinstance(d, dict):
            return d
    except Exception:
        pass
    return {k: float(v) for k, v in re.findall(r"'(\w+)'\s*:\s*([\d.eE+\-]+)", str(s))}


def _ranks_from_scores(scores_dict, all_keys):
    """
    Assign ranks 1..N to keys present in scores_dict (by descending score).
    Keys absent from scores_dict get rank N+1 (tied last).
    Returns dict {key: rank}.
    """
    present  = sorted(scores_dict, key=scores_dict.get, reverse=True)
    absent   = [k for k in all_keys if k not in scores_dict]
    ordered  = present + absent
    return {k: i+1 for i, k in enumerate(ordered)}


def _normalise_scores(scores_dict):
    """Min-max normalise a score dict to [0,1] for visual comparison only."""
    vals = list(scores_dict.values())
    lo, hi = min(vals), max(vals)
    if hi == lo:
        return {k: 0.5 for k in scores_dict}
    return {k: (v - lo) / (hi - lo) for k, v in scores_dict.items()}


def load_old(path):
    df = pd.read_csv(path)

    # Keep only HMM=3, CNN=4, RNN=5 — drop any other model IDs
    valid_models = [3, 4, 5]
    before = len(df)
    df = df[df["Model"].isin(valid_models)].reset_index(drop=True)
    dropped = before - len(df)
    if dropped:
        print(f"  Dropped {dropped} rows with Model not in {valid_models} "
              f"(kept {len(df)} rows with Model ∈ {{3=HMM, 4=CNN, 5=RNN}})")

    records = []
    for _, row in df.iterrows():
        model_id = int(row["Model"])
        ff_id    = int(row["FeatureFilter"])
        fi_raw   = _parse_fi_str(row.get("feature_imp", "{}"))

        # translate short → full names
        fi_full  = {SHORT_TO_FULL[k]: v for k, v in fi_raw.items() if k in SHORT_TO_FULL}
        reported = set(fi_full.keys())

        ranks    = _ranks_from_scores(fi_full, ALL_FEATS)
        normed   = _normalise_scores(fi_full) if fi_full else {}

        records.append({
            "task":      int(row["Task"]),
            "paradigm":  int(row["Paradigm"]),
            "model":     OLD_MODEL_MAP.get(model_id, str(model_id)),
            "ff":        OLD_FF_MAP.get(ff_id, str(ff_id)),
            "ba":        float(row.get("ba", 0)),
            "fi_raw":    fi_full,        # original scores (partial)
            "fi_norm":   normed,         # min-max normalised (partial)
            "ranks":     ranks,          # 1-18, absent feats tied-last
            "reported":  reported,       # set of full names actually given
            "n_rep":     len(reported),
        })
    return records


def load_new(base_dir):
    files = glob.glob(os.path.join(base_dir, "task*", "paradigm*", "*", "summary.json"),
                      recursive=True)
    if not files:
        files = glob.glob(os.path.join(base_dir, "**", "summary.json"), recursive=True)

    records = []
    for fpath in sorted(files):
        try:
            d = json.load(open(fpath))
        except Exception as e:
            print(f"  SKIP {fpath}: {e}")
            continue

        cfg  = d.get("config", {})
        res  = d.get("results", {})
        fi   = res.get("feature_importance", {})

        fi_full  = {k: fi.get(k, 0.0) for k in ALL_FEATS}
        ranks    = _ranks_from_scores(fi_full, ALL_FEATS)
        normed   = _normalise_scores(fi_full) if fi_full else {}

        records.append({
            "task":     int(cfg.get("task", res.get("task", 0))),
            "paradigm": int(cfg.get("paradigm", res.get("paradigm", 0))),
            "model":    cfg.get("model", res.get("model", "")).upper(),
            "method":   cfg.get("method", res.get("preprocessing_method", "")),
            "ba":       (res.get("metrics") or {}).get("ba"),
            "fi_raw":   fi_full,
            "fi_norm":  normed,
            "ranks":    ranks,
            "reported": set(ALL_FEATS),   # new always has all 18
            "n_rep":    18,
        })
    return records


def match(old_records, new_records):
    new_idx = {}
    for n in new_records:
        new_idx.setdefault((n["task"], n["paradigm"], n["model"]), []).append(n)

    pairs, unmatched = [], 0
    for o in old_records:
        key  = (o["task"], o["paradigm"], o["model"])
        news = new_idx.get(key, [])
        if not news:
            unmatched += 1
            continue
        for n in news:
            rep = o["reported"]          # features old actually reported

            # ── Spearman over reported features only ─────────────────────
            if len(rep) >= 3:
                o_r = [o["ranks"][f] for f in ALL_FEATS if f in rep]
                n_r = [n["ranks"][f] for f in ALL_FEATS if f in rep]
                rho, pval = spearmanr(o_r, n_r)
                rho  = None if np.isnan(rho)  else round(rho,  4)
                pval = None if np.isnan(pval) else round(pval, 4)
            else:
                rho = pval = None

            # ── Rank deltas per feature ───────────────────────────────────
            deltas = {f: (o["ranks"][f] - n["ranks"][f]) for f in ALL_FEATS}

            # ── Top-K overlap (restricted to reported) ────────────────────
            def topk_overlap(K):
                o_top = {f for f in rep if o["ranks"][f] <= K}
                n_top = {f for f in rep if n["ranks"][f] <= K}
                return o_top & n_top, o_top, n_top

            t3_hit, t3_o, t3_n = topk_overlap(3)
            t5_hit, t5_o, t5_n = topk_overlap(5)

            pairs.append({
                "task":      o["task"],
                "paradigm":  o["paradigm"],
                "model":     o["model"],
                "old_ff":    o["ff"],
                "new_method":n["method"],
                "old_ba":    o["ba"],
                "new_ba":    n["ba"],
                "n_rep":     o["n_rep"],
                "reported":  rep,
                "rho":       rho,
                "pval":      pval,
                "deltas":    deltas,          # {feat: signed rank delta}
                "old_ranks": o["ranks"],
                "new_ranks": n["ranks"],
                "old_fi_norm": o["fi_norm"],
                "new_fi_norm": n["fi_norm"],
                "old_fi_raw":  o["fi_raw"],
                "new_fi_raw":  n["fi_raw"],
                "t3_agree": sorted(t3_hit,   key=lambda f: o["ranks"][f]),
                "t3_only_old": sorted(t3_o - t3_hit, key=lambda f: o["ranks"][f]),
                "t3_only_new": sorted(t3_n - t3_hit, key=lambda f: n["ranks"][f]),
                "t5_agree": sorted(t5_hit,   key=lambda f: o["ranks"][f]),
                "t5_only_old": sorted(t5_o - t5_hit, key=lambda f: o["ranks"][f]),
                "t5_only_new": sorted(t5_n - t5_hit, key=lambda f: n["ranks"][f]),
            })

    if unmatched:
        print(f"  ⚠  {unmatched} old records had no match in new experiments")
    return pairs


# ── SHEET 1 : Raw Scores Side-by-Side ────────────────────────────────────────

def sheet_scores(wb, pairs):
    ws = wb.create_sheet("Scores Side-by-Side")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:AO1")
    c = ws["A1"]
    c.value = "Normalised Sensor Scores — Old (gold) vs New (blue)  |  Both min-max scaled to [0,1] for visual comparison  |  Note: raw scales differ (old ~1–4, new ~0.01–0.11)"
    c.font      = Font(bold=True, name="Arial", size=11, color="1F3864")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # Row 2: fixed info headers, then OLD group header, then NEW group header
    n_info = 5   # Task | Paradigm | Model | Old FF | New Method
    info_hdrs = ["Task","Paradigm","Model","Old FF","New Method"]

    # row 2 — OLD / NEW spanning headers
    ws.merge_cells(start_row=2, start_column=n_info+1,
                   end_row=2,   end_column=n_info+18)
    c = ws.cell(row=2, column=n_info+1, value="OLD (normalised, partial — grey = not reported)")
    _hdr(c, fill=OLD_HDR_FILL, size=10)

    ws.merge_cells(start_row=2, start_column=n_info+19,
                   end_row=2,   end_column=n_info+36)
    c = ws.cell(row=2, column=n_info+19, value="NEW (normalised, all 18 features)")
    _hdr(c, fill=NEW_HDR_FILL, size=10)

    ws.row_dimensions[2].height = 22

    # row 3 — individual feature names
    for col, h in enumerate(info_hdrs, 1):
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        _hdr(ws.cell(row=2, column=col, value=h))

    for i, feat in enumerate(ALL_FEATS):
        short = FULL_TO_SHORT.get(feat, feat)
        group = GROUPS.get(feat, "")
        c_old = ws.cell(row=3, column=n_info+1+i,    value=short)
        c_new = ws.cell(row=3, column=n_info+19+i,   value=short)
        _hdr(c_old, fill=OLD_HDR_FILL, size=8)
        _hdr(c_new, fill=NEW_HDR_FILL, size=8)

    ws.row_dimensions[3].height = 36

    for r_idx, p in enumerate(
        sorted(pairs, key=lambda x: (x["task"], x["paradigm"], x["model"])), 1
    ):
        row = r_idx + 3
        info_vals = [p["task"], p["paradigm"], p["model"], p["old_ff"], p["new_method"]]
        for col, val in enumerate(info_vals, 1):
            _dat(ws.cell(row=row, column=col), val=val)

        for i, feat in enumerate(ALL_FEATS):
            # OLD score
            c_old = ws.cell(row=row, column=n_info+1+i)
            if feat in p["reported"]:
                v = p["old_fi_norm"].get(feat, 0)
                _dat(c_old, val=round(v, 4), fill=OLD_DATA_FILL, fmt="0.0000")
            else:
                _dat(c_old, val="—", fill=GREY_FILL, color="888888")

            # NEW score
            c_new = ws.cell(row=row, column=n_info+19+i)
            v = p["new_fi_norm"].get(feat, 0)
            _dat(c_new, val=round(v, 4), fill=NEW_DATA_FILL, fmt="0.0000")

        ws.row_dimensions[row].height = 15

    # column widths
    for i, w in enumerate([6,9,7,12,12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(n_info+1, n_info+37):
        ws.column_dimensions[get_column_letter(i)].width = 7.5

    # color scale on old and new score blocks separately
    last = len(pairs) + 3
    ws.conditional_formatting.add(
        f"{get_column_letter(n_info+1)}4:{get_column_letter(n_info+18)}{last}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max",   end_color="F4B942"))
    ws.conditional_formatting.add(
        f"{get_column_letter(n_info+19)}4:{get_column_letter(n_info+36)}{last}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max",   end_color="2E75B6"))

    ws.freeze_panes = f"{get_column_letter(n_info+1)}4"


# ── SHEET 2 : Rank Comparison ─────────────────────────────────────────────────

def sheet_ranks(wb, pairs):
    ws = wb.create_sheet("Rank Comparison")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:BH1")
    ws["A1"].value = "Rank Comparison (1=most important, 18=least)  |  Grey/italic = not reported by old CSV  |  Δ = Old rank − New rank (green≈0, red=large gap)"
    ws["A1"].font      = Font(bold=True, name="Arial", size=11, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    n_info = 6   # Task|Paradigm|Model|OldFF|NewMethod|ρ(reported)
    info_hdrs = ["Task","Paradigm","Model","Old FF","New Method","Spearman ρ\n(reported)"]

    # row 2: info merge + feature group spans (3 sub-cols each: Old|New|Δ)
    for col, h in enumerate(info_hdrs, 1):
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        _hdr(ws.cell(row=2, column=col, value=h))

    for i, feat in enumerate(ALL_FEATS):
        start = n_info + 1 + i*3
        ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=start+2)
        gf = GROUP_FILLS.get(GROUPS.get(feat,""), NEUTRAL_FILL)
        short = FULL_TO_SHORT.get(feat, feat)
        c = ws.cell(row=2, column=start, value=short)
        c.font      = Font(bold=True, name="Arial", size=8)
        c.fill      = gf
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = THIN_B
        for sub, lbl in enumerate(["Old","New","Δ"]):
            sc = ws.cell(row=3, column=start+sub, value=lbl)
            sc.font      = Font(bold=True, name="Arial", size=8,
                                color="7F6000" if lbl=="Old" else "1F4E79" if lbl=="New" else "000000")
            sc.fill      = OLD_DATA_FILL if lbl=="Old" else NEW_DATA_FILL if lbl=="New" else NEUTRAL_FILL
            sc.alignment = Alignment(horizontal="center", vertical="center")
            sc.border    = THIN_B

    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 18

    for r_idx, p in enumerate(
        sorted(pairs, key=lambda x: (x["task"], x["paradigm"], x["model"])), 1
    ):
        row = r_idx + 3
        rho = p["rho"]
        info_vals = [p["task"], p["paradigm"], p["model"], p["old_ff"], p["new_method"], rho]
        for col, val in enumerate(info_vals, 1):
            c = ws.cell(row=row, column=col)
            if col == 6:
                _dat(c, val=val, bold=True, fill=_rho_fill(rho), fmt="0.00")
            else:
                _dat(c, val=val)
            if isinstance(val, float) and col != 6:
                c.number_format = "0.000"

        for i, feat in enumerate(ALL_FEATS):
            start   = n_info + 1 + i*3
            o_rank  = p["old_ranks"][feat]
            n_rank  = p["new_ranks"][feat]
            delta   = p["deltas"][feat]
            reported = feat in p["reported"]

            c_old = ws.cell(row=row, column=start)
            c_new = ws.cell(row=row, column=start+1)
            c_dlt = ws.cell(row=row, column=start+2)

            if reported:
                _dat(c_old, val=o_rank, fill=OLD_DATA_FILL)
                _dat(c_new, val=n_rank, fill=NEW_DATA_FILL)
                _dat(c_dlt, val=delta,  fill=_delta_fill(delta), bold=(abs(delta)>4))
            else:
                # greyed — old didn't report this feature
                for c in [c_old, c_new, c_dlt]:
                    _dat(c, val="·", fill=GREY_FILL, color="AAAAAA")

        ws.row_dimensions[row].height = 15

    info_widths = [6, 9, 7, 12, 12, 10]
    for i, w in enumerate(info_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(n_info+1, n_info+1+18*3):
        ws.column_dimensions[get_column_letter(i)].width = 3.8

    ws.freeze_panes = f"{get_column_letter(n_info+1)}4"


# ── SHEET 3 : Top-K Sensor Agreement ─────────────────────────────────────────

def sheet_topk(wb, pairs):
    ws = wb.create_sheet("Top-K Agreement")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:R1")
    ws["A1"].value = "Top-K Sensor Agreement — which sensors old & new BOTH put in their top-K (restricted to reported features)"
    ws["A1"].font      = Font(bold=True, name="Arial", size=11, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 22

    hdrs = ["Task","Paradigm","Model","Old FF","# Rep",
            "Top-3 Agree","Top-3 Old only","Top-3 New only","Top-3 %",
            "Top-5 Agree","Top-5 Old only","Top-5 New only","Top-5 %",
            "Spearman ρ"]
    for col, h in enumerate(hdrs, 1):
        _hdr(ws.cell(row=2, column=col, value=h), size=9)
    ws.row_dimensions[2].height = 30

    def _feat_list(feats):
        return ", ".join(FULL_TO_SHORT.get(f, f) for f in feats) if feats else "—"

    for r_idx, p in enumerate(
        sorted(pairs, key=lambda x: (x["task"], x["paradigm"], x["model"])), 1
    ):
        row = r_idx + 2
        n   = max(p["n_rep"], 1)
        t3p = len(p["t3_agree"]) / min(3, n)
        t5p = len(p["t5_agree"]) / min(5, n)

        vals = [
            p["task"], p["paradigm"], p["model"], p["old_ff"], p["n_rep"],
            _feat_list(p["t3_agree"]), _feat_list(p["t3_only_old"]),
            _feat_list(p["t3_only_new"]), t3p,
            _feat_list(p["t5_agree"]), _feat_list(p["t5_only_old"]),
            _feat_list(p["t5_only_new"]), t5p,
            p["rho"],
        ]
        pct_cols  = {9, 13}
        align_pct = {9: t3p, 13: t5p}
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col)
            if col in pct_cols:
                pct = align_pct[col]
                fill = GREEN_FILL if pct>=0.67 else AMBER_FILL if pct>=0.40 else RED_FILL
                _dat(c, val=val, fill=fill, bold=True, fmt="0%")
            elif col == 14:   # rho
                _dat(c, val=val, fill=_rho_fill(val), bold=True, fmt="0.00")
            elif col in {6,7,8,10,11,12}:
                align = "left"
                fill  = GREEN_FILL if col in {6,10} and val != "—" else \
                        OLD_DATA_FILL if col in {7,11} else \
                        NEW_DATA_FILL if col in {8,12} else None
                _dat(c, val=val, align=align, fill=fill, size=8)
            elif col == 5 and val < 18:
                _dat(c, val=val, fill=AMBER_FILL)
            else:
                _dat(c, val=val)
        ws.row_dimensions[row].height = 15

    widths = [6,9,7,12,6, 28,24,24,8, 32,28,28,8, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


# ── SHEET 4 : Alignment Heatmap ──────────────────────────────────────────────

def sheet_heatmap(wb, pairs):
    """
    Rows = sensors (18), Columns = experiments.
    Cell value = |rank delta| (0 = perfectly aligned, 17 = completely opposite).
    Green = aligned, Red = misaligned. Grey = not reported by old.
    """
    ws = wb.create_sheet("Alignment Heatmap")
    ws.sheet_view.showGridLines = False

    sorted_pairs = sorted(pairs, key=lambda x: (x["task"], x["paradigm"], x["model"]))
    n_exp = len(sorted_pairs)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4+n_exp)
    ws["A1"].value = (
        "Alignment Heatmap  |  Rows = sensors  |  Cols = experiments  |  "
        "Cell = |Old rank − New rank|  |  0 (green) = perfectly aligned  |  "
        "Grey = not reported by old CSV"
    )
    ws["A1"].font      = Font(bold=True, name="Arial", size=11, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 22

    # Row 2: experiment labels
    for col, p in enumerate(sorted_pairs, 5):
        lbl = f"T{p['task']}\nP{p['paradigm']}\n{p['model']}"
        c   = ws.cell(row=2, column=col, value=lbl)
        _hdr(c, fill=HDR_FILL, size=7)
    ws.row_dimensions[2].height = 42

    # Column 1-4: Sensor | Short | Group | Avg|Δ|
    for col, h in enumerate(["Sensor (full)", "Short", "Group", "Avg |Δ|"], 1):
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        _hdr(ws.cell(row=2, column=col, value=h))

    # Row 3: experiment sub-label (ρ value)
    for col, p in enumerate(sorted_pairs, 5):
        rho_lbl = f"ρ={p['rho']:.2f}" if p["rho"] is not None else "ρ=N/A"
        c = ws.cell(row=3, column=col, value=rho_lbl)
        _hdr(c, fill=_rho_fill(p["rho"]), size=7)
    ws.row_dimensions[3].height = 16

    # Data rows: one per sensor
    for feat_row, feat in enumerate(ALL_FEATS):
        row   = feat_row + 4
        short = FULL_TO_SHORT.get(feat, feat)
        group = GROUPS.get(feat, "")
        gfill = GROUP_FILLS.get(group, NEUTRAL_FILL)

        abs_deltas = []
        for col, p in enumerate(sorted_pairs, 5):
            delta     = p["deltas"][feat]
            reported  = feat in p["reported"]
            abs_delta = abs(delta)
            c = ws.cell(row=row, column=col)
            if not reported:
                _dat(c, val="", fill=GREY_FILL)
            else:
                abs_deltas.append(abs_delta)
                _dat(c, val=abs_delta, fill=_delta_fill(delta))
                c.number_format = "0"

        avg_delta = round(np.mean(abs_deltas), 2) if abs_deltas else None

        _dat(ws.cell(row=row, column=1), val=feat,  align="left", fill=gfill, size=8)
        _dat(ws.cell(row=row, column=2), val=short, align="left", fill=gfill, size=9, bold=True)
        _dat(ws.cell(row=row, column=3), val=group, align="left", fill=gfill, size=8)
        c_avg = ws.cell(row=row, column=4)
        if avg_delta is not None:
            _dat(c_avg, val=avg_delta, fill=_delta_fill(avg_delta), bold=True, fmt="0.0")
        else:
            _dat(c_avg, val="N/A", fill=GREY_FILL)

        ws.row_dimensions[row].height = 15

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 11
    ws.column_dimensions["D"].width = 7
    for i in range(5, 5+n_exp):
        ws.column_dimensions[get_column_letter(i)].width = 6

    # color scale on delta values
    last_feat_row = len(ALL_FEATS) + 3
    ws.conditional_formatting.add(
        f"E4:{get_column_letter(4+n_exp)}{last_feat_row}",
        ColorScaleRule(start_type="num",  start_value=0,  start_color="63BE7B",
                       mid_type="num",    mid_value=6,    mid_color="FFEB84",
                       end_type="num",    end_value=17,   end_color="F8696B"))
    ws.freeze_panes = "E4"


# ── SHEET 5 : Spearman per Experiment ────────────────────────────────────────

def sheet_spearman(wb, pairs):
    ws = wb.create_sheet("Spearman per Experiment")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    ws["A1"].value = "Spearman Rank Correlation (computed over reported features only)  |  ρ≥0.70 strong  |  0.40–0.70 moderate  |  <0.40 weak"
    ws["A1"].font      = Font(bold=True, name="Arial", size=11, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 22

    hdrs = ["Task","Paradigm","Model","Old FF","New Method","# Reported","Old BA","New BA","Spearman ρ","Interpretation"]
    for col, h in enumerate(hdrs, 1):
        _hdr(ws.cell(row=2, column=col, value=h), size=9)
    ws.row_dimensions[2].height = 28

    interp_map = [
        (0.70, "✓ Strong agreement"),
        (0.40, "~ Moderate agreement"),
        (-1.0, "✗ Weak / disagreement"),
    ]
    def _interp(rho):
        if rho is None: return "N/A (< 3 reported)"
        for thr, label in interp_map:
            if rho >= thr: return label
        return "✗ Weak / disagreement"

    sorted_pairs = sorted(pairs, key=lambda p: (p["rho"] or -99), reverse=True)

    for r_idx, p in enumerate(sorted_pairs, 1):
        row  = r_idx + 2
        rho  = p["rho"]
        fill = _rho_fill(rho)
        vals = [
            p["task"], p["paradigm"], p["model"],
            p["old_ff"], p["new_method"],
            p["n_rep"], p["old_ba"], p["new_ba"],
            rho, _interp(rho),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col)
            if col == 9:
                _dat(c, val=val, fill=fill, bold=True, fmt="0.0000")
            elif col == 10:
                _dat(c, val=val, fill=fill, align="left")
            elif col == 6 and isinstance(val, int) and val < 18:
                _dat(c, val=val, fill=AMBER_FILL, bold=True)
            elif col in (7,8) and isinstance(val, float):
                _dat(c, val=val, fmt="0.000")
            else:
                _dat(c, val=val)
        ws.row_dimensions[row].height = 15

    widths = [6,9,7,12,12,8,8,8,12,26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last = len(pairs) + 2
    ws.conditional_formatting.add(f"I3:I{last}",
        ColorScaleRule(start_type="num", start_value=-1, start_color="F8696B",
                       mid_type="num",   mid_value=0,    mid_color="FFEB84",
                       end_type="num",   end_value=1,    end_color="63BE7B"))
    ws.freeze_panes = "A3"


# ── SHEET 6 : Sensor-Level Summary ───────────────────────────────────────────

def sheet_sensor_summary(wb, pairs):
    ws = wb.create_sheet("Sensor-Level Summary")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:K1")
    ws["A1"].value = "Per-Sensor Summary  |  Avg Old Rank, Avg New Rank, Avg |Δ|, % experiments where both agree (top-5)"
    ws["A1"].font      = Font(bold=True, name="Arial", size=11, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 22

    hdrs = ["Sensor (full)","Short","Group",
            "Avg Old Rank","Avg New Rank","Avg |Δ| Rank",
            "Max |Δ|","% Both Top-5","% Both Top-3",
            "# Times Reported","Alignment"]
    for col, h in enumerate(hdrs, 1):
        _hdr(ws.cell(row=2, column=col, value=h), size=9)
    ws.row_dimensions[2].height = 30

    rows_data = []
    for feat in ALL_FEATS:
        reported_pairs = [p for p in pairs if feat in p["reported"]]
        n_rep          = len(reported_pairs)

        if reported_pairs:
            avg_old = np.mean([p["old_ranks"][feat] for p in reported_pairs])
            avg_new = np.mean([p["new_ranks"][feat] for p in reported_pairs])
            avg_abs = np.mean([abs(p["deltas"][feat]) for p in reported_pairs])
            max_abs = max( abs(p["deltas"][feat]) for p in reported_pairs)
        else:
            avg_old = avg_new = avg_abs = max_abs = None

        # % of pairs where both put this sensor in top-5 (among reported)
        t5_agree = sum(1 for p in reported_pairs
                       if feat in p["reported"]
                       and p["old_ranks"][feat] <= 5
                       and p["new_ranks"][feat] <= 5)
        t3_agree = sum(1 for p in reported_pairs
                       if feat in p["reported"]
                       and p["old_ranks"][feat] <= 3
                       and p["new_ranks"][feat] <= 3)
        pct_t5 = t5_agree / max(n_rep, 1)
        pct_t3 = t3_agree / max(n_rep, 1)

        short = FULL_TO_SHORT.get(feat, feat)
        group = GROUPS.get(feat, "")

        if avg_abs is None:
            alignment = "Not reported"
        elif avg_abs <= 2:  alignment = "✓ Strongly aligned"
        elif avg_abs <= 5:  alignment = "~ Moderately aligned"
        else:               alignment = "✗ Poorly aligned"

        rows_data.append({
            "feat": feat, "short": short, "group": group,
            "avg_old": avg_old, "avg_new": avg_new, "avg_abs": avg_abs,
            "max_abs": max_abs, "pct_t5": pct_t5, "pct_t3": pct_t3,
            "n_rep": n_rep, "alignment": alignment,
        })

    # Sort by avg_abs ascending (best aligned first)
    rows_data.sort(key=lambda x: (x["avg_abs"] is None, x["avg_abs"] or 99))

    for r_idx, rd in enumerate(rows_data, 1):
        row   = r_idx + 2
        gfill = GROUP_FILLS.get(rd["group"], NEUTRAL_FILL)
        avg_abs = rd["avg_abs"]
        a_fill  = GREEN_FILL if avg_abs is not None and avg_abs<=2 else \
                  AMBER_FILL if avg_abs is not None and avg_abs<=5 else RED_FILL

        vals = [rd["feat"], rd["short"], rd["group"],
                rd["avg_old"], rd["avg_new"], avg_abs,
                rd["max_abs"], rd["pct_t5"], rd["pct_t3"],
                rd["n_rep"], rd["alignment"]]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col)
            if col in (1,2,3):
                _dat(c, val=val, align="left", fill=gfill)
            elif col in (4,5):
                _dat(c, val=round(val,2) if val is not None else "N/A", fmt="0.00")
            elif col == 6:
                _dat(c, val=round(val,2) if val is not None else "N/A",
                     fill=a_fill, bold=True, fmt="0.00")
            elif col == 7:
                _dat(c, val=int(val) if val is not None else "N/A")
            elif col in (8,9):
                fill = GREEN_FILL if val>=0.5 else AMBER_FILL if val>=0.25 else RED_FILL
                _dat(c, val=val, fill=fill, fmt="0%")
            elif col == 10:
                fill = AMBER_FILL if isinstance(val,int) and val<len(pairs) else None
                _dat(c, val=val, fill=fill)
            elif col == 11:
                _dat(c, val=val, fill=a_fill, align="left")
            else:
                _dat(c, val=val)
        ws.row_dimensions[row].height = 16

    widths = [24,9,12,11,11,10,8,10,10,12,22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last = len(rows_data) + 2
    ws.conditional_formatting.add(f"F3:F{last}",
        ColorScaleRule(start_type="num", start_value=0,  start_color="63BE7B",
                       mid_type="num",   mid_value=6,    mid_color="FFEB84",
                       end_type="num",   end_value=14,   end_color="F8696B"))
    ws.freeze_panes = "A3"


# ── SHEET 7 : Experiment Dashboard ───────────────────────────────────────────

def sheet_dashboard(wb, pairs, old_records, new_records):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 32

    ws.merge_cells("A1:C1")
    ws["A1"].value = "Sensor Alignment Dashboard"
    ws["A1"].font      = Font(bold=True, name="Arial", size=15, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 32

    def sec(row, title):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row=row, column=1, value=title)
        c.font      = Font(bold=True, name="Arial", size=11, color="FFFFFF")
        c.fill      = SUBHDR_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22

    def kv(row, key, val, val_fill=None, note=None):
        ck = ws.cell(row=row, column=2, value=key)
        cv = ws.cell(row=row, column=3, value=val)
        ck.font = Font(bold=True, name="Arial", size=10)
        cv.font = Font(name="Arial", size=10)
        ck.alignment = Alignment(horizontal="left", vertical="center")
        cv.alignment = Alignment(horizontal="left", vertical="center")
        if val_fill: cv.fill = val_fill
        if note:
            cn = ws.cell(row=row, column=4, value=note)
            cn.font = Font(italic=True, name="Arial", size=9, color="595959")
        ws.row_dimensions[row].height = 18

    rhos   = [p["rho"] for p in pairs if p["rho"] is not None]
    n_reps = [p["n_rep"] for p in pairs]
    all_abs_deltas = [abs(d) for p in pairs for f,d in p["deltas"].items() if f in p["reported"]]

    r = 3
    sec(r, "1. Overview"); r+=1
    kv(r,"Old CSV experiments:", len(old_records)); r+=1
    kv(r,"New JSON experiments:", len(new_records)); r+=1
    kv(r,"Matched pairs (exact Task × Paradigm × Model):", len(pairs)); r+=1
    kv(r,"Avg features reported per old experiment:",
       f"{np.mean(n_reps):.1f} / 18",
       val_fill=AMBER_FILL if np.mean(n_reps)<18 else GREEN_FILL); r+=1
    kv(r,"Note: old CSV is partial (top-k only); grey cells = not reported",""); r+=2

    sec(r, "2. Rank Correlation (Spearman ρ, over reported features only)"); r+=1
    if rhos:
        kv(r,"Mean ρ:",   f"{np.mean(rhos):.3f}",
           val_fill=_rho_fill(np.mean(rhos))); r+=1
        kv(r,"Median ρ:", f"{np.median(rhos):.3f}"); r+=1
        kv(r,"Min / Max ρ:", f"{np.min(rhos):.3f}  /  {np.max(rhos):.3f}"); r+=1
        pct_strong = sum(x>=0.7 for x in rhos)/len(rhos)
        pct_mod    = sum(x>=0.4 for x in rhos)/len(rhos)
        kv(r,"% experiments with ρ ≥ 0.70 (strong):",
           f"{pct_strong*100:.1f}%",
           val_fill=GREEN_FILL if pct_strong>0.5 else RED_FILL); r+=1
        kv(r,"% experiments with ρ ≥ 0.40 (moderate+):",
           f"{pct_mod*100:.1f}%"); r+=1
    r+=1

    sec(r, "3. Raw Rank Delta Statistics"); r+=1
    kv(r,"Avg |rank delta| across all (sensor, experiment) pairs:",
       f"{np.mean(all_abs_deltas):.2f}  (scale: 0 = perfect, 17 = worst)"); r+=1
    kv(r,"Median |rank delta|:", f"{np.median(all_abs_deltas):.2f}"); r+=1
    kv(r,"% (sensor, experiment) pairs with |Δ| ≤ 2 (closely aligned):",
       f"{100*sum(x<=2 for x in all_abs_deltas)/len(all_abs_deltas):.1f}%",
       val_fill=GREEN_FILL); r+=1
    kv(r,"% (sensor, experiment) pairs with |Δ| > 7 (strongly misaligned):",
       f"{100*sum(x>7 for x in all_abs_deltas)/len(all_abs_deltas):.1f}%",
       val_fill=RED_FILL); r+=1
    r+=1

    sec(r, "4. Most Consistently Aligned Sensors (low avg |Δ|)"); r+=1
    sensor_stats = {}
    for feat in ALL_FEATS:
        rep_pairs = [p for p in pairs if feat in p["reported"]]
        if rep_pairs:
            sensor_stats[feat] = np.mean([abs(p["deltas"][feat]) for p in rep_pairs])
    for feat, avg in sorted(sensor_stats.items(), key=lambda x: x[1])[:6]:
        short = FULL_TO_SHORT.get(feat, feat)
        fill  = GREEN_FILL if avg<=2 else AMBER_FILL if avg<=5 else RED_FILL
        kv(r, f"  {short}  ({feat}):", f"avg |Δ| = {avg:.2f}", val_fill=fill); r+=1
    r+=1

    sec(r, "5. Most Misaligned Sensors (high avg |Δ|)"); r+=1
    for feat, avg in sorted(sensor_stats.items(), key=lambda x: x[1], reverse=True)[:6]:
        short = FULL_TO_SHORT.get(feat, feat)
        fill  = GREEN_FILL if avg<=2 else AMBER_FILL if avg<=5 else RED_FILL
        kv(r, f"  {short}  ({feat}):", f"avg |Δ| = {avg:.2f}", val_fill=fill); r+=1
    r+=1

    sec(r, "6. Interpretation Guide"); r+=1
    notes = [
        ("Spearman ρ ≥ 0.70", "Strong agreement — both systems rank sensors similarly"),
        ("Spearman ρ 0.40–0.70", "Moderate agreement — top sensors align, ordering differs"),
        ("Spearman ρ < 0.40", "Weak agreement — sensor rankings are inconsistent"),
        ("|Δ| = 0", "Perfectly aligned rank for that sensor in that experiment"),
        ("|Δ| ≤ 2", "Closely aligned (within 2 ranks)"),
        ("|Δ| > 7", "Strongly misaligned (very different importance assigned)"),
        ("Grey cells", "Sensor not reported by old CSV — excluded from ρ calculation"),
    ]
    for key, val in notes:
        kv(r, key, val); r+=1

    ws.column_dimensions["D"].width = 50


# ── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print("Loading old CSV …")
    old_recs = load_old(OLD_CSV)
    print(f"  {len(old_recs)} records  |  avg features reported: "
          f"{np.mean([r['n_rep'] for r in old_recs]):.1f}/18")

    print("Loading new JSON experiments …")
    new_recs = load_new(EXPERIMENTS_DIR)
    print(f"  {len(new_recs)} records")

    print("Matching (Task × Paradigm × Model) …")
    pairs = match(old_recs, new_recs)
    print(f"  {len(pairs)} matched pairs")

    wb = Workbook()
    wb.remove(wb.active)

    sheet_scores(wb, pairs)
    sheet_ranks(wb, pairs)
    sheet_topk(wb, pairs)
    sheet_heatmap(wb, pairs)
    sheet_spearman(wb, pairs)
    sheet_sensor_summary(wb, pairs)
    sheet_dashboard(wb, pairs, old_recs, new_recs)

    wb.save(OUTPUT_FILE)
    print(f"\nSaved → {OUTPUT_FILE}")

    # console summary
    rhos = [p["rho"] for p in pairs if p["rho"] is not None]
    if rhos:
        all_d = [abs(d) for p in pairs for f,d in p["deltas"].items() if f in p["reported"]]
        print(f"\nAlignment summary ({len(pairs)} pairs, {len(rhos)} with ρ):")
        print(f"  Mean Spearman ρ : {np.mean(rhos):.3f}")
        print(f"  Mean |Δ rank|   : {np.mean(all_d):.2f}")
        print(f"  % |Δ| ≤ 2       : {100*sum(x<=2 for x in all_d)/len(all_d):.1f}%")
        print(f"  Strong (ρ≥0.70) : {sum(x>=0.7 for x in rhos)}/{len(rhos)}")


if __name__ == "__main__":
    main()