"""
add_survey_columns.py
=====================
Reads SurveyResponse.csv for each subject and adds 11 individual question
columns (q1_jar … q11_sleep) plus DASH_survey formula to Sheet1 of
xdash_px_details.xlsx.

CSV format (Timestamp, Question, Response):
    182.68, Q1, NoDifficulty
    223.03, Q2, MildDifficulty
    ...

Response → numeric:
    NoDifficulty / None             → 1
    MildDifficulty / Mild           → 2
    ModerateDifficulty / Moderate   → 3
    SevereDifficulty / Severe       → 4
    Unable / Extreme / CantSleep    → 5

DASH_survey = ((mean of Q1-Q11) - 1) × 25

Usage:
    python add_survey_columns.py \
        --xlsx data/xdash_px_details.xlsx \
        --data data/xdash_data \
        --out  data/xdash_px_details_updated.xlsx
"""
import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Constants ─────────────────────────────────────────────────────────────────

Q_COLS = [
    "q1_jar", "q2_key", "q3_household", "q4_back_wash", "q5_knife",
    "q6_recreational", "q7_social", "q8_work",
    "q9_pain", "q10_tingling", "q11_sleep",
]

# Normalised text → 1-5  (keys: lowercase, no spaces/underscores/apostrophes)
RESPONSE_MAP = {
    "nodifficulty": 1, "none": 1, "no": 1, "1": 1,
    "milddifficulty": 2, "mild": 2, "2": 2,
    "moderatedifficulty": 3, "moderate": 3, "3": 3,
    "severedifficulty": 4, "severe": 4, "4": 4,
    "unable": 5, "extreme": 5, "5": 5,
    "cantsleep": 5, "cantmove": 5, "somuchdifficulty": 5,
}


# ── Response text → int ───────────────────────────────────────────────────────

def parse_response(text: str):
    # Already numeric?
    try:
        v = int(float(text.strip()))
        if 1 <= v <= 5:
            return v
    except (ValueError, AttributeError):
        pass

    # Normalise and look up
    key = (text.strip().lower()
           .replace(" ", "").replace("_", "")
           .replace("'", "").replace("\u2019", "").replace("\u2018", ""))

    if key in RESPONSE_MAP:
        return RESPONSE_MAP[key]

    # Prefix match for long strings like "somuchdifficultythaticantmove"
    for k, v in RESPONSE_MAP.items():
        if len(k) >= 6 and key.startswith(k):
            return v

    return None


# ── CSV loader ────────────────────────────────────────────────────────────────

def load_survey(csv_path: Path):
    """Return list of 11 values (int or None) for Q1-Q11, or None on failure."""
    if not csv_path.exists():
        return None

    responses = {}  # "Q1".."Q11" → int

    try:
        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            all_rows = [r for r in reader if any(c.strip() for c in r)]

        if not all_rows:
            return None

        # Detect header
        first = [c.strip().lower() for c in all_rows[0]]
        has_header = any(h in ("question", "response", "timestamp") for h in first)
        data_rows = all_rows[1:] if has_header else all_rows

        # Column indices: default Timestamp=0, Question=1, Response=2
        q_idx, r_idx = 1, 2
        if has_header:
            for i, h in enumerate(first):
                if h == "question":
                    q_idx = i
                elif h == "response":
                    r_idx = i

        for row in data_rows:
            if len(row) <= max(q_idx, r_idx):
                continue
            q_label = row[q_idx].strip().upper()
            r_text  = row[r_idx].strip()
            numeric = parse_response(r_text)
            if q_label and numeric is not None:
                responses[q_label] = numeric

    except Exception as e:
        print(f"  [ERROR] {csv_path}: {e}")
        return None

    if not responses:
        print(f"  [WARN] No valid responses in {csv_path}")
        return None

    result = []
    for i in range(1, 12):
        val = responses.get(f"Q{i}")
        if val is None:
            print(f"  [WARN] {csv_path.parent.name}: Q{i} not found")
        result.append(val)

    return result


# ── Path resolver ─────────────────────────────────────────────────────────────

def find_survey(data_dir: Path, sid: str):
    """Try all case combinations to locate SurveyResponse.csv."""
    prefix = "px" if sid.upper().startswith("PX") else "fx"
    for p in (prefix, prefix.upper()):
        for s in (sid, sid.upper(), sid.lower()):
            path = data_dir / p / s / "SurveyResponse.csv"
            if path.exists():
                return path
    # Return primary candidate for error message
    return data_dir / prefix / sid / "SurveyResponse.csv"


# ── Styles ────────────────────────────────────────────────────────────────────

def hdr(cell, label):
    cell.value     = label
    cell.font      = Font(bold=True, name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color="D9E1F2")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = Border(bottom=Side(style="thin"), right=Side(style="thin"))


def note(cell, text):
    cell.value     = text
    cell.font      = Font(italic=True, name="Arial", size=9, color="595959")
    cell.fill      = PatternFill("solid", start_color="FFF2CC")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def data(cell, value):
    cell.value     = value
    cell.font      = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="data/xdash_data/xdash_px_details.xlsx")
    ap.add_argument("--data", default="data/xdash_data")
    ap.add_argument("--out",  default=None)
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    data_dir  = Path(args.data)
    out_path  = Path(args.out) if args.out else xlsx_path

    print(f"Loading: {xlsx_path}")
    wb = load_workbook(xlsx_path)
    ws = wb["Sheet1"]

    # ── Find start column ────────────────────────────────────────────────────
    existing = {cell.value: cell.column for cell in ws[2] if cell.value}

    if "q1_jar" in existing:
        q_start = existing["q1_jar"]
        print(f"Survey columns exist at {get_column_letter(q_start)} — refreshing values.")
    else:
        q_start = ws.max_column + 1
        print(f"Inserting survey columns at {get_column_letter(q_start)}")

        # Row 1 — note in first new col only
        note(ws.cell(row=1, column=q_start),
             "Survey: 1=No difficulty  2=Mild  3=Moderate  4=Severe  5=Unable/Extreme")
        ws.row_dimensions[1].height = 28

        # Row 2 — q1..q11 headers
        for i, name in enumerate(Q_COLS):
            hdr(ws.cell(row=2, column=q_start + i), name)
            ws.column_dimensions[get_column_letter(q_start + i)].width = 13

        # DASH_survey header
        dash_col = q_start + len(Q_COLS)
        hdr(ws.cell(row=2, column=dash_col), "DASH_survey")
        ws.column_dimensions[get_column_letter(dash_col)].width = 14

    dash_col = q_start + len(Q_COLS)

    # ── Subject rows ─────────────────────────────────────────────────────────
    subjects = {
        str(ws.cell(row=r, column=1).value).strip(): r
        for r in range(3, ws.max_row + 1)
        if ws.cell(row=r, column=1).value
    }
    print(f"Subjects in sheet: {len(subjects)}")

    n_ok = n_miss = 0

    for sid, excel_row in sorted(subjects.items()):
        csv_path  = find_survey(data_dir, sid)
        responses = load_survey(csv_path)

        if responses is None:
            print(f"  [MISSING] {sid:6s}  →  {csv_path}")
            n_miss += 1
            for i in range(len(Q_COLS)):
                data(ws.cell(row=excel_row, column=q_start + i), None)
            ws.cell(row=excel_row, column=dash_col).value = None
            continue

        n_ok += 1
        for i, val in enumerate(responses):
            data(ws.cell(row=excel_row, column=q_start + i), val)

        # DASH formula
        c1 = get_column_letter(q_start)
        c2 = get_column_letter(q_start + len(Q_COLS) - 1)
        dc = ws.cell(row=excel_row, column=dash_col)
        dc.value     = f"=((AVERAGE({c1}{excel_row}:{c2}{excel_row}))-1)*25"
        dc.font      = Font(name="Arial", size=10)
        dc.alignment = Alignment(horizontal="center")

    wb.save(out_path)

    print(f"\nSaved  →  {out_path}")
    print(f"  Found   : {n_ok}  subjects with survey data")
    print(f"  Missing : {n_miss} subjects")
    col_range = f"{get_column_letter(q_start)}–{get_column_letter(dash_col)}"
    print(f"\nColumns written ({col_range}):")
    for i, name in enumerate(Q_COLS):
        print(f"  {get_column_letter(q_start+i)}: {name}")
    print(f"  {get_column_letter(dash_col)}: DASH_survey  [formula]")


if __name__ == "__main__":
    main()