"""
hmm_laterality_analysis.py
===========================
Post-hoc laterality analysis for HMM results across all 24 task×paradigm experiments.

Three analyses:
  1. Laterality–Feature Importance Alignment
     For each unilateral patient (n=34), test whether the HMM assigns higher
     importance to ipsilateral vs contralateral hand channels.
     Output: per-task summary + subject-level table.

  2. Dominant-Hand vs Injured-Side Cross-Tabulation
     Cross-tab dia_side x hand_xr (DOMINANT hand) for patients.
     NOTE: hand_xr = dominant hand, not the hand used during the XR task.
     Inconsistently recorded -- missing for some patients (shown as Unknown).
     For each cell, report mean HMM probability and classification accuracy.
     Output: cross-tab heatmap table across all tasks.

  3. HMM Probability vs Self-Reported Difficulty (Spearman correlation)
     Correlate y_proba with q* task scores per task, across all 24
     task x paradigm cells (all-subjects and within-patient families).
     p-values are Benjamini-Hochberg FDR-corrected separately within each
     family; significance assessed at FDR = 0.05 (q < 0.05). Uncorrected
     p-values are reported alongside the corrected q-values.
     Also examines misclassified patients' q* scores.
     Output: correlation table + misclassification insight table.

Usage:
    python hmm_laterality_analysis.py
    python hmm_laterality_analysis.py --dataset xdash --px-details data/xdash_px_details.xlsx
    python hmm_laterality_analysis.py --paradigm 1   # only paradigm 1
    python hmm_laterality_analysis.py --task 1 2 3   # only tasks 1-3

Output:
    storage/results/<dataset>/hmm/laterality/HMM_Laterality_Analysis.xlsx  (5-sheet workbook)
"""

import argparse
import glob
import json
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from scipy import stats
from statsmodels.stats.multitest import multipletests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

# ── Constants ─────────────────────────────────────────────────────────────────

TASKS      = list(range(1, 7))
PARADIGMS  = list(range(1, 5))

TASK_LABELS = {
    1: "Jar Opening",  2: "Key Turning",   3: "Cleaning",
    4: "Back Washing", 5: "Cutting",        6: "Hammering",
}
PARADIGM_LABELS = {
    1: "Patients vs Controls", 2: "RCT vs Controls",
    3: "Other vs Controls",    4: "RCT vs Other",
}
PARADIGM_G1 = {1: "Patients", 2: "RCT", 3: "Other", 4: "RCT"}
PARADIGM_G0 = {1: "Controls", 2: "Controls", 3: "Controls", 4: "Other"}

TASK_Q_COL = {
    1: "q1_jar",     2: "q2_key",     3: "q3_household",
    4: "q4_back_wash", 5: "q5_knife", 6: "q6_recreational",
}

# Channel-group index ranges — these assume the standard xdash sensor layout
# (head, left hand, right hand, 6 DoF each). channel_names themselves are
# loaded per-dataset in main() via load_dataset_config(args.dataset)["channels"].
HEAD_IDX  = list(range(0, 6))
LEFT_IDX  = list(range(6, 12))
RIGHT_IDX = list(range(12, 18))

DIA_NAMES = {
    0: "Healthy/Control", 1: "Rotator Cuff Tear",
    2: "Glenohumeral Arthritis", 3: "Biceps Tendonitis", 4: "Bursitis",
}

# Colour palette
C = {
    "navy":    "1F4E79", "blue":  "2E75B6", "teal":  "1F6B6B",
    "purple":  "4A3060", "green": "375623", "amber": "7B5C1E",
    "red":     "7B2C2C", "gray":  "595959",
    "r_light": "FCE4D6", "l_light": "E2EFDA",
    "b_light": "DDEBF7", "h_light": "FFF2CC",
    "miss":    "FFD7D7", "hit":   "E2EFDA",
    "bil":     "F2F2F2", "ctrl":  "EBF3FB",
    "white":   "FFFFFF", "offwhite": "F9F9F9",
}

# ── Style helpers ─────────────────────────────────────────────────────────────

def _bdr(color="BBBBBB"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_bdr():
    t = Side(style="medium", color="1F4E79")
    n = Side(style="thin",   color="BBBBBB")
    return Border(left=t, right=t, top=t, bottom=t)

def hdr(ws, r, c, val, bg, fc="FFFFFF", bold=True, size=9, wrap=True, ha="center"):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = Font(bold=bold, name="Arial", size=size, color=fc)
    cell.fill      = PatternFill("solid", start_color=bg, end_color=bg)
    cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    cell.border    = _bdr()
    return cell

def dat(ws, r, c, val, bg=None, bold=False, fc="000000", ha="center", fmt=None, wrap=False):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = Font(name="Arial", size=9, bold=bold, color=fc)
    cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    cell.border    = _bdr()
    if bg:
        cell.fill = PatternFill("solid", start_color=bg, end_color=bg)
    if fmt:
        cell.number_format = fmt
    return cell

def title_row(ws, r, ncols, text, bg="1F4E79", size=11):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=text)
    c.font      = Font(bold=True, name="Arial", size=size, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color=bg, end_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 22

def note_row(ws, r, ncols, text, bg="F2F2F2"):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=text)
    c.font      = Font(italic=True, name="Arial", size=8, color="595959")
    c.fill      = PatternFill("solid", start_color=bg, end_color=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 28

def grad_scale(ws, col_letter, r1, r2, lo="FFFFFF", hi="1F4E79"):
    ws.conditional_formatting.add(
        f"{col_letter}{r1}:{col_letter}{r2}",
        ColorScaleRule(start_type="min", start_color=lo,
                       end_type="max",   end_color=hi))

def set_col_widths(ws, widths):
    """widths: list of (col_letter_or_int, width)"""
    for col, w in widths:
        letter = get_column_letter(col) if isinstance(col, int) else col
        ws.column_dimensions[letter].width = w


# ── Data loading ──────────────────────────────────────────────────────────────

def _find(pattern):
    matches = sorted(glob.glob(pattern, recursive=True))
    return Path(matches[-1]) if matches else None

def load_results_json(hmm_dir: Path, t: int, p: int):
    for pat in [
        str(hmm_dir / f"task{t}" / f"paradigm{p}" / "HMM*" / "results" / f"results_T{t}_P{p}_HMM_variable_length.json"),
        str(hmm_dir / "**" / f"results_T{t}_P{p}_HMM_variable_length.json"),
    ]:
        f = _find(pat)
        if f:
            return json.load(open(f))
    return None

def load_checkpoint(hmm_dir: Path, t: int, p: int):
    for pat in [
        str(hmm_dir / f"task{t}" / f"paradigm{p}" / "HMM*" / "model_checkpoints" / f"HMM_T{t}_P{p}_BA*.json"),
        str(hmm_dir / "**" / f"HMM_T{t}_P{p}_BA*.json"),
    ]:
        f = _find(pat)
        if f:
            return json.load(open(f))
    return None

def load_px_details(path: Path) -> pd.DataFrame | None:
    if not path.exists():
        print(f"  [WARN] px_details not found: {path}")
        return None
    for h in range(6):
        df = pd.read_excel(path, sheet_name="Sheet1", header=h,
                           dtype={"id": str})
        if "dia_code" in df.columns:
            df["id"] = df["id"].str.strip().str.upper()
            return df
    return None

def get_ckpt_subject_order(g1_ids, g0_ids):
    """Reproduce the np.unique() ordering used when saving predictions."""
    raw = ([f"g0_{i}_{sid}" for i, sid in enumerate(g0_ids)] +
           [f"g1_{i}_{sid}" for i, sid in enumerate(g1_ids)])
    return [s.split("_", 2)[2] for s in np.unique(raw)]


# ── Patient metadata ──────────────────────────────────────────────────────────

def build_patient_meta(df_px: pd.DataFrame) -> pd.DataFrame:
    """
    Return a tidy DataFrame of patient records with laterality columns.
    Columns: id, age, sex, dia_code, diagnosis, dia_side_r, dia_side_l,
             hand_xr (DOMINANT hand -- inconsistently recorded), C_DASH, X_DASH,
             avg_dash, laterality, is_bilateral, is_right_only, is_left_only,
             dom_congruent (dominant hand == injured side; NaN if hand_xr missing)
    """
    px = df_px[df_px["id"].str.startswith("PX")].copy()

    # Normalise laterality columns
    for col in ["dia_side_r", "dia_side_l"]:
        if col in px.columns:
            px[col] = pd.to_numeric(px[col], errors="coerce").fillna(0).astype(int)
        else:
            px[col] = 0

    # hand_xr = DOMINANT hand. Inconsistently recorded -- NaN where absent.
    # Normalise to "R", "L", "B", or NaN regardless of how it was entered.
    if "hand_xr" in px.columns:
        raw = px["hand_xr"].copy()

        # Diagnostic: show unique raw values so coding can be verified
        unique_raw = raw.dropna().unique().tolist()
        print(f"  [hand_xr] raw unique values in px: {unique_raw}")

        # Normalise: cast to str, strip, upper, then map all known variants
        norm = raw.astype(str).str.strip().str.upper()

        # Map every plausible encoding → canonical R / L / B
        HAND_MAP = {
            # Already canonical
            "R": "R", "L": "L", "B": "B",
            # Full words
            "RIGHT": "R", "LEFT": "L", "BOTH": "B", "BILATERAL": "B",
            # Numbers — this dataset uses 1=Right, 0=Left
            "0": "L", "1": "R", "2": "L", "3": "B",
            # Nullish strings  -> will become NaN below
            "NAN": None, "NONE": None, "NAT": None, "": None, "NA": None,
        }
        px["hand_xr"] = norm.map(lambda v: HAND_MAP.get(v, None))

        # Warn about any values that weren't mapped
        unmapped = norm[px["hand_xr"].isna() & raw.notna() &
                        ~norm.isin(["NAN", "NONE", "NAT", "", "NA"])]
        if len(unmapped):
            print(f"  [hand_xr] WARNING: unmapped values (add to HAND_MAP): "
                  f"{unmapped.unique().tolist()}")
        n_ok = px["hand_xr"].notna().sum()
        print(f"  [hand_xr] {n_ok}/{len(px)} patients have a valid dominant-hand value")
    else:
        print("  [hand_xr] column not found in px_details -- all Unknown")
        px["hand_xr"] = np.nan
    px["hand_xr_available"] = px["hand_xr"].notna()

    # Derived laterality category
    def _lat(row):
        r, l = row["dia_side_r"], row["dia_side_l"]
        if r == 1 and l == 1:
            return "Bilateral"
        elif r == 1:
            return "Right"
        elif l == 1:
            return "Left"
        else:
            return "Unknown"

    px["laterality"]    = px.apply(_lat, axis=1)
    px["is_bilateral"]  = px["laterality"] == "Bilateral"
    px["is_right_only"] = px["laterality"] == "Right"
    px["is_left_only"]  = px["laterality"] == "Left"
    px["is_unilateral"] = px["laterality"].isin(["Right", "Left"])

    # dom_congruent: dominant hand == injured side.
    # True  = dominant arm is injured (max functional interference).
    # False = non-dominant arm injured (may compensate via healthy dominant arm).
    # NaN   = bilateral pathology, or hand_xr not recorded.
    def _dom_congruent(row):
        lat = row["laterality"]
        hxr = row["hand_xr"]
        if lat == "Bilateral" or pd.isna(hxr):
            return np.nan
        if lat == "Right" and hxr == "R":
            return True    # dominant R, injured R
        if lat == "Left" and hxr == "L":
            return True    # dominant L, injured L
        if lat in ("Right", "Left") and hxr in ("R", "L"):
            return False   # dominant != injured
        return np.nan

    px["dom_congruent"] = px.apply(_dom_congruent, axis=1)

    px["diagnosis"] = px["dia_code"].map(DIA_NAMES).fillna("Unknown")
    if "c and x" in px.columns:
        px = px.rename(columns={"c and x": "avg_dash"})
    elif "avg_dash" not in px.columns:
        if "C_DASH" in px.columns and "X_DASH" in px.columns:
            px["avg_dash"] = (pd.to_numeric(px["C_DASH"], errors="coerce") +
                              pd.to_numeric(px["X_DASH"], errors="coerce")) / 2

    return px.reset_index(drop=True)


def build_subject_pred_map(results, ckpt, df_px):
    """
    Returns dict: subject_id.upper() → {y_true, y_pred, y_proba, feature_importance}
    Uses checkpoint predictions (most complete) with results as fallback.
    """
    preds_src = (ckpt or {}).get("predictions") or results.get("predictions", {})
    y_true  = np.array(preds_src.get("y_true", []))
    y_pred  = np.array(preds_src.get("y_pred", []))
    y_proba = np.array(preds_src.get("y_proba", []))

    feat_imp = results.get("feature_importance", {})

    if df_px is None or len(y_true) == 0:
        return {}, feat_imp

    # Reconstruct subject order
    px    = df_px[df_px["id"].str.startswith("PX")]["id"].tolist()
    fx    = [r["id"] for _, r in df_px[~df_px["id"].str.startswith("PX")].iterrows()]
    g1_ids = px   # paradigm 1: all patients = g1
    g0_ids = fx

    ordered = get_ckpt_subject_order(g1_ids, g0_ids)
    pred_map = {}
    for i, sid in enumerate(ordered):
        if i < len(y_true):
            pred_map[sid.upper()] = {
                "y_true":  int(y_true[i]),
                "y_pred":  int(y_pred[i]),
                "y_proba": float(y_proba[i]) if i < len(y_proba) else np.nan,
            }
    return pred_map, feat_imp


# ═══════════════════════════════════════════════════════════════════════════════
# ANALYSIS 1 — Laterality × Feature Importance
# ═══════════════════════════════════════════════════════════════════════════════

def compute_laterality_importance(all_results: list, px_meta: pd.DataFrame,
                                    channel_names: list) -> pd.DataFrame:
    """
    For each task×paradigm, for each unilateral patient:
      - ipsi_rank:   mean rank (1=best) of ipsilateral hand channels
      - contra_rank: mean rank of contralateral hand channels
      - head_rank:   mean rank of head channels
      - ipsi_imp:    sum importance of ipsilateral hand channels
      - contra_imp:  sum importance of contralateral channels
      - head_imp:    sum importance of head channels

    Returns tidy DataFrame with one row per patient×task×paradigm.
    """
    rows = []
    id_to_lat = dict(zip(px_meta["id"].str.upper(), px_meta["laterality"]))

    for entry in all_results:
        t, p    = entry["task"], entry["paradigm"]
        feat_imp = entry.get("feature_importance", {})
        if not feat_imp:
            continue

        # Build importance vector in channel_names order
        imp_vec = np.array([feat_imp.get(ch, 0.0) for ch in channel_names])

        # Ranks (1 = most important)
        ranks = imp_vec.argsort()[::-1].argsort() + 1  # 1-based

        left_imp  = imp_vec[LEFT_IDX].sum()
        right_imp = imp_vec[RIGHT_IDX].sum()
        head_imp  = imp_vec[HEAD_IDX].sum()
        left_rank_mean  = ranks[LEFT_IDX].mean()
        right_rank_mean = ranks[RIGHT_IDX].mean()
        head_rank_mean  = ranks[HEAD_IDX].mean()

        # Top-3 features
        top3 = sorted(feat_imp.items(), key=lambda x: x[1], reverse=True)[:3]
        top3_str = ", ".join(f"{k}({v:.3f})" for k, v in top3)

        # Per-patient analysis
        pred_map = entry.get("pred_map", {})
        for sid_up, preds in pred_map.items():
            if preds.get("y_true") != 1:  # patients only
                continue
            lat = id_to_lat.get(sid_up, "Unknown")
            if lat not in ("Right", "Left"):
                continue

            if lat == "Right":
                ipsi_imp, ipsi_rank = right_imp, right_rank_mean
                contra_imp, contra_rank = left_imp, left_rank_mean
            else:
                ipsi_imp, ipsi_rank = left_imp, left_rank_mean
                contra_imp, contra_rank = right_imp, right_rank_mean

            rows.append({
                "task":        t,
                "task_label":  TASK_LABELS[t],
                "paradigm":    p,
                "subject_id":  sid_up,
                "laterality":  lat,
                "y_pred":      preds.get("y_pred"),
                "y_proba":     preds.get("y_proba"),
                "correct":     preds.get("y_true") == preds.get("y_pred"),
                "ipsi_imp":    round(ipsi_imp,  4),
                "contra_imp":  round(contra_imp, 4),
                "head_imp":    round(head_imp,   4),
                "ipsi_rank":   round(ipsi_rank,  2),
                "contra_rank": round(contra_rank, 2),
                "head_rank":   round(head_rank_mean,   2),
                "ipsi_dominates": ipsi_imp > contra_imp,
                "head_dominates": head_imp > max(ipsi_imp, contra_imp),
                "top3_features": top3_str,
            })

    return pd.DataFrame(rows)


def compute_task_laterality_summary(df: pd.DataFrame, paradigm: int = 1) -> pd.DataFrame:
    """Aggregate laterality importance by task for a given paradigm."""
    df_p = df[df["paradigm"] == paradigm] if "paradigm" in df.columns else df
    rows = []
    for t in TASKS:
        sub = df_p[df_p["task"] == t]
        if sub.empty:
            continue
        n_right = len(sub[sub["laterality"] == "Right"])
        n_left  = len(sub[sub["laterality"] == "Left"])
        n_total = len(sub)

        n_ipsi_dom  = sub["ipsi_dominates"].sum()
        n_head_dom  = sub["head_dominates"].sum()
        n_correct   = sub["correct"].sum()
        pct_ipsi    = n_ipsi_dom / n_total if n_total else np.nan
        pct_head    = n_head_dom / n_total if n_total else np.nan
        pct_correct = n_correct  / n_total if n_total else np.nan

        rows.append({
            "task":            t,
            "task_label":      TASK_LABELS[t],
            "n_unilateral_px": n_total,
            "n_right":         n_right,
            "n_left":          n_left,
            "mean_ipsi_imp":   round(sub["ipsi_imp"].mean(),  4),
            "mean_contra_imp": round(sub["contra_imp"].mean(), 4),
            "mean_head_imp":   round(sub["head_imp"].mean(),   4),
            "pct_ipsi_dom":    round(pct_ipsi,    3),
            "pct_head_dom":    round(pct_head,    3),
            "pct_correct":     round(pct_correct, 3),
            "interpretation": (
                "Head/trunk compensation dominates" if pct_head > 0.5
                else "Ipsilateral restriction dominates" if pct_ipsi > 0.5
                else "Contralateral/mixed signal"
            ),
        })
    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════════════════
# ANALYSIS 2 — Dominant-Hand x Injured-Side Cross-Tabulation
# ═══════════════════════════════════════════════════════════════════════════════

def compute_used_hand_crosstab(all_results: list, px_meta: pd.DataFrame) -> tuple:
    """
    Returns (subject_df, crosstab_df).

    Cross-tabulates injured side (dia_side_r/l) against dominant hand (hand_xr).
    hand_xr is the DOMINANT hand, not the hand used during the task, and is
    inconsistently recorded. Rows with missing hand_xr get cell='<Lat>_Unknown'
    and are included in per-subject table but excluded from aggregated means.

    subject_df: one row per patient x task x paradigm with cell label.
    crosstab_df: aggregated by cell x task (paradigm 1 only, known hand_xr only).
    """
    id_to_row = {r["id"].upper(): r for _, r in px_meta.iterrows()}

    subj_rows = []
    for entry in all_results:
        t, p = entry["task"], entry["paradigm"]
        pred_map = entry.get("pred_map", {})

        for sid_up, preds in pred_map.items():
            if preds.get("y_true") != 1:
                continue
            meta = id_to_row.get(sid_up)
            if meta is None:
                continue

            lat    = meta["laterality"]
            hand   = str(meta.get("hand_xr", "")).strip().upper()
            if hand in ("", "NAN", "NONE", "NAT"):
                hand = np.nan

            # Cell = injured_side x dominant_hand relationship
            if lat == "Bilateral":
                cell = "Bilateral"
            elif pd.isna(hand):
                cell = f"{lat}_Unknown"           # hand_xr not recorded
            elif (lat == "Right" and hand == "R") or (lat == "Left" and hand == "L"):
                cell = f"{lat}_DomInjured"        # dominant = injured side
            elif (lat == "Right" and hand == "L") or (lat == "Left" and hand == "R"):
                cell = f"{lat}_DomUnaffected"     # dominant = healthy side
            else:
                cell = f"{lat}_{hand}"

            subj_rows.append({
                "task":       t,
                "task_label": TASK_LABELS[t],
                "paradigm":   p,
                "subject_id": sid_up,
                "laterality": lat,
                "dominant_hand":     hand if not pd.isna(hand) else "Unknown",
                "hand_xr_recorded":  not pd.isna(hand),
                "cell":       cell,
                "y_proba":    preds.get("y_proba", np.nan),
                "correct":    preds.get("y_true") == preds.get("y_pred"),
                "diagnosis":  meta.get("diagnosis", ""),
            })

    subj_df = pd.DataFrame(subj_rows)
    if subj_df.empty:
        return subj_df, pd.DataFrame()

    # Aggregate per paradigm × cell × task (known hand_xr only)
    xtab_rows = []
    for p in PARADIGMS:
        dfp = subj_df[subj_df["paradigm"] == p]
        for t in TASKS:
            tsub = dfp[dfp["task"] == t]
            for cell in ["Right_DomInjured", "Right_DomUnaffected",
                         "Left_DomInjured",  "Left_DomUnaffected",  "Bilateral"]:
                csub = tsub[tsub["cell"] == cell]
                n = len(csub)
                xtab_rows.append({
                    "paradigm":    p,
                    "task":        t,
                    "task_label":  TASK_LABELS[t],
                    "cell":        cell,
                    "n":           n,
                    "mean_proba":  round(csub["y_proba"].mean(), 3) if n else np.nan,
                    "pct_correct": round(csub["correct"].mean(), 3) if n else np.nan,
                    "n_correct":   int(csub["correct"].sum()) if n else 0,
                })
    xtab_df = pd.DataFrame(xtab_rows)
    return subj_df, xtab_df


# ═══════════════════════════════════════════════════════════════════════════════
# ANALYSIS 3 — HMM Probability vs Self-Reported Difficulty
# ═══════════════════════════════════════════════════════════════════════════════

def compute_proba_difficulty_correlation(all_results: list,
                                          px_meta: pd.DataFrame,
                                          df_px_full: pd.DataFrame) -> tuple:
    """
    Returns (corr_df, subj_df, misclass_df).

    corr_df:   Spearman rho + p-value per task×paradigm (patients only).
    subj_df:   per-subject proba vs q* for all tasks (paradigm 1).
    misclass_df: misclassified patients' q* scores.
    """
    id_to_row = {r["id"].upper(): r for _, r in df_px_full.iterrows()}

    subj_rows = []
    for entry in all_results:
        t, p = entry["task"], entry["paradigm"]
        q_col = TASK_Q_COL.get(t)
        pred_map = entry.get("pred_map", {})

        for sid_up, preds in pred_map.items():
            meta = id_to_row.get(sid_up)
            q_score = None
            if meta is not None and q_col and q_col in meta:
                q_raw = meta[q_col]
                q_score = float(q_raw) if pd.notna(q_raw) else None

            subj_rows.append({
                "task":       t,
                "task_label": TASK_LABELS[t],
                "paradigm":   p,
                "subject_id": sid_up,
                "y_true":     preds.get("y_true"),
                "y_pred":     preds.get("y_pred"),
                "y_proba":    preds.get("y_proba", np.nan),
                "correct":    preds.get("y_true") == preds.get("y_pred"),
                "q_score":    q_score,
                "group":      "Patient" if preds.get("y_true") == 1 else "Control",
                "diagnosis":  id_to_row.get(sid_up, {}).get("diagnosis", ""),
                "laterality": px_meta[px_meta["id"] == sid_up]["laterality"].values[0]
                              if sid_up in px_meta["id"].values else "Unknown",
            })

    subj_df = pd.DataFrame(subj_rows)
    if subj_df.empty:
        return pd.DataFrame(), subj_df, pd.DataFrame()

    # Spearman correlations (all subjects, patients only)
    corr_rows = []
    for t in TASKS:
        for p in PARADIGMS:
            sub = subj_df[(subj_df["task"] == t) & (subj_df["paradigm"] == p)]
            px_sub = sub[sub["y_true"] == 1].dropna(subset=["y_proba", "q_score"])
            all_sub = sub.dropna(subset=["y_proba", "q_score"])

            def _spearman(df):
                if len(df) < 4:
                    return np.nan, np.nan
                rho, pv = stats.spearmanr(df["y_proba"], df["q_score"])
                return round(float(rho), 3), round(float(pv), 4)

            rho_all,  pv_all  = _spearman(all_sub)
            rho_px,   pv_px   = _spearman(px_sub)

            corr_rows.append({
                "task":          t,
                "task_label":    TASK_LABELS[t],
                "paradigm":      p,
                "paradigm_label": PARADIGM_LABELS[p],
                "n_all":         len(all_sub),
                "n_patients":    len(px_sub),
                "rho_all":       rho_all,
                "pv_all":        pv_all,
                "rho_patients":  rho_px,
                "pv_patients":   pv_px,
            })

    corr_df = pd.DataFrame(corr_rows)

    # ── Benjamini-Hochberg FDR correction ──────────────────────────────────
    # Applied separately within the all-subjects family and the
    # within-patient family, across all 24 task x paradigm cells.
    # Significance is assessed at FDR = 0.05 (q < 0.05); uncorrected
    # p-values (pv_all / pv_patients) are kept alongside for reference.
    def _bh_correct(pvals: pd.Series) -> pd.Series:
        q = pd.Series(np.nan, index=pvals.index)
        mask = pvals.notna()
        if mask.sum() > 0:
            _, q_vals, _, _ = multipletests(
                pvals[mask].values, alpha=0.05, method="fdr_bh")
            q[mask] = q_vals
        return q

    corr_df["q_all"]      = _bh_correct(corr_df["pv_all"]).round(4)
    corr_df["q_patients"] = _bh_correct(corr_df["pv_patients"]).round(4)
    corr_df["sig_all"]      = corr_df["q_all"].apply(
        lambda q: "✓" if pd.notna(q) and q < 0.05 else "")
    corr_df["sig_patients"] = corr_df["q_patients"].apply(
        lambda q: "✓" if pd.notna(q) and q < 0.05 else "")

    # Misclassified patient insight
    miss_rows = []
    for entry in all_results:
        t, p = entry["task"], entry["paradigm"]
        q_col = TASK_Q_COL.get(t)
        pred_map = entry.get("pred_map", {})
        for sid_up, preds in pred_map.items():
            yt, yp = preds.get("y_true"), preds.get("y_pred")
            if yt != 1 or yp != 0:   # only FN (patient misclassified as control)
                continue
            meta = id_to_row.get(sid_up, {})
            q_score = None
            if q_col and q_col in meta:
                q_raw = meta[q_col]
                q_score = float(q_raw) if pd.notna(q_raw) else None

            miss_rows.append({
                "task":       t,
                "task_label": TASK_LABELS[t],
                "paradigm":   p,
                "subject_id": sid_up,
                "diagnosis":  meta.get("diagnosis", ""),
                "laterality": px_meta[px_meta["id"] == sid_up]["laterality"].values[0]
                              if sid_up in px_meta["id"].values else "Unknown",
                "dominant_hand": str(meta.get("hand_xr", "")).strip().upper(),
                "hand_xr_recorded": (pd.notna(meta.get("hand_xr")) and
                                     str(meta.get("hand_xr","")).strip().upper()
                                     not in ("", "NAN", "NONE")),
                "y_proba":    preds.get("y_proba", np.nan),
                "q_score":    q_score,
                "q_col":      q_col,
                "c_dash":     meta.get("C_DASH"),
                "x_dash":     meta.get("X_DASH"),
                "note": (
                    "Low self-reported difficulty — consistent with HMM error"
                    if q_score is not None and q_score <= 2
                    else "High self-reported difficulty — HMM missed clear pathology"
                    if q_score is not None and q_score >= 4
                    else ""
                ),
            })

    misclass_df = pd.DataFrame(miss_rows)
    return corr_df, subj_df, misclass_df


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL SHEET WRITERS
# ═══════════════════════════════════════════════════════════════════════════════

def sheet_overview(ws, px_meta: pd.DataFrame):
    """Sheet 1 — Cohort overview and laterality breakdown."""
    NCOLS = 10

    title_row(ws, 1, NCOLS, "XDash HMM — Laterality Analysis Overview", "1F4E79", size=12)
    note_row(ws, 2, NCOLS,
             "Cohort breakdown by injured side and dominant hand (hand_xr). "
             "Bilateral = pathology on both sides (n=6). "
             "hand_xr = DOMINANT hand, NOT the hand used during the XR task. "
             "Inconsistently recorded -- missing for some patients. "
             "Dom-Injured = dominant arm is the injured side. "
             "Dom-Unaffected = non-dominant arm injured.")

    # ── Cohort summary table ──
    r = 4
    hdr(ws, r, 1, "Laterality Group",  C["navy"],  size=9)
    hdr(ws, r, 2, "N",                  C["navy"],  size=9)
    hdr(ws, r, 3, "% of Patients",      C["navy"],  size=9)
    hdr(ws, r, 4, "Diagnoses",           C["navy"],  size=9, ha="left")
    hdr(ws, r, 5, "Mean C_DASH",         C["navy"],  size=9)
    hdr(ws, r, 6, "Mean X_DASH",         C["navy"],  size=9)
    hdr(ws, r, 7, "Mean Avg DASH",       C["navy"],  size=9)
    ws.row_dimensions[r].height = 18

    groups = [
        ("Right (only)",  px_meta[px_meta["is_right_only"]],  C["r_light"]),
        ("Left (only)",   px_meta[px_meta["is_left_only"]],   C["l_light"]),
        ("Bilateral",     px_meta[px_meta["is_bilateral"]],   C["bil"]),
        ("Unknown",       px_meta[px_meta["laterality"] == "Unknown"], C["offwhite"]),
        ("ALL PATIENTS",  px_meta,                             "F0F0F0"),
    ]
    n_px = len(px_meta)
    for i, (label, sub, bg) in enumerate(groups):
        rr = r + 1 + i
        n  = len(sub)
        diag_str = ", ".join(sorted(sub["diagnosis"].value_counts().index.tolist()))
        dat(ws, rr, 1, label,                bg, bold=(label == "ALL PATIENTS"), ha="left")
        dat(ws, rr, 2, n,                    bg)
        dat(ws, rr, 3, f"{n/n_px*100:.1f}%" if n_px else "—", bg)
        dat(ws, rr, 4, diag_str,             bg, ha="left", wrap=True)
        dat(ws, rr, 5, round(pd.to_numeric(sub["C_DASH"],   errors="coerce").mean(), 1) if n else "—", bg, fmt="0.0")
        dat(ws, rr, 6, round(pd.to_numeric(sub["X_DASH"],   errors="coerce").mean(), 1) if n else "—", bg, fmt="0.0")
        dat(ws, rr, 7, round(pd.to_numeric(sub["avg_dash"], errors="coerce").mean(), 1) if n else "—", bg, fmt="0.0")
        ws.row_dimensions[rr].height = 16

    # ── Congruence sub-table ──
    r2 = r + len(groups) + 2
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=NCOLS)
    hdr(ws, r2, 1, "Dominant-Hand vs Injured-Side (Unilateral Patients)", C["teal"], size=10)

    r2 += 1
    for lbl in ["Congruent Column", "Compensated Column", "Unknown"]:
        hdr(ws, r2, 1, "Category",  C["teal"])
        hdr(ws, r2, 2, "N",         C["teal"])
        hdr(ws, r2, 3, "% Unilat.", C["teal"])
        hdr(ws, r2, 4, "Description", C["teal"], ha="left")
        ws.row_dimensions[r2].height = 16
        break

    uni  = px_meta[px_meta["is_unilateral"]]
    cong = uni[uni["dom_congruent"] == True]
    comp = uni[uni["dom_congruent"] == False]
    unkn = uni[uni["dom_congruent"].isna()]
    n_uni = len(uni)
    for label, sub, bg, desc in [
        ("Dom-Injured (dominant = injured side)", cong, C["r_light"],
         "Dominant arm has pathology — maximum interference with habitual movement patterns"),
        ("Dom-Unaffected (dominant = healthy side)", comp, C["l_light"],
         "Non-dominant arm injured — patient can rely on healthy dominant arm; subtler signal"),
        ("Unknown (hand_xr not recorded)",       unkn, C["bil"],
         "hand_xr missing — excluded from Dom-Injured/Dom-Unaffected groupings"),
        ("All Unilateral",                       uni,  "F0F0F0", ""),
    ]:
        r2 += 1
        n = len(sub)
        dat(ws, r2, 1, label,                     bg, bold=(label == "All Unilateral"), ha="left")
        dat(ws, r2, 2, n,                          bg)
        dat(ws, r2, 3, f"{n/n_uni*100:.1f}%" if n_uni else "—", bg)
        dat(ws, r2, 4, desc,                       bg, ha="left", wrap=True)
        ws.row_dimensions[r2].height = 16

    set_col_widths(ws, [(1, 32), (2, 8), (3, 14), (4, 55), (5, 12), (6, 12), (7, 12)])
    ws.freeze_panes = "A4"


def sheet_analysis1(ws, summary_df: pd.DataFrame, detail_df: pd.DataFrame, paradigm: int = 1):
    """Laterality × Feature Importance for one paradigm."""
    NCOLS = 13
    p_label = PARADIGM_LABELS[paradigm]

    title_row(ws, 1, NCOLS,
              f"Analysis 1 — Laterality × Feature Importance  |  P{paradigm}: {p_label}", "1F4E79")
    note_row(ws, 2, NCOLS,
             "For each task, do HMM feature importances point to the INJURED side (ipsilateral) "
             "or the HEAD/TRUNK (compensation)? "
             "Ipsi_dominates = ipsilateral hand importance > contralateral. "
             "Head_dominates = head importance > both hand groups. "
             "Ranks: lower = more important (1 = top feature).")

    # ── Task summary ──
    r = 4
    title_row(ws, r, NCOLS, f"Task-Level Summary — Paradigm {paradigm}: {p_label}", C["blue"], size=10)
    r += 1

    COLS = [
        ("Task", 14), ("N Unilateral Px", 10), ("N Right / Left", 12),
        ("Mean Ipsi Imp", 12), ("Mean Contra Imp", 12), ("Mean Head Imp", 12),
        ("% Ipsi Dominant", 13), ("% Head Dominant", 13),
        ("% Correct (LOO)", 13), ("Interpretation", 38),
    ]
    for ci, (lbl, w) in enumerate(COLS):
        hdr(ws, r, ci+1, lbl, C["blue"])
        ws.column_dimensions[get_column_letter(ci+1)].width = w
    ws.row_dimensions[r].height = 18
    r += 1

    if not summary_df.empty:
        for _, row in summary_df.iterrows():
            bg = C["r_light"] if row.get("pct_head_dom", 0) > 0.5 else C["l_light"]
            dat(ws, r, 1,  f"T{row['task']}: {row['task_label']}", bg, bold=True, ha="left")
            dat(ws, r, 2,  row["n_unilateral_px"],                   bg)
            dat(ws, r, 3,  f"{row['n_right']}R / {row['n_left']}L",  bg)
            dat(ws, r, 4,  row["mean_ipsi_imp"],                      bg, fmt="0.0000")
            dat(ws, r, 5,  row["mean_contra_imp"],                    bg, fmt="0.0000")
            dat(ws, r, 6,  row["mean_head_imp"],                      bg, fmt="0.0000")
            dat(ws, r, 7,  f"{row['pct_ipsi_dom']*100:.1f}%",        bg)
            dat(ws, r, 8,  f"{row['pct_head_dom']*100:.1f}%",        bg)
            dat(ws, r, 9,  f"{row['pct_correct']*100:.1f}%",         bg)
            dat(ws, r, 10, row["interpretation"],                      bg, ha="left", wrap=True)
            ws.row_dimensions[r].height = 18
            r += 1

    if not summary_df.empty:
        data_r1 = r - len(summary_df)
        data_r2 = r - 1
        for col_letter in ["D", "E", "F"]:
            grad_scale(ws, col_letter, data_r1, data_r2)

    # ── Per-subject detail table ──
    r += 2
    title_row(ws, r, NCOLS,
              f"Per-Subject Detail — Paradigm {paradigm} Unilateral Patients", C["purple"], size=10)
    r += 1

    DCOLS = [
        ("Task", 14), ("Subject", 10), ("Laterality", 10), ("Correct?", 8),
        ("HMM Proba", 10), ("Ipsi Imp", 10), ("Contra Imp", 10), ("Head Imp", 10),
        ("Ipsi Rank", 10), ("Head Rank", 10), ("Ipsi Dom?", 9), ("Head Dom?", 9),
        ("Top-3 Features", 45),
    ]
    for ci, (lbl, w) in enumerate(DCOLS):
        hdr(ws, r, ci+1, lbl, C["purple"])
        ws.column_dimensions[get_column_letter(ci+1)].width = w
    ws.row_dimensions[r].height = 18
    r += 1

    if not detail_df.empty:
        dfp = detail_df[detail_df["paradigm"] == paradigm].sort_values(
            ["task", "laterality", "subject_id"])
        for _, row in dfp.iterrows():
            correct = row.get("correct", True)
            bg = C["miss"] if not correct else (C["r_light"] if row["laterality"] == "Right"
                                                else C["l_light"])
            dat(ws, r, 1,  f"T{row['task']}: {row['task_label']}", bg, ha="left")
            dat(ws, r, 2,  row["subject_id"],                       bg)
            dat(ws, r, 3,  row["laterality"],                       bg)
            dat(ws, r, 4,  "✓" if correct else "✗",                 bg, bold=not correct)
            dat(ws, r, 5,  row["y_proba"],                           bg, fmt="0.000")
            dat(ws, r, 6,  row["ipsi_imp"],                          bg, fmt="0.0000")
            dat(ws, r, 7,  row["contra_imp"],                        bg, fmt="0.0000")
            dat(ws, r, 8,  row["head_imp"],                          bg, fmt="0.0000")
            dat(ws, r, 9,  row["ipsi_rank"],                         bg, fmt="0.00")
            dat(ws, r, 10, row["head_rank"],                         bg, fmt="0.00")
            dat(ws, r, 11, "Yes" if row["ipsi_dominates"] else "No", bg)
            dat(ws, r, 12, "Yes" if row["head_dominates"] else "No", bg)
            dat(ws, r, 13, row["top3_features"],                     bg, ha="left", wrap=True)
            ws.row_dimensions[r].height = 16
            r += 1

    ws.freeze_panes = "A5"


def sheet_analysis2(ws, subj_df: pd.DataFrame, xtab_df: pd.DataFrame, paradigm: int = 1):
    """Dominant-Hand × Injured-Side Cross-Tabulation for one paradigm."""
    NCOLS = 9
    p_label = PARADIGM_LABELS[paradigm]

    title_row(ws, 1, NCOLS,
              f"Analysis 2 — Dominant-Hand x Injured-Side  |  P{paradigm}: {p_label}", "1F4E79")
    note_row(ws, 2, NCOLS,
             "hand_xr = DOMINANT hand (not the hand used during the task). Inconsistently recorded. "
             "Dom-Injured = dominant arm is injured (max functional interference with habitual movement). "
             "Dom-Unaffected = non-dominant arm injured (patient can rely on healthy dominant arm). "
             "Unknown = hand_xr missing -- subjects appear in per-subject table but excluded from aggregated means. "
             "Hypothesis: Dom-Injured patients show stronger HMM signal (higher proba) because "
             "pathology directly disrupts their primary movement strategy.")

    # ── Cross-tab pivot: rows=cell, cols=task ──
    r = 4
    CELLS_ORDER = ["Right_DomInjured", "Right_DomUnaffected",
                   "Left_DomInjured",  "Left_DomUnaffected",  "Bilateral"]
    CELL_BG = {
        "Right_DomInjured":    C["r_light"],
        "Right_DomUnaffected": "FADBD8",
        "Left_DomInjured":     C["l_light"],
        "Left_DomUnaffected":  "D5F5E3",
        "Bilateral":           C["bil"],
    }
    CELL_LABEL = {
        "Right_DomInjured":    "Right — Dom-Injured\n(dom R = injured R)",
        "Right_DomUnaffected": "Right — Dom-Unaffected\n(dom L, injured R)",
        "Left_DomInjured":     "Left — Dom-Injured\n(dom L = injured L)",
        "Left_DomUnaffected":  "Left — Dom-Unaffected\n(dom R, injured L)",
        "Bilateral":           "Bilateral",
    }

    # Header row: task labels
    hdr(ws, r, 1, "Cell", C["navy"], size=9)
    hdr(ws, r, 2, "N",    C["navy"], size=9)
    for ti, t in enumerate(TASKS):
        # Mean proba
        hdr(ws, r, 3 + ti*2,     f"T{t} {TASK_LABELS[t][:8]}\nMean Proba", C["navy"])
        hdr(ws, r, 3 + ti*2 + 1, f"T{t}\n% Correct",                        C["navy"])
    ws.row_dimensions[r].height = 32
    r += 1

    if not xtab_df.empty:
        xtab_p = xtab_df[xtab_df["paradigm"] == paradigm]
        subj_p = subj_df[subj_df["paradigm"] == paradigm] if not subj_df.empty else subj_df
        for cell in CELLS_ORDER:
            bg = CELL_BG.get(cell, C["offwhite"])
            n_px = len(subj_p[subj_p["cell"] == cell]["subject_id"].unique())
            dat(ws, r, 1, CELL_LABEL[cell], bg, ha="left", wrap=True, bold=True)
            dat(ws, r, 2, n_px, bg)

            for ti, t in enumerate(TASKS):
                row_data = xtab_p[(xtab_p["cell"] == cell) & (xtab_p["task"] == t)]
                if row_data.empty:
                    dat(ws, r, 3 + ti*2,     "—", bg)
                    dat(ws, r, 3 + ti*2 + 1, "—", bg)
                else:
                    proba = row_data.iloc[0]["mean_proba"]
                    pct_c = row_data.iloc[0]["pct_correct"]
                    n_c   = row_data.iloc[0]["n_correct"]
                    n_all = row_data.iloc[0]["n"]
                    dat(ws, r, 3 + ti*2,
                        round(proba, 3) if pd.notna(proba) else "—", bg, fmt="0.000")
                    dat(ws, r, 3 + ti*2 + 1,
                        f"{pct_c*100:.0f}%\n({n_c}/{n_all})" if pd.notna(pct_c) else "—",
                        bg, wrap=True)

            ws.row_dimensions[r].height = 30
            r += 1

    # ── Per-subject table ──
    r += 2
    title_row(ws, r, NCOLS,
              f"Per-Subject Detail — Paradigm {paradigm}", C["teal"], size=10)
    r += 1

    SCOLS = [("Task", 14), ("Subject", 10), ("Laterality", 10),
             ("Dom. Hand\n(hand_xr)", 12), ("Recorded?", 9),
             ("Cell", 26), ("HMM Proba", 10), ("Correct?", 8), ("Diagnosis", 22)]
    for ci, (lbl, w) in enumerate(SCOLS):
        hdr(ws, r, ci+1, lbl, C["teal"])
        ws.column_dimensions[get_column_letter(ci+1)].width = w
    ws.row_dimensions[r].height = 18
    r += 1

    if not subj_df.empty:
        dfp = subj_df[subj_df["paradigm"] == paradigm].sort_values(["cell", "task", "subject_id"])
        for _, row in dfp.iterrows():
            bg = CELL_BG.get(row["cell"], C["offwhite"])
            correct = row.get("correct", True)
            if not correct:
                bg = C["miss"]
            dat(ws, r, 1, f"T{row['task']}: {row['task_label']}", bg, ha="left")
            dat(ws, r, 2, row["subject_id"], bg)
            dat(ws, r, 3, row["laterality"], bg)
            dat(ws, r, 4, row.get("dominant_hand", "—"), bg)
            dat(ws, r, 5, "Yes" if row.get("hand_xr_recorded") else "No", bg,
                fc="375623" if row.get("hand_xr_recorded") else "595959")
            dat(ws, r, 6, row["cell"],        bg, ha="left")
            dat(ws, r, 7, row["y_proba"],     bg, fmt="0.000")
            dat(ws, r, 8, "✓" if correct else "✗", bg, bold=not correct)
            dat(ws, r, 9, row["diagnosis"],   bg, ha="left")
            ws.row_dimensions[r].height = 16
            r += 1

    # Column widths for pivot
    set_col_widths(ws, [(1, 26), (2, 6)] +
                   [(3 + i, 11) for i in range(len(TASKS) * 2)])
    ws.freeze_panes = "A5"


def sheet_analysis3_corr(ws, corr_df: pd.DataFrame):
    """Sheet 4 — Spearman Correlation table."""
    NCOLS = 13

    title_row(ws, 1, NCOLS,
              "Analysis 3 — HMM Probability vs Self-Reported Difficulty  (Spearman ρ)", "1F4E79")
    note_row(ws, 2, NCOLS,
             "Spearman ρ between HMM y_proba and task-specific Q score (1=no difficulty … 5=unable). "
             "rho_all = all subjects (patients + controls). rho_patients = patients only. "
             "p-values are corrected for multiple comparisons across all 24 task x paradigm cells "
             "using Benjamini-Hochberg FDR, applied separately within the all-subjects and "
             "within-patient families. ✓ = q < 0.05 (FDR-corrected). Uncorrected p-values shown alongside. "
             "Strong positive correlation = HMM confidence aligns with self-reported impairment. "
             "Paradigm 1 (all patients vs controls) shown first for interpretability.")

    r = 4
    COLS = [
        ("Task", 14), ("Paradigm", 22), ("N All", 7), ("N Patients", 9),
        ("ρ All Subjects", 13), ("p All", 9), ("q All (FDR)", 10), ("Sig?", 5),
        ("ρ Patients Only", 13), ("p Patients", 9), ("q Patients (FDR)", 10), ("Sig?", 5),
        ("Interpretation", 38),
    ]
    for ci, (lbl, w) in enumerate(COLS):
        hdr(ws, r, ci+1, lbl, C["navy"])
        ws.column_dimensions[get_column_letter(ci+1)].width = w
    ws.row_dimensions[r].height = 18
    r += 1

    def _interp(rho, qv):
        if pd.isna(rho) or pd.isna(qv):
            return "Insufficient data"
        sig = qv < 0.05
        strength = ("|ρ|>0.5 strong" if abs(rho) > 0.5 else
                    "|ρ|>0.3 moderate" if abs(rho) > 0.3 else "weak")
        direction = "positive (HMM↑ with difficulty↑)" if rho > 0 else "negative"
        return f"{strength}, {direction}{' *' if sig else ''}"

    if not corr_df.empty:
        prev_t = None
        for _, row in corr_df.sort_values(["task", "paradigm"]).iterrows():
            t = row["task"]
            bg = C["b_light"] if t % 2 == 0 else C["offwhite"]
            is_p1 = row["paradigm"] == 1
            bold_it = is_p1

            dat(ws, r, 1,  f"T{t}: {row['task_label']}", bg, bold=bold_it, ha="left")
            dat(ws, r, 2,  f"P{row['paradigm']}: {row['paradigm_label']}", bg, ha="left")
            dat(ws, r, 3,  row["n_all"],       bg)
            dat(ws, r, 4,  row["n_patients"],  bg)
            dat(ws, r, 5,  row["rho_all"],     bg, fmt="0.000",
                bold=(pd.notna(row["rho_all"]) and abs(row["rho_all"]) > 0.3))
            dat(ws, r, 6,  row["pv_all"],      bg, fmt="0.0000")
            dat(ws, r, 7,  row["q_all"],       bg, fmt="0.0000")
            dat(ws, r, 8,  row["sig_all"],     bg, bold=True,
                fc="375623" if row["sig_all"] == "✓" else "000000")
            dat(ws, r, 9,  row["rho_patients"],bg, fmt="0.000",
                bold=(pd.notna(row["rho_patients"]) and abs(row["rho_patients"]) > 0.3))
            dat(ws, r, 10, row["pv_patients"], bg, fmt="0.0000")
            dat(ws, r, 11, row["q_patients"],  bg, fmt="0.0000")
            dat(ws, r, 12, row["sig_patients"],bg, bold=True,
                fc="375623" if row["sig_patients"] == "✓" else "000000")
            dat(ws, r, 13, _interp(row["rho_patients"], row["q_patients"]),
                bg, ha="left", wrap=True)
            ws.row_dimensions[r].height = 16

            if t != prev_t:
                ws.row_dimensions[r].height = 18
            prev_t = t
            r += 1

    # Gradient on rho columns
    if not corr_df.empty:
        grad_scale(ws, "E", 5, r-1, "FFC7CE", "C6EFCE")
        grad_scale(ws, "I", 5, r-1, "FFC7CE", "C6EFCE")

    ws.freeze_panes = "A5"


def sheet_analysis3_misclass(ws, misclass_df: pd.DataFrame, subj_df: pd.DataFrame):
    """Sheet 5 — Misclassification insight (FN patients) + top-task q* comparison."""
    NCOLS = 11

    title_row(ws, 1, NCOLS,
              "Analysis 3b — Misclassified Patients (False Negatives) & Q-Score Insight", "7B2C2C")
    note_row(ws, 2, NCOLS,
             "False Negatives = patients predicted as controls by LOO-CV HMM. "
             "If a misclassified patient also reports LOW task difficulty (Q=1-2), "
             "the HMM error is clinically coherent — patient moves near-normally for this task. "
             "If Q is HIGH (4-5) but HMM still misclassifies, the HMM missed genuine impairment.")

    # ── FN table ──
    r = 4
    hdr(ws, r, 1,  "Task",        C["red"])
    hdr(ws, r, 2,  "Paradigm",    C["red"])
    hdr(ws, r, 3,  "Subject ID",  C["red"])
    hdr(ws, r, 4,  "Diagnosis",   C["red"])
    hdr(ws, r, 5,  "Laterality",  C["red"])
    hdr(ws, r, 6,  "Dom. Hand\n(hand_xr)", C["red"])
    hdr(ws, r, 7,  "HMM Proba",   C["red"])
    hdr(ws, r, 8,  "Q Score\n(task difficulty)", C["red"])
    hdr(ws, r, 9,  "C_DASH",      C["red"])
    hdr(ws, r, 10, "X_DASH",      C["red"])
    hdr(ws, r, 11, "Interpretation", C["red"])
    ws.row_dimensions[r].height = 28
    r += 1

    if not misclass_df.empty:
        for _, row in misclass_df.sort_values(["task", "paradigm", "subject_id"]).iterrows():
            q = row.get("q_score")
            if q is not None and pd.notna(q):
                bg = C["l_light"] if q <= 2 else C["miss"] if q >= 4 else C["h_light"]
            else:
                bg = C["offwhite"]

            dat(ws, r, 1,  f"T{row['task']}: {row['task_label']}", bg, ha="left")
            dat(ws, r, 2,  f"P{row['paradigm']}",                   bg)
            dat(ws, r, 3,  row["subject_id"],                        bg)
            dat(ws, r, 4,  row["diagnosis"],                         bg, ha="left")
            dat(ws, r, 5,  row["laterality"],                        bg)
            dat(ws, r, 6,  row.get("dominant_hand", "—"),            bg)
            dat(ws, r, 7,  row["y_proba"],                           bg, fmt="0.000")
            dat(ws, r, 8,  q if pd.notna(q) else "—",               bg, bold=True,
                fc=("375623" if (q is not None and pd.notna(q) and q <= 2)
                    else "7B2C2C" if (q is not None and pd.notna(q) and q >= 4)
                    else "000000"))
            dat(ws, r, 9,  row.get("c_dash", "—"),                   bg, fmt="0.0")
            dat(ws, r, 10, row.get("x_dash", "—"),                   bg, fmt="0.0")
            dat(ws, r, 11, row.get("note", ""),                       bg, ha="left", wrap=True)
            ws.row_dimensions[r].height = 18
            r += 1
    else:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
        ws.cell(row=r, column=1, value="No false negatives found in loaded results.").font = \
            Font(italic=True, name="Arial", size=9)
        r += 1

    # ── Q-score distribution by group per task (paradigm 1) ──
    r += 2
    title_row(ws, r, NCOLS, "Q-Score Distribution by Group (Paradigm 1, all tasks)", C["amber"], size=10)
    r += 1
    note_row(ws, r, NCOLS,
             "Mean task-specific Q score (1=no difficulty, 5=unable) for patients vs controls per task. "
             "Controls expected near 1. Large patient–control gap = more discriminable task for self-report.")
    r += 1

    hdr(ws, r, 1, "Task",           C["amber"])
    hdr(ws, r, 2, "N Patients",     C["amber"])
    hdr(ws, r, 3, "N Controls",     C["amber"])
    hdr(ws, r, 4, "Mean Q Patients",C["amber"])
    hdr(ws, r, 5, "Mean Q Controls",C["amber"])
    hdr(ws, r, 6, "Q Gap",          C["amber"])
    hdr(ws, r, 7, "Mean Proba Px",  C["amber"])
    hdr(ws, r, 8, "Mean Proba Ctrl",C["amber"])
    ws.row_dimensions[r].height = 18
    r += 1

    if not subj_df.empty:
        df1 = subj_df[subj_df["paradigm"] == 1]
        for t in TASKS:
            tsub = df1[df1["task"] == t].dropna(subset=["q_score"])
            px_s  = tsub[tsub["y_true"] == 1]
            ct_s  = tsub[tsub["y_true"] == 0]
            q_px  = round(px_s["q_score"].mean(), 2)  if len(px_s)  else np.nan
            q_ct  = round(ct_s["q_score"].mean(), 2)  if len(ct_s)  else np.nan
            q_gap = round(q_px - q_ct, 2)              if (pd.notna(q_px) and pd.notna(q_ct)) else np.nan
            pr_px = round(px_s["y_proba"].mean(), 3)  if len(px_s)  else np.nan
            pr_ct = round(ct_s["y_proba"].mean(), 3)  if len(ct_s)  else np.nan

            gap_bg = (C["r_light"] if (pd.notna(q_gap) and q_gap > 1.5)
                      else C["h_light"] if (pd.notna(q_gap) and q_gap > 0.8)
                      else C["offwhite"])
            dat(ws, r, 1, f"T{t}: {TASK_LABELS[t]}", gap_bg, ha="left", bold=True)
            dat(ws, r, 2, len(px_s),  gap_bg)
            dat(ws, r, 3, len(ct_s),  gap_bg)
            dat(ws, r, 4, q_px,       gap_bg, fmt="0.00")
            dat(ws, r, 5, q_ct,       gap_bg, fmt="0.00")
            dat(ws, r, 6, q_gap,      gap_bg, fmt="0.00",
                bold=(pd.notna(q_gap) and q_gap > 1.0))
            dat(ws, r, 7, pr_px,      gap_bg, fmt="0.000")
            dat(ws, r, 8, pr_ct,      gap_bg, fmt="0.000")
            ws.row_dimensions[r].height = 16
            r += 1

    set_col_widths(ws, [(1, 18), (2, 10), (3, 10), (4, 12), (5, 12),
                        (6, 10), (7, 12), (8, 12), (9, 12), (10, 12), (11, 42)])
    ws.freeze_panes = "A5"


# ═══════════════════════════════════════════════════════════════════════════════
# MASTER LOADER — collect all task×paradigm results
# ═══════════════════════════════════════════════════════════════════════════════

def load_all_results(hmm_dir: Path, df_px: pd.DataFrame,
                     tasks: list, paradigms: list) -> list:
    """
    Load results + checkpoint for every task×paradigm.
    Returns list of dicts with pred_map and feature_importance attached.
    """
    all_results = []
    px_ids  = df_px[df_px["id"].str.startswith("PX")]["id"].tolist()
    fx_ids  = df_px[~df_px["id"].str.startswith("PX")]["id"].tolist()

    for t in tasks:
        for p in paradigms:
            results = load_results_json(hmm_dir, t, p)
            if results is None:
                print(f"  [SKIP] T{t}_P{p} — no results JSON")
                continue
            ckpt = load_checkpoint(hmm_dir, t, p)
            print(f"  [OK]   T{t}_P{p}  BA={results['metrics'].get('ba', '?')}")

            # Determine g1/g0 for this paradigm
            if p == 1:
                g1_ids, g0_ids = px_ids, fx_ids
            elif p == 2:
                rct = df_px[(df_px["id"].str.startswith("PX")) &
                            (df_px["dia_code"] == 1)]["id"].tolist()
                g1_ids, g0_ids = rct, fx_ids
            elif p == 3:
                oth = df_px[(df_px["id"].str.startswith("PX")) &
                            (df_px["dia_code"] != 1)]["id"].tolist()
                g1_ids, g0_ids = oth, fx_ids
            elif p == 4:
                rct = df_px[(df_px["id"].str.startswith("PX")) &
                            (df_px["dia_code"] == 1)]["id"].tolist()
                oth = df_px[(df_px["id"].str.startswith("PX")) &
                            (df_px["dia_code"] != 1)]["id"].tolist()
                g1_ids, g0_ids = rct, oth

            # Build prediction map
            preds_src = (ckpt or {}).get("predictions") or results.get("predictions", {})
            y_true  = np.array(preds_src.get("y_true",  []))
            y_pred  = np.array(preds_src.get("y_pred",  []))
            y_proba = np.array(preds_src.get("y_proba", []))

            ordered = get_ckpt_subject_order(g1_ids, g0_ids)
            pred_map = {}
            for i, sid in enumerate(ordered):
                if i < len(y_true):
                    pred_map[sid.upper()] = {
                        "y_true":  int(y_true[i]),
                        "y_pred":  int(y_pred[i]),
                        "y_proba": float(y_proba[i]) if i < len(y_proba) else np.nan,
                    }

            all_results.append({
                "task":               t,
                "paradigm":           p,
                "metrics":            results.get("metrics", {}),
                "best_params":        results.get("best_params", {}),
                "feature_importance": results.get("feature_importance", {}),
                "pred_map":           pred_map,
            })

    return all_results


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(description="HMM Laterality Analysis — 3 post-hoc analyses")
    ap.add_argument("--dataset",     default="xdash",
                    help="Dataset name (must match datasets/ folder). Default: xdash")
    ap.add_argument("--hmm-dir",     default=None,
                    help="Root of HMM experiment outputs "
                         "(default: storage/results/<dataset>/experiments)")
    ap.add_argument("--px-details",  default="data/xdash_px_details.xlsx",
                    help="Path to xdash_px_details.xlsx")
    ap.add_argument("--out",         default=None,
                    help="Output directory (default: storage/results/<dataset>/hmm/laterality)")
    ap.add_argument("--task",        nargs="+", type=int, default=TASKS,
                    choices=TASKS, help="Tasks to include (default: all 1-6)")
    ap.add_argument("--paradigm",    nargs="+", type=int, default=PARADIGMS,
                    choices=PARADIGMS, help="Paradigms to include (default: all 1-4)")
    args = ap.parse_args()

    for _root in (Path.cwd(), *Path(__file__).resolve().parents):
        if (_root / "config" / "paths.py").exists():
            sys.path.insert(0, str(_root))
            break
    from config.paths import get_experiments_dir, get_results_dir
    from dataio.ingestion import load_dataset_config

    hmm_dir = Path(args.hmm_dir) if args.hmm_dir else get_experiments_dir(args.dataset)
    out_dir = Path(args.out) if args.out else get_results_dir(args.dataset) / "hmm" / "laterality"
    out_dir.mkdir(parents=True, exist_ok=True)

    channel_names = load_dataset_config(args.dataset)["channels"]

    print(f"\n{'='*60}")
    print(f"  HMM Laterality Analysis")
    print(f"  Tasks:    {args.task}")
    print(f"  Paradigms:{args.paradigm}")
    print(f"  HMM dir:  {hmm_dir}")
    print(f"{'='*60}\n")

    # Load patient metadata
    df_px = load_px_details(Path(args.px_details))
    if df_px is None:
        print("ERROR: Cannot load px_details. Aborting.")
        return

    px_meta = build_patient_meta(df_px)
    print(f"Cohort: {len(px_meta)} patients")
    print(px_meta["laterality"].value_counts().to_string())
    print()

    # Load all HMM results
    print("Loading HMM results...")
    all_results = load_all_results(hmm_dir, df_px, args.task, args.paradigm)
    print(f"\nLoaded {len(all_results)} / {len(args.task)*len(args.paradigm)} experiments\n")

    if not all_results:
        print("No results loaded. Check --hmm-dir path.")
        return

    # ── Run analyses ──────────────────────────────────────────────────────────
    print("Running Analysis 1: Laterality × Feature Importance...")
    detail_df = compute_laterality_importance(all_results, px_meta, channel_names)
    # summary_df is now computed per-paradigm inside the workbook builder below

    print("Running Analysis 2: Dominant-Hand x Injured-Side Cross-Tabulation...")
    subj_df_2, xtab_df = compute_used_hand_crosstab(all_results, px_meta)

    print("Running Analysis 3: HMM Probability vs Self-Reported Difficulty...")
    corr_df, subj_df_3, misclass_df = compute_proba_difficulty_correlation(
        all_results, px_meta, df_px)

    # ── Build workbook ────────────────────────────────────────────────────────
    print("\nBuilding Excel workbook...")
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Overview"
    sheet_overview(ws1, px_meta)

    # Analysis 1 — 4 sheets (one per paradigm)
    for p in PARADIGMS:
        p_label = PARADIGM_LABELS[p]
        ws = wb.create_sheet(f"A1-P{p} Lat×Importance")
        summary_p = compute_task_laterality_summary(detail_df, paradigm=p)
        sheet_analysis1(ws, summary_p, detail_df, paradigm=p)

    # Analysis 2 — 4 sheets (one per paradigm)
    for p in PARADIGMS:
        ws = wb.create_sheet(f"A2-P{p} DomHand×Injured")
        sheet_analysis2(ws, subj_df_2, xtab_df, paradigm=p)

    # Analysis 3 — correlation + misclassification (single sheets, all paradigms)
    ws_corr = wb.create_sheet("A3a - Proba vs Difficulty")
    ws_miss = wb.create_sheet("A3b - Misclassification")
    sheet_analysis3_corr(ws_corr, corr_df)
    sheet_analysis3_misclass(ws_miss, misclass_df, subj_df_3)

    out_path = out_dir / "HMM_Laterality_Analysis.xlsx"
    wb.save(out_path)
    print(f"\n{'='*60}")
    print(f"  Output: {out_path}")
    print(f"  Sheets:")
    for ws in wb.worksheets:
        print(f"    • {ws.title}")
    print(f"{'='*60}\n")

    # ── Print quick summary to console ───────────────────────────────────────
    if not detail_df.empty:
        print("Analysis 1 — Task laterality summary (paradigm 1):")
        summary_p1 = compute_task_laterality_summary(detail_df, paradigm=1)
        print(summary_p1[["task_label", "n_unilateral_px", "mean_ipsi_imp",
                           "mean_head_imp", "pct_ipsi_dom", "pct_head_dom",
                           "pct_correct", "interpretation"]].to_string(index=False))
        print()

    if not corr_df.empty:
        p1_corr = corr_df[corr_df["paradigm"] == 1][
            ["task_label", "n_patients", "rho_patients", "pv_patients",
             "q_patients", "sig_patients"]]
        print("Analysis 3 — Spearman ρ (patients only, paradigm 1):")
        print(p1_corr.to_string(index=False))
        print()

    if not misclass_df.empty:
        print(f"Analysis 3b — {len(misclass_df)} false negative patient observations found")
        print(misclass_df[["task_label", "subject_id", "diagnosis",
                            "q_score", "note"]].to_string(index=False))


if __name__ == "__main__":
    main()