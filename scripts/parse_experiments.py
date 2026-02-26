import json
import glob
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils.dataframe import dataframe_to_rows

# ── Config ────────────────────────────────────────────────────────────────────
EXPERIMENTS_DIR = "experiments_from_hpc"
OUTPUT_FILE = "nn-models-results/xdash_results.xlsx"

FEATURE_COLS = [
    "head_pos_x","head_pos_y","head_pos_z",
    "head_rot_x","head_rot_y","head_rot_z",
    "right_hand_pos_x","right_hand_pos_y","right_hand_pos_z",
    "right_hand_rot_x","right_hand_rot_y","right_hand_rot_z",
    "left_hand_pos_x","left_hand_pos_y","left_hand_pos_z",
    "left_hand_rot_x","left_hand_rot_y","left_hand_rot_z",
]

TASK_NAMES = {
    1:"jar_opening", 2:"key_turning", 3:"cleaning",
    4:"back_washing", 5:"cutting", 6:"hammering"
}
PARADIGM_NAMES = {
    1:"patients_vs_controls", 2:"rct_vs_controls",
    3:"other_conditions_vs_controls", 4:"rct_vs_other_conditions"
}

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
SUBHDR_FILL   = PatternFill("solid", fgColor="2E75B6")   # mid blue
SECTION_FILLS = {
    "CNN": PatternFill("solid", fgColor="D6E4F7"),
    "RNN": PatternFill("solid", fgColor="E2EFDA"),
    "HMM": PatternFill("solid", fgColor="FFF2CC"),
}
HIGH_RISK_FILL = PatternFill("solid", fgColor="FFB3B3")
MED_RISK_FILL  = PatternFill("solid", fgColor="FFE5A0")
LOW_RISK_FILL  = PatternFill("solid", fgColor="C6EFCE")

THIN = Side(style="thin", color="BFBFBF")
THICK = Side(style="medium", color="1F3864")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def hdr_style(cell, size=11):
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=size)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER

def subhdr_style(cell):
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill = SUBHDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

def data_style(cell, bold=False, align="center", fill=None):
    cell.font = Font(bold=bold, name="Arial", size=9)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill

def pct_fmt(cell):
    cell.number_format = "0.0%"

# ── Parse ─────────────────────────────────────────────────────────────────────
def load_all_summaries(base_dir):
    rows = []
    pattern = os.path.join(base_dir, "task*", "paradigm*", "*_*", "summary.json")
    files = glob.glob(pattern, recursive=True)
    if not files:
        # Try flat layout
        pattern = os.path.join(base_dir, "**", "summary.json")
        files = glob.glob(pattern, recursive=True)

    print(f"Found {len(files)} summary.json files")

    for fpath in sorted(files):
        try:
            with open(fpath) as f:
                d = json.load(f)
        except Exception as e:
            print(f"  SKIP {fpath}: {e}")
            continue

        cfg = d.get("config", {})
        res = d.get("results", {})
        evl = d.get("evaluation", {})
        diag = d.get("diagnostics", {})
        metrics = res.get("metrics", {})
        fi = res.get("feature_importance", {})
        bp = res.get("best_params", {})

        task = cfg.get("task", res.get("task"))
        paradigm = cfg.get("paradigm", res.get("paradigm"))
        model = cfg.get("model", res.get("model", "")).upper()

        row = {
            "task":          task,
            "task_name":     TASK_NAMES.get(task, f"task{task}"),
            "paradigm":      paradigm,
            "paradigm_name": PARADIGM_NAMES.get(paradigm, f"paradigm{paradigm}"),
            "model":         model,
            "method":        cfg.get("method", res.get("preprocessing_method", "")),
            # metrics
            "ba":            metrics.get("ba"),
            "recall":        metrics.get("recall"),
            "precision":     metrics.get("precision"),
            "f1":            metrics.get("f1"),
            "auc":           metrics.get("auc"),
            "auc_ci_low":    metrics.get("auc_ci_low"),
            "auc_ci_high":   metrics.get("auc_ci_high"),
            # diagnostics
            "overfitting_risk":    diag.get("overfitting_risk"),
            "generalization_gap":  diag.get("generalization_gap"),
            # best params (store as string for display)
            "best_params_str": json.dumps(bp, separators=(",", ":")),
        }

        # Feature importances — ranked position
        sorted_feats = sorted(fi.items(), key=lambda x: x[1], reverse=True)
        for rank, (feat, score) in enumerate(sorted_feats, 1):
            row[f"fi_{feat}"] = score
            row[f"rank_{feat}"] = rank

        rows.append(row)

    return pd.DataFrame(rows)


# ── Sheet builders ────────────────────────────────────────────────────────────

def build_summary_sheet(wb, df):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:R1")
    ws["A1"] = "XDash Experiment Results — Summary"
    ws["A1"].font = Font(bold=True, name="Arial", size=14, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        "Task", "Task Name", "Paradigm", "Paradigm Name", "Model",
        "BA", "Recall", "Precision", "F1", "AUC",
        "AUC CI Low", "AUC CI High",
        "Overfitting Risk", "Gen. Gap",
        "Top-1 Feature", "Top-2 Feature", "Top-3 Feature", "Top-4 Feature",
        "Top-5 Feature", "Top-6 Feature", "Top-7 Feature", "Top-8 Feature",
        "Top-9 Feature", "Top-10 Feature", "Top-11 Feature", "Top-12 Feature", 
        "Top-13 Feature", "Top-14 Feature", "Top-15 Feature", "Top-16 Feature", 
        "Top-17 Feature", "Top-18 Feature"
    ]
    ws.append([])  # blank row 2
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        hdr_style(cell)
    ws.row_dimensions[3].height = 30

    metric_cols = {"BA":6,"Recall":7,"Precision":8,"F1":9,"AUC":10}
    feat_cols_idx = {"Top-1 Feature":15,"Top-2 Feature":16,"Top-3 Feature":17}

    df_sorted = df.sort_values(["task","paradigm","model"]).reset_index(drop=True)

    for r_idx, row in df_sorted.iterrows():
        excel_row = r_idx + 4
        model = row.get("model","")
        fill = SECTION_FILLS.get(model)

        vals = [
            row.get("task"), row.get("task_name"),
            row.get("paradigm"), row.get("paradigm_name"),
            model,
            row.get("ba"), row.get("recall"), row.get("precision"),
            row.get("f1"), row.get("auc"),
            row.get("auc_ci_low"), row.get("auc_ci_high"),
            row.get("overfitting_risk"), row.get("generalization_gap"),
        ]

        # All features by rank
        rank_cols = {feat: row.get(f"rank_{feat}") for feat in FEATURE_COLS if pd.notna(row.get(f"rank_{feat}"))}
        all = sorted(rank_cols, key=rank_cols.get)
        vals += all + [""] * (3 - len(all))

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            data_style(cell, fill=fill)
            if col in range(6, 13) and isinstance(val, float):
                cell.number_format = "0.000"
            if col == 13:  # risk
                risk = str(val or "")
                if risk == "HIGH":   cell.fill = HIGH_RISK_FILL
                elif risk == "MEDIUM": cell.fill = MED_RISK_FILL
                elif risk == "LOW":  cell.fill = LOW_RISK_FILL
                cell.font = Font(bold=True, name="Arial", size=9)
            if col == 14 and isinstance(val, float):
                cell.number_format = "0.0000"

        ws.row_dimensions[excel_row].height = 16

    # Column widths
    widths = [6,16,8,28,6, 7,7,7,7,7, 9,9, 14,10, 22,22,22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Conditional formatting on BA column
    last_row = len(df_sorted) + 3
    ws.conditional_formatting.add(
        f"F4:F{last_row}",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="num", mid_value=0.7, mid_color="FFEB84",
            end_type="max", end_color="63BE7B"
        )
    )
    ws.freeze_panes = "A4"


def build_metrics_sheet(wb, df):
    ws = wb.create_sheet("Metrics by Task-Paradigm")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L1")
    ws["A1"] = "Performance Metrics — All Models × Tasks × Paradigms"
    ws["A1"].font = Font(bold=True, name="Arial", size=13, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["Task","Paradigm","Model","BA","Recall","Precision","F1",
               "AUC","AUC CI Low","AUC CI High","Overfitting Risk","Gen. Gap"]
    ws.append([])
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        hdr_style(cell)
    ws.row_dimensions[3].height = 30

    df_sorted = df.sort_values(["task","paradigm","model"]).reset_index(drop=True)
    prev_tp = None

    for r_idx, row in df_sorted.iterrows():
        excel_row = r_idx + 4
        tp = (row.get("task"), row.get("paradigm"))
        model = row.get("model","")
        fill = SECTION_FILLS.get(model)

        vals = [
            row.get("task"), row.get("paradigm"), model,
            row.get("ba"), row.get("recall"), row.get("precision"), row.get("f1"),
            row.get("auc"), row.get("auc_ci_low"), row.get("auc_ci_high"),
            row.get("overfitting_risk"), row.get("generalization_gap"),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            data_style(cell, fill=fill)
            if col in range(4, 11) and isinstance(val, float):
                cell.number_format = "0.000"
            if col == 11:
                risk = str(val or "")
                if risk == "HIGH":   cell.fill = HIGH_RISK_FILL
                elif risk == "MEDIUM": cell.fill = MED_RISK_FILL
                elif risk == "LOW":  cell.fill = LOW_RISK_FILL
                cell.font = Font(bold=True, name="Arial", size=9)
            if col == 12 and isinstance(val, float):
                cell.number_format = "0.0000"

        # Separator line between task-paradigm groups
        if prev_tp and prev_tp != tp:
            for col in range(1, 13):
                ws.cell(row=excel_row, column=col).border = Border(
                    top=Side(style="medium", color="1F3864"),
                    left=THIN, right=THIN, bottom=THIN
                )
        prev_tp = tp
        ws.row_dimensions[excel_row].height = 16

    widths = [6, 9, 7, 7, 7, 9, 7, 7, 9, 9, 14, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last_row = len(df_sorted) + 3
    ws.conditional_formatting.add(f"D4:D{last_row}",
        ColorScaleRule(start_type="min", start_color="F8696B",
                       mid_type="num", mid_value=0.7, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B"))
    ws.freeze_panes = "A4"


def build_feature_importance_sheet(wb, df):
    ws = wb.create_sheet("Feature Importance")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:V1")
    ws["A1"] = "Feature Importance Scores — All Experiments"
    ws["A1"].font = Font(bold=True, name="Arial", size=13, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    info_cols = ["Task", "Task Name", "Paradigm", "Model"]
    headers = info_cols + FEATURE_COLS
    ws.append([])
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        hdr_style(cell, size=9)
    ws.row_dimensions[3].height = 40

    df_sorted = df.sort_values(["task","paradigm","model"]).reset_index(drop=True)

    for r_idx, row in df_sorted.iterrows():
        excel_row = r_idx + 4
        model = row.get("model","")
        fill = SECTION_FILLS.get(model)

        info_vals = [row.get("task"), row.get("task_name"), row.get("paradigm"), model]
        for col, val in enumerate(info_vals, 1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            data_style(cell, fill=fill, bold=(col==4))

        for col_offset, feat in enumerate(FEATURE_COLS):
            col = col_offset + 5
            val = row.get(f"fi_{feat}")
            cell = ws.cell(row=excel_row, column=col, value=val)
            data_style(cell)
            if isinstance(val, float):
                cell.number_format = "0.0000"

        ws.row_dimensions[excel_row].height = 16

    # Widths
    info_widths = [6, 16, 9, 7]
    for i, w in enumerate(info_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(5, 5 + len(FEATURE_COLS)):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # Color scale on feature scores
    last_row = len(df_sorted) + 3
    fi_range = f"E4:{get_column_letter(4+len(FEATURE_COLS))}{last_row}"
    ws.conditional_formatting.add(fi_range,
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max", end_color="2E75B6"))
    ws.freeze_panes = "E4"


def build_feature_ranks_sheet(wb, df):
    ws = wb.create_sheet("Feature Rankings")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:V1")
    ws["A1"] = "Feature Rankings (1 = most important) — All Experiments"
    ws["A1"].font = Font(bold=True, name="Arial", size=13, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    info_cols = ["Task", "Task Name", "Paradigm", "Model"]
    headers = info_cols + FEATURE_COLS
    ws.append([])
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        hdr_style(cell, size=9)
    ws.row_dimensions[3].height = 40

    df_sorted = df.sort_values(["task","paradigm","model"]).reset_index(drop=True)

    for r_idx, row in df_sorted.iterrows():
        excel_row = r_idx + 4
        model = row.get("model","")
        fill = SECTION_FILLS.get(model)

        info_vals = [row.get("task"), row.get("task_name"), row.get("paradigm"), model]
        for col, val in enumerate(info_vals, 1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            data_style(cell, fill=fill, bold=(col==4))

        for col_offset, feat in enumerate(FEATURE_COLS):
            col = col_offset + 5
            val = row.get(f"rank_{feat}")
            cell = ws.cell(row=excel_row, column=col, value=int(val) if pd.notna(val) else "")
            data_style(cell)

        ws.row_dimensions[excel_row].height = 16

    info_widths = [6, 16, 9, 7]
    for i, w in enumerate(info_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(5, 5 + len(FEATURE_COLS)):
        ws.column_dimensions[get_column_letter(i)].width = 16

    # Color scale: rank 1=dark green, 18=light (reversed)
    last_row = len(df_sorted) + 3
    fi_range = f"E4:{get_column_letter(4+len(FEATURE_COLS))}{last_row}"
    ws.conditional_formatting.add(fi_range,
        ColorScaleRule(start_type="num", start_value=1, start_color="63BE7B",
                       end_type="num", end_value=18, end_color="FFFFFF"))
    ws.freeze_panes = "E4"


def build_model_pivot_sheet(wb, df):
    """Best BA per task × paradigm for each model, side by side."""
    ws = wb.create_sheet("Model Comparison Pivot")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:M1")
    ws["A1"] = "Best Balanced Accuracy — Task × Paradigm × Model"
    ws["A1"].font = Font(bold=True, name="Arial", size=13, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    models = sorted(df["model"].dropna().unique())
    # Headers: Task | Paradigm | CNN | RNN | HMM | Best Model
    col_headers = ["Task", "Paradigm"] + models + ["Best Model", "Best BA"]
    ws.append([])
    for col, h in enumerate(col_headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        hdr_style(cell)
    ws.row_dimensions[3].height = 30

    tasks = sorted(df["task"].dropna().unique())
    paradigms = sorted(df["paradigm"].dropna().unique())

    excel_row = 4
    for task in tasks:
        for paradigm in paradigms:
            sub = df[(df["task"]==task) & (df["paradigm"]==paradigm)]
            ba_by_model = {m: sub[sub["model"]==m]["ba"].max() for m in models}
            valid = {m: v for m,v in ba_by_model.items() if pd.notna(v)}
            best_model = max(valid, key=valid.get) if valid else ""
            best_ba    = valid.get(best_model)

            row_vals = [task, paradigm] + [ba_by_model.get(m) for m in models] + [best_model, best_ba]
            for col, val in enumerate(row_vals, 1):
                cell = ws.cell(row=excel_row, column=col, value=val)
                data_style(cell)
                if isinstance(val, float):
                    cell.number_format = "0.000"
                if col == len(col_headers) - 1 and val:  # best model col
                    cell.fill = SECTION_FILLS.get(str(val), PatternFill())
                    cell.font = Font(bold=True, name="Arial", size=9)

            ws.row_dimensions[excel_row].height = 16
            excel_row += 1

    for i, w in enumerate([6, 9] + [9]*len(models) + [12, 9], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last_row = excel_row - 1
    for col_offset, _ in enumerate(models):
        col_letter = get_column_letter(3 + col_offset)
        ws.conditional_formatting.add(f"{col_letter}4:{col_letter}{last_row}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="num", mid_value=0.7, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B"))
    ws.freeze_panes = "A4"


def build_diagnostics_sheet(wb, df):
    ws = wb.create_sheet("Diagnostics")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"] = "Overfitting Diagnostics — All Experiments"
    ws["A1"].font = Font(bold=True, name="Arial", size=13, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["Task", "Task Name", "Paradigm", "Paradigm Name", "Model",
               "Overfitting Risk", "Generalization Gap"]
    ws.append([])
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        hdr_style(cell)
    ws.row_dimensions[3].height = 30

    df_sorted = df.sort_values(["task","paradigm","model"]).reset_index(drop=True)
    for r_idx, row in df_sorted.iterrows():
        excel_row = r_idx + 4
        model = row.get("model","")
        vals = [
            row.get("task"), row.get("task_name"),
            row.get("paradigm"), row.get("paradigm_name"),
            model, row.get("overfitting_risk"), row.get("generalization_gap"),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            fill = SECTION_FILLS.get(model)
            data_style(cell, fill=fill)
            if col == 6:
                risk = str(val or "")
                if risk == "HIGH":    cell.fill = HIGH_RISK_FILL
                elif risk == "MEDIUM": cell.fill = MED_RISK_FILL
                elif risk == "LOW":   cell.fill = LOW_RISK_FILL
                cell.font = Font(bold=True, name="Arial", size=9)
            if col == 7 and isinstance(val, float):
                cell.number_format = "0.0000"
        ws.row_dimensions[excel_row].height = 16

    for i, w in enumerate([6,16,9,30,7,14,14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last_row = len(df_sorted) + 3
    ws.conditional_formatting.add(f"G4:G{last_row}",
        ColorScaleRule(start_type="min", start_color="63BE7B",
                       end_type="max", end_color="F8696B"))
    ws.freeze_panes = "A4"


def build_legend_sheet(wb):
    ws = wb.create_sheet("Legend")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50

    ws["A1"] = "Legend & Notes"
    ws["A1"].font = Font(bold=True, name="Arial", size=14, color="1F3864")
    ws.row_dimensions[1].height = 26

    entries = [
        ("SECTION",""),
        ("Color Coding — Models",""),
        ("CNN rows","Light blue"),
        ("RNN rows","Light green"),
        ("HMM rows","Light yellow"),
        ("",""),
        ("Color Coding — Risk",""),
        ("HIGH overfitting","Red background"),
        ("MEDIUM overfitting","Amber background"),
        ("LOW overfitting","Green background"),
        ("",""),
        ("Metrics",""),
        ("BA","Balanced Accuracy (primary metric)"),
        ("AUC","Area Under ROC Curve"),
        ("AUC CI","95% Bootstrap Confidence Interval"),
        ("Gen. Gap","Generalization Gap (|train − val|)"),
        ("",""),
        ("Feature Rankings",""),
        ("Rank 1","Most discriminative feature (darkest green)"),
        ("Rank 18","Least discriminative feature"),
        ("",""),
        ("Sheets",""),
        ("Summary","One row per experiment, top-3 features"),
        ("Metrics by Task-Paradigm","All metrics with risk colour coding"),
        ("Feature Importance","Raw importance scores (higher = more important)"),
        ("Feature Rankings","Rank 1–18 per experiment"),
        ("Model Comparison Pivot","Best BA per task × paradigm, side-by-side"),
        ("Diagnostics","Overfitting risk + generalization gap"),
    ]

    fills = {
        "CNN rows": SECTION_FILLS["CNN"],
        "RNN rows": SECTION_FILLS["RNN"],
        "HMM rows": SECTION_FILLS["HMM"],
        "HIGH overfitting": HIGH_RISK_FILL,
        "MEDIUM overfitting": MED_RISK_FILL,
        "LOW overfitting": LOW_RISK_FILL,
    }

    for r, (k, v) in enumerate(entries, 3):
        ca = ws.cell(row=r, column=1, value=k)
        cb = ws.cell(row=r, column=2, value=v)
        ca.font = Font(name="Arial", size=10, bold=(v==""))
        cb.font = Font(name="Arial", size=10)
        if k in fills:
            ca.fill = fills[k]
            cb.fill = fills[k]
        if v == "":
            ca.font = Font(bold=True, name="Arial", size=10, color="1F3864")
        ws.row_dimensions[r].height = 16


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    df = load_all_summaries(EXPERIMENTS_DIR)
    if df.empty:
        print("No data loaded. Check EXPERIMENTS_DIR path.")
        return

    print(f"Loaded {len(df)} experiments")
    print(f"Models: {sorted(df['model'].dropna().unique())}")
    print(f"Tasks:  {sorted(df['task'].dropna().unique())}")

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_summary_sheet(wb, df)
    build_metrics_sheet(wb, df)
    build_feature_importance_sheet(wb, df)
    build_feature_ranks_sheet(wb, df)
    build_model_pivot_sheet(wb, df)
    build_diagnostics_sheet(wb, df)
    build_legend_sheet(wb)

    wb.save(OUTPUT_FILE)
    print(f"\nSaved → {OUTPUT_FILE}")


if __name__ == "__main__":
    main()