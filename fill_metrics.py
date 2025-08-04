import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ─────────────────────────────────────────────────────────────────────────────
# 1) LOAD YOUR “LONG” RESULTS (from CSV or Python list of dicts)
# ─────────────────────────────────────────────────────────────────────────────

# TODO: change here
df_res = pd.read_csv("rcvni.csv")

# ─────────────────────────────────────────────────────────────────────────────
# 2) MAP NUMERIC OR INTERNAL CODES TO EXACT EXCEL LABELS
# ─────────────────────────────────────────────────────────────────────────────

# paradigm_map = {
#     2: "1. Rotatary Cuff vs No Injury",
#     1: "2. Injury vs No Injury (# 37 vs 17)",
#     3: "4. Other Injury vs No Injury",
#     4: "3. Rotatary vs Other Injury"
# }


# attempt_map = {
#     2: "One attempts in LL",
#     1: "One attempts in LU",
#     3: "One attempts in LL+LU",
#     4: "Multiple attempts in LL+LU"
# }

model_map = {
    1: "LR (LOO)",
    2: "RF (LOO)",
    3: "KNN (LOO)",
    4: "HMM (LOO)"
}

# TODO: change here
feature_filter_map = {
    1: "Time series (with padding) [length]",
    2: "Time series (with truncating) [length]",
    3: "Time series (with DTW) [length]",
    4: "Time series (with variable) [length]"
}

df_res["AttemptLabel"] = df_res["OneAttempt"].map(attempt_map)
df_res["ParadigmLabel"] = df_res["Paradigm"].map(paradigm_map)
df_res["ModelLabel"] = df_res["Model"].map(model_map)
df_res["FeatureFilterLabel"] = df_res["FeatureFilter"].map(feature_filter_map)
print(df_res)

# ─────────────────────────────────────────────────────────────────────────────
# 3) READ YOUR EXISTING EXCEL “TEMPLATE”
# ─────────────────────────────────────────────────────────────────────────────

# Adjust the sheet_name= if your sheet is named differently.
df_tpl = pd.read_excel("features-result-table.xlsx")

# For clarity, let’s display the first few columns so you can see column names:
print("=== TEMPLATED SHEET COLUMNS ===")
print(df_tpl.columns.tolist())
print("===============================")


# ─────────────────────────────────────────────────────────────────────────────
# 4) LOOP OVER EACH ROW OF THE TEMPLATE AND “FILL IN” THE AUC/BA/RECALL CELLS
# ─────────────────────────────────────────────────────────────────────────────

print("\n=== RESULT KEYS ===")
print(df_res.columns.tolist())
print("===================\n")

filled_cells = []

# Now iterate row by row through the template:
for idx, tpl_row in df_tpl.iterrows():
    # 1) Read “Filter – Number of attempts” from the template
    attempt_label = tpl_row["Filter - Number of attempts"]
    model_label = tpl_row["Model used"]
    feature_filter_label = tpl_row["Features"]

    # 2) For *each* of the four paradigm‐columns, check if we have a matching row in df_res
    for paradigm_col in paradigm_map.values():        
        match = df_res[
            (df_res["ParadigmLabel"] == paradigm_col) & 
            (df_res["FeatureFilterLabel"] == feature_filter_label) &
            (df_res["ModelLabel"] == model_label) &
            (df_res["AttemptLabel"] == attempt_label)
        ]
        if match.shape[0] == 0:
            # No matching result for this (paradigm,attempt). Skip.
            continue

        print("@@@@@@@@@@@@@@", idx, tpl_row["Filter - Number of attempts"], tpl_row["Model used"], tpl_row["Features"])
        print("********heyyyyyyyyyyyy", paradigm_col, attempt_label, model_label, feature_filter_label)


        
        # If there are multiple matches (e.g. different FeatureFilter/Model combos),
        # you’ll want further logic here (e.g. pick the one where Model==“LR-80/20” or something).
        # For now we just grab the first:
        row = match.iloc[0]
        auc_val    = row["auc"]
        ba_val     = row["ba"]
        recall_val = row["recall"]


        cell_value = df_tpl.at[idx, paradigm_col]
        print("before cell val", cell_value)
        if isinstance(cell_value, str) and cell_value.startswith("Recall"):
            df_tpl.at[idx, paradigm_col] = "Recall="+str(round(recall_val, 3))
        elif isinstance(cell_value, str) and cell_value.startswith("BA"):
            df_tpl.at[idx, paradigm_col] = "BA="+str(round(ba_val, 3))
        elif isinstance(cell_value, str) and cell_value.startswith("AUC"):
            number_strings = re.findall(r"\d+\.\d+", auc_val)
            numbers = [float(x) for x in number_strings]
            df_tpl.at[idx, paradigm_col] = "AUC="+str(round(numbers[0], 3))+"["+str(round(numbers[1], 3))+"-"+str(round(numbers[2], 3))+"]"
        print("after cell val", df_tpl.at[idx, paradigm_col])

        filled_cells.append((idx, paradigm_col))

# # ─────────────────────────────────────────────────────────────────────────────
# # 5) SAVE THE UPDATED EXCEL
# # ─────────────────────────────────────────────────────────────────────────────

output_filename = "features-result-table.xlsx"
df_tpl.to_excel(output_filename, index=False)
print(f"\n✅ Wrote updated sheet to '{output_filename}'")

# ─────────────────────────────────────────────────────────────────────────────
# 5) REOPEN WITH openpyxl AND APPLY FILL TO RECORDED CELLS
# ─────────────────────────────────────────────────────────────────────────────

wb = load_workbook(output_filename)
ws = wb.active

# Create a yellow fill style
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Map column names to Excel column letters
# (e.g. "1. Rotator Cuff vs No Injury" → "D" if that is column D in the sheet)
col_letter_map = {}
for col_idx, col_name in enumerate(df_tpl.columns, start=1):
    col_letter_map[col_name] = ws.cell(row=1, column=col_idx).column_letter

# Now iterate over filled_cells to color each one
for row_idx, col_name in filled_cells:
    # Excel row = DataFrame index + 2 (because header is row 1, and index=0→Excel row 2)
    excel_row    = row_idx + 2
    excel_column = col_letter_map[col_name]
    ws[f"{excel_column}{excel_row}"].fill = yellow_fill

# Save the workbook
wb.save(output_filename)
print(f"✅ Filled cells highlighted in yellow and saved to '{output_filename}'")